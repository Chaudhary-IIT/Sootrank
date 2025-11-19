from collections import defaultdict
from io import BytesIO, TextIOWrapper
import json
import io
import os
import razorpay
from django.forms import modelformset_factory
from django.utils.safestring import mark_safe
from django.db.models import Sum
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.template.loader import render_to_string
from decimal import Decimal, ROUND_HALF_UP
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.db import IntegrityError, transaction
from django.db.models import Prefetch, Sum,IntegerField, BigIntegerField,Count, Avg, F, FloatField, ExpressionWrapper,Case, When, Value
from django.utils.text import slugify
from django.http import Http404, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden,JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import Workbook
from Registration.models import Branch, Category, Course, CourseBranch, Department, ProgramRequirement, Student, Faculty, Admins, StudentCourse, AssessmentComponent,AssessmentScore, Attendance, Timetable,FeeRecord,GradingPolicy
from django.contrib.auth import authenticate , login as auth_login
from django.contrib.auth.hashers import make_password,check_password
from django.db.models import Count, Q, OuterRef, Exists ,Max
from django.views.decorators.http import require_POST,require_http_methods
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.conf import settings
from urllib.parse import urlencode
from django.core.files.storage import default_storage
import re
from Registration.forms import FacultyEditForm, StudentEditForm
import csv
from django.http import JsonResponse
import pandas as pd
from dataclasses import dataclass
from datetime import datetime
from django.db.models import Count, Q
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from io import TextIOWrapper
import csv
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from openpyxl import load_workbook

EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+-]+@iitmandi\.ac\.in$')
SLOTS = ["A","B","C","D","E","F","G","H","FS"]
CATEGORY_MAP = {
    "DC": "DC",
    "DE": "DE",
    "IC": "IC",
    "HSS": "HSS",
    "FE": "FE",
    "IKS": "IKS",
    "ISTP": "ISTP",
    "MTP": "MTP",
}

CATEGORIES = [
    {"code": "DC", "label": "Disciplinary Core (DC)"},
    {"code": "DE", "label": "Disciplinary Elective (DE)"},
    {"code": "IC", "label": "Institute Core (IC)"},
    {"code": "HSS", "label": "Humanities and Social Science (HSS)"},
    {"code": "FE", "label": "Free Elective (FE)"},
    {"code": "IKS", "label": "Indian Knowledge System (IKS)"},
    {"code": "ISTP", "label": "Interactive Socio-Technical Practicum (ISTP)"},
    {"code": "MTP", "label": "Major Technical Project (MTP)"},
]
CATEGORY_HEADERS = ["DC", "DE", "IC", "HSS", "FE", "IKS", "ISTP", "MTP"]
HEADER_TO_CATEGORY = {h: h for h in CATEGORY_HEADERS}

GRADE_BOUNDARIES = [
    ("A", 85),
    ("B", 75),
    ("C", 65),
    ("D", 55),
    ("E", 45),
    ("F", 0),
]
PASS_MIN_PERCENT = 45  # for letter-graded; PF handled by is_pass_fail

GRADE_POINTS = {
    "A": 10,
    "A-": 9,
    "B": 8,
    "B-": 7,
    "C": 6,
    "C-": 5,
    "D": 4,
    "F": 0,
    "FS": 0,
}

CATEGORY_PRIORITY = {
    "IC": 1,
    "DC": 2,
    "DE": 3,
    "HSS": 4,
    "IKS": 5,
    "FE": 6,
}

# Slot mapping for timetable
SLOT_SCHEDULE = {
    "A": [("Monday", "8–8:50 AM"), ("Tuesday", "11–11:50 AM"), ("Thursday", "9–9:50 AM")],
    "B": [("Monday", "9–9:50 AM"), ("Tuesday", "12–12:50 PM"), ("Thursday", "10–10:50 AM")],
    "C": [("Monday", "10–10:50 AM"), ("Thursday", "11–11:50 AM"), ("Wednesday", "8–8:50 AM")],
    "D": [("Monday", "11–11:50 AM"), ("Wednesday", "9–9:50 AM"), ("Thursday", "12–12:50 PM")],
    "E": [("Monday", "12–12:50 PM"), ("Wednesday", "11–11:50 AM"), ("Friday", "9–9:50 AM")],
    "F": [("Tuesday", "8–8:50 AM"), ("Wednesday", "10–10:50 AM"), ("Friday", "11–11:50 AM")],
    "G": [("Tuesday", "9–9:50 AM"), ("Thursday", "8–8:50 AM"), ("Friday", "10–10:50 AM")],
    "H": [("Tuesday", "10–10:50 AM"), ("Wednesday", "12–12:50 PM"), ("Friday", "8–8:50 AM")],
    "FS": [("Friday", "2–5 PM")],
}

# settings flag name
TEMP_PREREG_OPEN_FLAG = "PREREG_TEMP_OPEN"

def _deadline_open() -> bool:
    # Temporary override has highest priority
    temp = getattr(settings, TEMP_PREREG_OPEN_FLAG, None)
    if temp is True:
        return True
    if temp is False:
        return False

    # Date-based logic (existing)
    deadline = getattr(settings, "PREREG_DEADLINE", None)
    if not deadline:
        return True
    now = timezone.now()
    if timezone.is_naive(deadline):
        deadline = timezone.make_aware(deadline, timezone.get_current_timezone())
    return now <= deadline


def _safe_str(v):
    return "" if v is None else str(v).strip()

def _is_blank_or_nan(v):
    s = _safe_str(v)
    return s == "" or s.lower() in ("nan", "none", "null")

def _split_codes(cell):
    if _is_blank_or_nan(cell):
        return []
    s = str(cell)
    for sep in ["|", ";"]:
        s = s.replace(sep, ",")
    if "," in s:
        tokens = s.split(",")
    else:
        tokens = s.split()
    return [t.strip().upper() for t in tokens if t.strip()]

@dataclass
class FacultyMini:
    first_name: str
    last_name: str | None
    email_id: str
    department: str

#---------------------------------------------------------------------------------------------------------------

def login(request):
    if request.method == "POST":
        identifier = request.POST['identifier'].lower()
        password = request.POST['password']

        try:
            if Admins.objects.filter(email_id=identifier).exists():
                Admin = Admins.objects.get(email_id=identifier)
                if check_password(password, Admin.password):
                    return redirect("/custom-admin/")
                else:
                    messages.error(request, "Invalid Admin Credentials")
                    return render(request, "login.html")
        except Admin.DoesNotExist:
            pass

        # 🔹 Case 1: Faculty
        try:
            if Faculty.objects.filter(email_id=identifier).exists():
                faculty = Faculty.objects.get(email_id=identifier)
                if check_password(password, faculty.password):
                    request.session["email_id"] = faculty.email_id  # stable across reloads
                    request.session["flash_ctx"] = {
                        "full_name": f"{faculty.first_name}",
                        "role_label": "Faculty",
                    }  # optional one-time
                    return render(request, "auth_successful.html", {"redirect_url": "/faculty_dashboard/"})
                else:
                    messages.error(request, "Invalid Faculty Credentials")
                    return render(request, "login.html")
        except Faculty.DoesNotExist:
            pass

        # 🔹 Case 2: Student
        try:
            if Student.objects.filter(roll_no=identifier).exists():
                student = Student.objects.get(roll_no=identifier)
            elif Student.objects.filter(email_id=identifier).exists():
                student = Student.objects.get(email_id=identifier)
            else:
                messages.error(request, "User Not Found")
                return render(request, "login.html")

            if check_password(password, student.password):
                request.session["roll_no"] = student.roll_no  # stable across reloads
                request.session["flash_ctx"] = {
                    "full_name": f"{student.first_name}",
                    "role_label": "Student",
                }  # optional one-time
                return render(request, "auth_successful.html", {"redirect_url": "/students_dashboard/"})
            else:
                messages.error(request, "Invalid Student Credentials")
        except Student.DoesNotExist:
            messages.error(request, "First Register Yourself")

    return render(request, "login.html")



def register(request):
    if request.method=='POST':
        firstname=request.POST['firstname']
        lastname=request.POST['lastname']
        roll_no=request.POST['roll_no']
        email=request.POST['email']
        department=request.POST['department']
        branch=request.POST['branch']
        password1=request.POST['password1']
        password2=request.POST['password2']
        hashed_password = make_password(password1)

        pattern1=re.compile(r'^(?:B|b|V|v|D|d|IM|im|MB|mb)\d{5}$')
        pattern2=re.compile(r'^(?:B|b|V|v|D|d|IM|im|MB|mb)\d{5}@students\.iitmandi\.ac\.in$')


        if(pattern1.match(roll_no) and pattern2.match(email) and (email[:len(roll_no)].lower()==roll_no.lower())):
            
            try:
                Student.objects.create(
                    first_name=firstname,
                    last_name=lastname,
                    email_id=email.lower(),
                    roll_no=roll_no.lower(),
                    password=hashed_password,
                    department=department,
                    branch=branch
                )
                messages.success(request, 'Registration successful! Please log in.')
                return redirect('/')  # Redirect to login page or another page
            except IntegrityError:
                messages.error(request, "Student with this email or roll number already exists.")
        else:
            messages.error(request, "Invalid Roll No. or Institute Email")

    return render(request, 'register.html')

def students_dashboard(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        # not logged in or session expired
        return redirect("/")

    student = get_object_or_404(Student, roll_no=roll_no)

    # ✅ Always calculate the current semester dynamically
    computed_semester = student.calculate_current_semester()

    flash = request.session.pop("flash_ctx", {})  # optional, disappears after first load
    semesters = (
        StudentCourse.objects.filter(student=student)
        .order_by('semester')
        .values_list('semester', flat=True)
        .distinct()
    )

    context = {
        "student": student,
        "computed_semester": computed_semester,  # ✅ pass live semester
        "full_name": flash.get("full_name"),
        "role_label": flash.get("role_label", "Student"),
        "semesters": semesters,
    }
    return render(request, "students_dashboard.html", context)


def faculty_dashboard(request):
    email_id = request.session.get("email_id")
    if not email_id:
        # not logged in or session expired
        return redirect("/")
    faculty = get_object_or_404(Faculty, email_id=email_id)

    flash = request.session.pop("flash_ctx", {})  # optional, disappears after first load
    context = {
        "faculty": faculty,
        "full_name": flash.get("full_name"),
        "role_label": flash.get("role_label", "Faculty"),
    }
    return render(request,'faculty_dashboard.html',context)

def pre_registration(request):
    """
    Pre-registration page:
     - preselected_by_slot comes from APPROVED enrollments (status='ENR') and PND (pending) for prefill.
     - new_cycle flag is True only when the student has no enrollments for current_sem
     - slot_map & slot_map_json exclude any courses the student has already PASSED
     - Mode removed (no course_mode selection here)
     - Pending (PND) rows count automatically toward total (Option A)
    """
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)
    branch = student.branch
    current_sem = student.calculate_current_semester()
    window_open = _deadline_open()

    # POST handling
    if request.method == "POST":
        if not window_open:
            messages.error(request, "Pre-registration window is closed.")
            return redirect("pre_registration")

        raw = request.POST.get("payload", "")
        try:
            payload = json.loads(raw) if raw else {}
        except Exception:
            messages.error(request, "Invalid submission payload.")
            return redirect("pre_registration")

        selections = payload.get("selections") or []
        if not isinstance(selections, list):
            messages.error(request, "Invalid selections.")
            return redirect("pre_registration")

        # Build course cache (codes -> Course)
        codes = [ (s.get("course_code") or "").strip().upper() for s in selections if s.get("course_code") ]
        course_map = {c.code.upper(): c for c in Course.objects.filter(code__in=codes)}

        created, updated, skipped_locked = 0, 0, 0
        with transaction.atomic():
            for sel in selections:
                slot = (sel.get("slot") or "").strip().upper()
                code = (sel.get("course_code") or "").strip().upper()
                if not slot or not code:
                    continue
                course = course_map.get(code)
                if not course or course.slot != slot:
                    continue

                # Remove non-enrolled duplicates for this slot (keep ENR)
                StudentCourse.objects.filter(
                    student=student, semester=current_sem, course__slot=slot
                ).exclude(status="ENR").delete()

                chosen_cat = (sel.get("category") or "").upper()
                if chosen_cat == "ALL":
                    # Try to pick a primary category if available
                    primary_cat = (
                        CourseBranch.objects
                        .filter(course=course, branch=branch)
                        .values_list("categories__code", flat=True)
                        .first()
                    )
                    chosen_cat = (primary_cat or "").upper()

                obj, created_now = StudentCourse.objects.get_or_create(
                    student=student,
                    course=course,
                    semester=current_sem,
                    defaults={
                        "status": "PND",
                        "type": chosen_cat or None,
                    }
                )

                if not created_now:
                    if obj.status == "ENR":
                        skipped_locked += 1
                        continue
                    # update pending entry with category if changed
                    obj.status = "PND"
                    obj.type = chosen_cat or obj.type
                    obj.save(update_fields=["status", "type"])
                    updated += 1
                else:
                    created += 1

        if created:
            messages.success(request, f"Submitted {created} request(s).")
        if updated:
            messages.info(request, f"Updated {updated} request(s).")
        if skipped_locked:
            pass
        return redirect("check_status_page")

    # --- GET: build options and prefill ---

    # compute passed courses codes (exclude them from future selections)
    passed_codes_qs = StudentCourse.objects.filter(
        student=student,
        status__in=["CMP"],
        outcome__iexact="PAS"
    ).values_list("course__code", flat=True).distinct()
    passed_codes = set([c.upper() for c in passed_codes_qs])

    # Build slot_map per category but exclude passed courses
    base_cb = CourseBranch.objects.filter(branch=branch)
    slot_map = {}
    CATEGORIES = getattr(settings, "COURSE_CATEGORIES", [
        {"code": "DC", "label": "Disciplinary Core (DC)"},
        {"code": "DE", "label": "Disciplinary Elective (DE)"},
        {"code": "IC", "label": "Institute Core (IC)"},
        {"code": "HSS", "label": "Humanities and Social Science (HSS)"},
        {"code": "FE", "label": "Free Elective (FE)"},
        {"code": "IKS", "label": "Indian Knowledge System (IKS)"},
        {"code": "ISTP", "label": "Interactive Socio-Technical Practicum (ISTP)"},
        {"code": "MTP", "label": "Major Technical Project (MTP)"},
    ])

    for slot in SLOTS:
        cat_map = {}
        for cat in [c["code"] for c in CATEGORIES]:
            qs = (
                Course.objects.filter(
                    slot=slot,
                    coursebranch__branch=branch,
                    coursebranch__categories__code=cat,
                )
                .exclude(code__in=passed_codes)
                .distinct()
            )
            cat_map[cat] = qs
        slot_map[slot] = cat_map

    # Build preselected_by_slot and locked_slots:
    # include ENR (locked) and PND (prefill) — user wanted PND to count automatically
    existing_approved = StudentCourse.objects.filter(
        student=student, semester=current_sem
    ).select_related("course")

    preselected_by_slot = {}
    locked_slots = set()
    for sc in existing_approved:
        if sc.course and sc.course.slot:
            # if ENR, lock; if PND, still preselect but not locked
            preselected_by_slot[sc.course.slot] = sc.course.code
            if sc.status == "ENR":
                locked_slots.add(sc.course.slot)

    # JSON-safe slot_map for frontend
    slot_map_json = {}
    for slot, cat_map in slot_map.items():
        slot_map_json[slot] = {}
        for cat, qs in cat_map.items():
            slot_map_json[slot][cat] = list(
                qs.values("code", "name", "credits")
            )

    # new_cycle: show banner only if the student has NO rows at all for current semester
    has_any_rows = StudentCourse.objects.filter(student=student, semester=current_sem).exists()
    new_cycle = not has_any_rows

    context = {
        "student": student,
        "categories": CATEGORIES,
        "slot_map": slot_map,
        "slot_map_json": mark_safe(json.dumps(slot_map_json)),
        "slots": SLOTS,
        "min_credit": 12,
        "max_credit": 22,
        "computed_semester": current_sem,
        "window_open": window_open,
        "preselected_by_slot": preselected_by_slot,
        "preselected_json": mark_safe(json.dumps(preselected_by_slot)),
        "locked_slots": list(locked_slots),
        "locked_slots_json": mark_safe(json.dumps(list(locked_slots))),
        "new_cycle": new_cycle,
    }
    return render(request, "registration/pre_registration.html", context)

@require_POST
def submit_preregistration(request):
    """Handle student pre-registration submissions"""
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)
    payload_raw = request.POST.get("payload", "")

    # --- Parse incoming JSON payload safely ---
    try:
        payload = json.loads(payload_raw) if payload_raw else {}
    except Exception:
        messages.error(request, "Invalid submission payload.")
        return redirect("prereg_page")

    semester = payload.get("semester")
    try:
        semester = int(semester)
    except Exception:
        messages.error(request, "Select a valid semester.")
        return redirect("prereg_page")

    selections = payload.get("selections") or []
    if not isinstance(selections, list) or not selections:
        messages.error(request, "No course selections found.")
        return redirect("prereg_page")

    # --- Prepare course map ---
    codes = [s.get("course_code", "").strip().upper() for s in selections if s.get("course_code")]
    codes = [c for c in codes if c]
    if not codes:
        messages.error(request, "Select at least one course.")
        return redirect("prereg_page")

    courses = {c.code.upper(): c for c in Course.objects.filter(code__in=codes)}
    missing = [c for c in codes if c not in courses]
    if missing:
        messages.error(request, f"Unknown course(s): {', '.join(missing)}")
        return redirect("prereg_page")

    # --- Credit total sanity check (optional backend validation) ---
    total_credits = sum(int(sel.get("credits") or 0) for sel in selections)
    # You can optionally check min/max credit logic here if needed.

    created, updated, skipped = 0, 0, 0

    with transaction.atomic():
        for sel in selections:
            code = (sel.get("course_code") or "").strip().upper()
            if not code:
                continue

            course = courses.get(code)
            if not course:
                continue

            chosen_cat = (sel.get("category") or "").upper() or None
            course_mode = sel.get("course_mode") or "REG"

            # Remove any older pending/dropped requests for this slot before saving
            StudentCourse.objects.filter(
                student=student,
                semester=semester,
                course__slot=course.slot,
                is_active_pre_reg=True,
            ).exclude(status="ENR").delete()

            # Create or update this record
            obj, is_created = StudentCourse.objects.get_or_create(
                student=student,
                course=course,
                semester=semester,
                is_active_pre_reg=True,
                defaults={
                    "status": "PND",
                    "course_mode": course_mode,
                    "type": chosen_cat,
                },
            )

            if not is_created:
                # If already exists but not finalized (not ENR), update to pending again
                if obj.status in ("DRP", "PND", "INS"):
                    obj.status = "PND"
                    obj.course_mode = course_mode
                    obj.type = chosen_cat
                    obj.is_active_pre_reg = True
                    obj.save(update_fields=["status", "course_mode", "type", "is_active_pre_reg"])
                    updated += 1
                else:
                    skipped += 1
            else:
                created += 1

    # --- Final message feedback ---
    if created:
        messages.success(request, f"Submitted {created} course request(s) for approval.")
    if updated:
        messages.info(request, f"Updated {updated} existing course(s).")
    if skipped:
        messages.warning(request, f"{skipped} course(s) already approved and not modified.")

    return redirect("check_status_page")

def check_status(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)
    current_sem = student.calculate_current_semester()

    # ✅ Fetch only *active* semester enrollments (fresh cycle only)
    enrollments = (
        StudentCourse.objects
        .filter(student=student, semester=current_sem, is_active_pre_reg=True)
        .select_related("course")
        .prefetch_related(Prefetch("course__faculties"))
        .order_by("course__slot", "course__code")
    )

    # ✅ Also filter pass/fail credit totals only from active records
    pf_total = (
        StudentCourse.objects
        .filter(student=student, course_mode="PF", is_active_pre_reg=True)
        .aggregate(total=Sum("course__credits"))
        .get("total") or 0
    )
    sem_pf_total = (
        StudentCourse.objects
        .filter(student=student, course_mode="PF", semester=current_sem, is_active_pre_reg=True)
        .aggregate(total=Sum("course__credits"))
        .get("total") or 0
    )

    # ✅ When admin resets pre-registration, there will be no active enrollments
    if not enrollments.exists():
        pass

    context = {
        "student": student,
        "enrollments": enrollments,
        "pf_total": pf_total,
        "sem_pf_total": sem_pf_total,
        "sem_display": current_sem,
    }
    return render(request, "registration/check_status.html", context)


@require_POST
def apply_mode_changes(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)
    raw = request.POST.get("payload") or ""
    try:
        data = json.loads(raw)
        changes = data.get("changes") or []
    except Exception:
        messages.error(request, "Invalid update payload.")
        return redirect("check_status_page")

    overall_pf_total = (
        StudentCourse.objects
        .filter(student=student, course_mode="PF")
        .aggregate(total=Sum("course__credits"))
        .get("total") or 0
    )

    applied, blocked = 0, 0
    for item in changes:
        sc_id = item.get("sc_id")
        new_mode = item.get("new_mode")
        if new_mode not in ["REG", "PF", "AUD"]:
            continue

        sc = StudentCourse.objects.select_related("course", "student").filter(id=sc_id, student=student).first()
        if not sc:
            continue

        # Skip approved courses
        if sc.status == "ENR":
            blocked += 1
            continue

        credits = sc.course.credits or 0
        current_mode = sc.course_mode

        # Skip no-change
        if new_mode == current_mode:
            continue

        # Calculate PF credit limits
        sem_pf_total = (
            StudentCourse.objects
            .filter(student=student, semester=sc.semester, course_mode="PF")
            .aggregate(total=Sum("course__credits"))
            .get("total") or 0
        )

        next_overall = overall_pf_total
        next_sem = sem_pf_total

        if new_mode == "PF" and current_mode != "PF":
            next_overall += credits
            next_sem += credits
        elif current_mode == "PF" and new_mode != "PF":
            next_overall -= credits
            next_sem -= credits

        if new_mode == "PF" and (next_overall > 9 or next_sem > 6):
            blocked += 1
            continue

        sc.course_mode = new_mode
        sc.save(update_fields=["course_mode"])
        applied += 1
        overall_pf_total = next_overall

    if applied:
        messages.success(request, f"Applied {applied} change(s).")
    if blocked:
        messages.warning(request, f"{blocked} change(s) blocked (approved or PF limit exceeded).")

    return redirect("check_status_page")


def registered_courses(request):
    return render(request, 'registration/registered_course.html')

# def course_request(request):
#     return render(request, 'instructor/course_request.html')


def student_profile(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)
    computed_semester = student.calculate_current_semester()

    cumulative = student.calculate_cumulative_metrics()
    cgpa = cumulative.get("CGPA", 0)
    total_earned = cumulative.get("TECR", 0)

    requirements = ProgramRequirement.objects.filter(branch=student.branch)
    total_required = sum(req.required_credits or 0 for req in requirements)

    # Base queryset: only CMP & PAS
    completed_qs = (
        StudentCourse.objects.filter(
            student=student,
            status="CMP",
            outcome="PAS",
        )
        .exclude(course_mode="AUD")   # ← Only exclude audit
        .select_related("course")
        .prefetch_related(
            Prefetch(
                "course__coursebranch_set",
                queryset=CourseBranch.objects.filter(branch=student.branch).prefetch_related("categories"),
                to_attr="cb_for_branch"
            )
        )
    )


    # Step 1: Determine each course’s primary (highest priority) category
    course_to_category = {}
    for sc in completed_qs:
        highest_cat = None
        highest_rank = float('inf')
        for cb in getattr(sc.course, "cb_for_branch", []):
            for cat in cb.categories.all():
                rank = CATEGORY_PRIORITY.get(cat.code.upper(), 999)
                if rank < highest_rank:
                    highest_rank = rank
                    highest_cat = cat.code.upper()
        if highest_cat:
            course_to_category[sc.course.code] = {
                "category": highest_cat,
                "course": sc.course,
                "semester": sc.semester,
            }

    # Step 2: Group courses by their final assigned category
    completed_by_category = {}
    for data in course_to_category.values():
        cat = data["category"]
        course = data["course"]
        if cat not in completed_by_category:
            completed_by_category[cat] = []
        completed_by_category[cat].append({
            "code": course.code,
            "name": course.name,
            "credits": course.credits,
            "slot": course.slot,
            "semester": data["semester"],
        })

    # Step 3: Compute category-wise credit progress
    categories_progress = []
    for req in requirements:
        cat_code = req.category
        cat_courses = completed_by_category.get(cat_code, [])
        completed_sum = sum(c["credits"] for c in cat_courses)
        remaining = max(req.required_credits - completed_sum, 0)
        percent = (completed_sum / req.required_credits * 100) if req.required_credits else 0
        categories_progress.append({
            "category": cat_code,
            "required": req.required_credits,
            "completed": completed_sum,
            "remaining": remaining,
            "percent": round(min(percent, 100), 1),
        })

    # Step 4: Totals
    remaining_total = max(total_required - total_earned, 0)
    overall_percent = round((total_earned / total_required) * 100, 1) if total_required else 0
    overall_percent = min(overall_percent, 100.0)

    context = {
        "student": student,
        "computed_semester": computed_semester,
        "cgpa": cgpa,
        "total_earned": total_earned,
        "total_required": total_required,
        "remaining": remaining_total,
        "overall_percent": overall_percent,
        "categories_progress": categories_progress,
        "completed_by_category": completed_by_category,
    }
    return render(request, "student_profile.html", context)

def faculty_profile(request):
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    faculty = get_object_or_404(Faculty, email_id=email_id)

    # All courses the faculty is teaching
    teaching_courses = Course.objects.filter(faculties=faculty).distinct()
    total_courses = teaching_courses.count()

    # Total unique students enrolled across all courses
    students_enrolled = (
        StudentCourse.objects.filter(course__in=teaching_courses, status="ENR")
        .values("student")
        .distinct()
        .count()
    )

    # Prepare chart data: { "Course Name": count_of_students }
    course_student_counts = {}
    for course in teaching_courses:
        count = (
            StudentCourse.objects.filter(course=course, status="ENR")
            .values("student")
            .distinct()
            .count()
        )
        course_student_counts[course.name] = count

    context = {
        "faculty": faculty,
        "total_courses": total_courses,
        "students_enrolled": students_enrolled,
        "teaching_courses": teaching_courses,
        "course_student_counts": course_student_counts, 
    }

    return render(request, "instructor/profile.html", context)

def student_edit_profile(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)

    if request.method == "POST":
        # Handle mobile number safely
        mobile_no = request.POST.get("mobile_no", "").strip()
        if mobile_no:
            try:
                student.mobile_no = int(mobile_no)
            except ValueError:
                # Ignore if not numeric (user error)
                pass
        else:
            student.mobile_no = None  # Allow clearing mobile number

        # Handle image upload
        if "profile_image" in request.FILES:
            student.profile_image = request.FILES["profile_image"]

        # Handle image removal
        elif "remove_image" in request.POST:
            if student.profile_image:
                student.profile_image.delete(save=False)
            student.profile_image = None

        student.save()
        return redirect(reverse("student_profile"))

    context = {"student": student}
    return render(request, "student_edit_profile.html", context)


def faculty_edit_profile(request):
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    faculty = get_object_or_404(Faculty, email_id=email_id)

    if request.method == "POST":
        # Handle mobile number safely
        mobile_no = request.POST.get("mobile_no", "").strip()
        if mobile_no:
            try:
                faculty.mobile_no = int(mobile_no)
            except ValueError:
                # Ignore if not numeric (user error)
                pass
        else:
            faculty.mobile_no = None  # Allow clearing mobile number

        # Handle image upload
        if "profile_image" in request.FILES:
            faculty.profile_image = request.FILES["profile_image"]

        # Handle image removal
        elif "remove_image" in request.POST:
            if faculty.profile_image:
                faculty.profile_image.delete(save=False)
            faculty.profile_image = None

        faculty.save()
        return redirect(reverse("faculty_profile"))

    # For a GET request, just pass the faculty object to the template
    context = {"faculty": faculty}
    return render(request, "instructor/edit_profile.html", context)



def custom_admin_home(request):
    return render(request, "admin/custom_admin_home.html")

def custom_admin_students(request):
    students = Student.objects.select_related("department","branch").all().order_by("roll_no")

    if request.method == 'POST':
        firstname = request.POST['firstname'].strip()
        lastname  = request.POST['lastname'].strip()
        roll_no   = request.POST['roll_no'].strip()
        email     = request.POST['email'].strip()
        password1 = request.POST['password']
        mobile_no = request.POST.get('mobile_no') or None

        # Get selected IDs from dropdowns
        dept_id = request.POST.get('department_id')
        branch_id = request.POST.get('branch_id')

        # Validate basic formats
        pattern1 = re.compile(r'^(?:B|b|V|v|D|d|IM|im|MB|mb)\d{5}$')
        pattern2 = re.compile(r'^(?:B|b|V|v|D|d|IM|im|MB|mb)\d{5}@students\.iitmandi\.ac\.in$')

        if pattern1.match(roll_no) and pattern2.match(email) and (email[:len(roll_no)].lower() == roll_no.lower()):
            try:
                dept = Department.objects.get(pk=dept_id) if dept_id else None
                branch = Branch.objects.get(pk=branch_id) if branch_id else None

                Student.objects.create(
                    first_name=firstname,
                    last_name=lastname,
                    email_id=email.lower(),
                    roll_no=roll_no.lower(),
                    password=make_password(password1),
                    department=dept,     # ForeignKey object, not string
                    branch=branch,       # ForeignKey object, not string
                    mobile_no=mobile_no,
                )
                messages.success(request, 'Registration successful.')
                return redirect('custom_admin_students')
            except Department.DoesNotExist:
                messages.error(request, "Selected department does not exist.")
            except Branch.DoesNotExist:
                messages.error(request, "Selected branch does not exist.")
            except IntegrityError:
                messages.error(request, "Student with this email or roll number already exists.")
        else:
            messages.error(request, "Invalid Roll No. or Institute Email")

    # Supply choices for dropdowns
    departments = Department.objects.all().order_by('code')
    branches = Branch.objects.all().order_by('name')
    return render(request, "admin/custom_admin_students.html", {
        "students": students,
        "departments": departments,
        "branches": branches,
    })

def custom_admin_students_bulk_add(request):
    if request.method == "POST":
        csv_file = request.FILES.get("csv_file")
        if not csv_file:
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect("custom_admin_students_bulk_add")

        temp_file_path = default_storage.save(f"temp/{csv_file.name}", csv_file)
        added_count = 0
        error_rows = []

        try:
            ext = csv_file.name.split('.')[-1].lower()
            if ext == 'csv':
                with default_storage.open(temp_file_path, mode='r') as file:
                    reader = csv.DictReader(file)
                    rows = list(reader)
            elif ext in ['xls', 'xlsx']:
                with default_storage.open(temp_file_path, 'rb') as f:
                    df = pd.read_excel(f)
                rows = df.to_dict(orient='records')
            else:
                messages.error(request, "Unsupported file format. Please upload CSV or Excel.")
                default_storage.delete(temp_file_path)
                return redirect("custom_admin_students_bulk_add")

            # Cache departments and branches for FK efficiency
            departments = {d.code: d for d in Department.objects.all()}
            branches = {b.name: b for b in Branch.objects.all()}

            # Required fields except last_name (nullable)
            required_fields = [
                "first_name", "email_id", "password",
                "roll_no", "department", "branch"
            ]

            def safe_str(val):
                return '' if val is None else str(val).strip()

            for idx, row in enumerate(rows, start=2):  # header = 1
                missing_fields = [f for f in required_fields if not row.get(f) or safe_str(row.get(f)) == '']
                if missing_fields:
                    error_rows.append(f"Row {idx}: Missing required fields: {', '.join(missing_fields)}.")
                    continue

                first_name = safe_str(row.get('first_name'))
                last_name = safe_str(row.get('last_name')) or None  # Nullable field
                email = safe_str(row.get('email_id')).lower()
                password_raw = row.get('password')
                roll_no = safe_str(row.get('roll_no')).lower()
                dept_code = safe_str(row.get('department')).upper()
                br_code = safe_str(row.get('branch'))

                dept = departments.get(dept_code)
                if not dept:
                    error_rows.append(f"Row {idx}: Department code '{dept_code}' not found.")
                    continue

                br = branches.get(br_code)
                if not br:
                    error_rows.append(f"Row {idx}: Branch code '{br_code}' not found.")
                    continue

                try:
                    student, created = Student.objects.get_or_create(
                        roll_no=roll_no,
                        defaults={
                            "first_name": first_name,
                            "last_name": last_name,
                            "email_id": email,
                            "password": make_password(password_raw),
                            "department": dept,
                            "branch": br,
                            "mobile_no": None,  # since no mobile column
                        },
                    )
                    if created:
                        added_count += 1
                    else:
                        error_rows.append(f"Row {idx}: Student with roll_no '{roll_no}' already exists.")
                except IntegrityError as e:
                    error_rows.append(f"Row {idx}: Integrity error: {str(e)}")
                except Exception as e:
                    error_rows.append(f"Row {idx}: Error saving student - {str(e)}")

        except Exception as e:
            messages.error(request, f"Failed to process file: {str(e)}")
            default_storage.delete(temp_file_path)
            return redirect("custom_admin_students_bulk_add")

        default_storage.delete(temp_file_path)

        if added_count:
            messages.success(request, f"Successfully added {added_count} students.")
        for error in error_rows:
            messages.error(request, error)
        return redirect("custom_admin_students_bulk_add")

    return render(request, "admin/bulk_add.html")


def delete_student_by_roll(request, roll_no):
    student = get_object_or_404(Student, roll_no=roll_no)
    if request.method == "POST":
        student.delete()
        return redirect('custom_admin_students')


def custom_admin_edit_student(request, roll_no):
    student = get_object_or_404(Student, roll_no=roll_no)

    if request.method == "POST":
        # Basic fields (safe .strip() and normalize where needed)
        new_roll   = (request.POST.get("roll_no") or "").strip().lower()
        student.roll_no   = new_roll or student.roll_no
        student.first_name = (request.POST.get("first_name") or "").strip()
        student.last_name  = (request.POST.get("last_name") or "").strip()
        student.email_id   = (request.POST.get("email_id") or "").strip().lower()

        # Department (optional)
        dept_id = request.POST.get("department")
        if dept_id:
            try:
                student.department = Department.objects.get(pk=dept_id)
            except Department.DoesNotExist:
                # keep previous value; optionally show a warning
                messages.warning(request, "Selected department not found. Keeping previous value.")
        else:
            student.department = None

        # Branch (optional)
        br_id = request.POST.get("branch")
        if br_id:
            try:
                student.branch = Branch.objects.get(pk=br_id)
            except Branch.DoesNotExist:
                messages.warning(request, "Selected branch not found. Keeping previous value.")
        else:
            student.branch = None

        # Mobile number — behave like student_edit_profile
        mobile_no = (request.POST.get("mobile_no") or "").strip()
        if mobile_no:
            try:
                student.mobile_no = int(mobile_no)
            except ValueError:
                # Ignore invalid mobile; do NOT block save or show an error
                pass
        else:
            # Allow clearing
            student.mobile_no = None

        # Profile image (optional upload/remove, just like profile view)
        if "profile_image" in request.FILES:
            student.profile_image = request.FILES["profile_image"]
        elif "remove_image" in request.POST:
            if student.profile_image:
                student.profile_image.delete(save=False)
            student.profile_image = None

        student.save()
        messages.success(request, "Student details updated successfully.")
        # Return to the same edit page (with possibly changed roll_no)
        return redirect("custom_admin_students")

    # GET: render form
    departments = Department.objects.all().order_by('code')
    branches = Branch.objects.all().order_by('name')
    return render(request, "admin/custom_admin_edit_student.html", {
        "student": student,
        "departments": departments,
        "branches": branches,
    })


def custom_admin_faculty(request):
    faculties = Faculty.objects.select_related("department").all().order_by("last_name", "first_name")
    departments = Department.objects.all().order_by('code')

    if request.method == 'POST':
        first_name = request.POST.get('firstname', '').strip()
        last_name = request.POST.get('lastname', '').strip()
        email_id = request.POST.get('email', '').strip().lower()
        department_id = request.POST.get('department')
        mobile_no = request.POST.get('mobile_no') or None
        raw_password = request.POST.get('password', '')

        # Basic validation
        if not EMAIL_RE.match(email_id):
            messages.error(request, "Invalid Institute Email")
            return redirect('custom_admin_faculty')

        # Resolve FK
        dept = None
        if department_id:
            try:
                dept = Department.objects.get(pk=department_id)
            except Department.DoesNotExist:
                messages.error(request, "Selected department does not exist.")
                return redirect('custom_admin_faculty')
        else:
            messages.error(request, "Please select a department.")
            return redirect('custom_admin_faculty')

        try:
            Faculty.objects.create(
                first_name=first_name,
                last_name=last_name,
                email_id=email_id,
                password=make_password(raw_password),
                department=dept,
                mobile_no=mobile_no,
            )
            messages.success(request, 'Faculty added successfully.')
            return redirect('custom_admin_faculty')
        except IntegrityError:
            messages.error(request, "Faculty with this email already exists.")
        except Exception as e:
            messages.error(request, f"Error adding faculty: {e}")

    return render(request, "admin/custom_admin_faculty.html", {
        "faculties": faculties,
        "departments": departments
    })


def custom_admin_faculty_bulk_add(request):
    if request.method == "POST":
        faculty_file = request.FILES.get("faculty_file")
        if not faculty_file:
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect("custom_admin_faculty_bulk_add")

        temp_file_path = default_storage.save(f"temp/{faculty_file.name}", faculty_file)
        added_count = 0
        error_rows = []

        try:
            ext = faculty_file.name.split('.')[-1].lower()
            if ext == 'csv':
                with default_storage.open(temp_file_path, mode='r') as file:
                    reader = csv.DictReader(file)
                    rows = list(reader)
            elif ext in ['xls', 'xlsx']:
                with default_storage.open(temp_file_path, 'rb') as f:
                    df = pd.read_excel(f)
                rows = df.to_dict(orient='records')
            else:
                messages.error(request, "Unsupported file format. Upload CSV or Excel.")
                return redirect("custom_admin_faculty_bulk_add")

            departments = {d.code: d for d in Department.objects.all()}

            def safe_str(val):
                return '' if val is None else str(val).strip()

            for idx, row in enumerate(rows, start=2):
                first_name = safe_str(row.get('first_name'))
                last_name = safe_str(row.get('last_name'))
                email = safe_str(row.get('email_id')).lower()
                password_raw = row.get('password')
                dept_code = safe_str(row.get('department')).upper()

                missing_fields = [f for f,v in [('first_name', first_name),  ('email_id', email), ('password', password_raw), ('department', dept_code)] if not v]
                if missing_fields:
                    error_rows.append(f"Row {idx}: Missing required fields: {', '.join(missing_fields)}.")
                    continue

                dept = departments.get(dept_code)
                if not dept:
                    error_rows.append(f"Row {idx}: Department code '{dept_code}' not found.")
                    continue

                mobile_no_raw = row.get('mobile_no')
                mobile_no = safe_str(mobile_no_raw) if mobile_no_raw else None

                try:
                    faculty, created = Faculty.objects.get_or_create(
                        email_id=email,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name if last_name else None,
                            'password': make_password(password_raw),
                            'department': dept,
                            'mobile_no': mobile_no if mobile_no else None,
                        }
                    )
                    if created:
                        added_count += 1
                    else:
                        error_rows.append(f"Row {idx}: Faculty with email '{email}' already exists.")
                except IntegrityError as e:
                    error_rows.append(f"Row {idx}: Integrity error: {str(e)}")
                except Exception as e:
                    error_rows.append(f"Row {idx}: Error saving faculty - {str(e)}")

        except Exception as e:
            messages.error(request, f"Failed to process file: {str(e)}")
            default_storage.delete(temp_file_path)
            return redirect("custom_admin_faculty_bulk_add")

        default_storage.delete(temp_file_path)

        if added_count:
            messages.success(request, f"Successfully added {added_count} faculties.")
        for err in error_rows:
            messages.error(request, err)
        return redirect("custom_admin_faculty_bulk_add")

    return render(request, "admin/custom_admin_faculty_bulk.html")



def delete_faculty(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)
    if request.method == "POST":
        faculty.delete()
        return redirect('custom_admin_faculty')

def custom_admin_edit_faculty(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)
    departments = Department.objects.all().order_by('code')

    if request.method == "POST":
        # Basic text fields
        faculty.first_name = (request.POST.get("first_name") or "").strip()
        faculty.last_name  = (request.POST.get("last_name") or "").strip()
        faculty.email_id   = (request.POST.get("email_id") or "").strip().lower()

        # Department (optional)
        dept_id = request.POST.get("department")
        if dept_id:
            try:
                faculty.department = Department.objects.get(pk=dept_id)
            except Department.DoesNotExist:
                messages.warning(request, "Selected department not found. Keeping previous value.")

        # Mobile number — parse like in student_edit_profile
        mobile_str = (request.POST.get("mobile_no") or "").strip()
        if mobile_str:
            try:
                faculty.mobile_no = int(mobile_str)
            except ValueError:
                pass
        else:
            # Allow clearing
            faculty.mobile_no = None

        # Password (only if provided)
        raw_pwd = (request.POST.get("password") or "").strip()
        if raw_pwd:
            faculty.password = make_password(raw_pwd)

        # Profile image upload/remove
        if "profile_image" in request.FILES:
            faculty.profile_image = request.FILES["profile_image"]
        elif "remove_image" in request.POST:
            if faculty.profile_image:
                faculty.profile_image.delete(save=False)
            faculty.profile_image = None

        faculty.save()
        messages.success(request, "Faculty details updated successfully.")
        return redirect("custom_admin_faculty")

    return render(request, "admin/custom_admin_edit_faculty.html", {
        "faculty": faculty,
        "departments": departments
    })

def custom_admin_branch(request):
    branches = Branch.objects.select_related("department").order_by("department__code", "name")
    departments = Department.objects.all().order_by("code")

    # derive choices from model (list of (code, label))
    branch_choices = Branch.BRANCHES

    if request.method == "POST":
        dept_id = request.POST.get("department_id")
        branch_code = request.POST.get("branch_code")  # e.g., "CSE"

        if not dept_id or not branch_code:
            messages.error(request, "Please select both Department and Branch.")
            return redirect("custom_admin_branch")

        try:
            dept = Department.objects.get(pk=dept_id)
        except Department.DoesNotExist:
            messages.error(request, "Selected department does not exist.")
            return redirect("custom_admin_branch")

        valid_codes = {c for c, _ in branch_choices}
        if branch_code not in valid_codes:
            messages.error(request, "Invalid branch selection.")
            return redirect("custom_admin_branch")

        try:
            # Branch.name stores the code because name has choices=BRANCHES
            obj, created = Branch.objects.get_or_create(name=branch_code, defaults={"department": dept})
            if not created:
                # if it exists but department differs, update it
                if obj.department_id != dept.id:
                    obj.department = dept
                    obj.save()
                    messages.success(request, f"Branch {obj.name} moved to {dept.code}.")
                else:
                    messages.info(request, "Branch already exists under this department.")
            else:
                messages.success(request, "Branch created successfully.")
        except IntegrityError:
            messages.error(request, "Branch with the same code already exists.")
        return redirect("custom_admin_branch")

    return render(request, "admin/custom_admin_branch.html", {
        "branches": branches,
        "departments": departments,
        "branch_choices": branch_choices,
    })


def ajax_branches_json(request):
    dept_id = request.GET.get("department_id")
    qs = Branch.objects.none()
    if dept_id:
        qs = Branch.objects.filter(department_id=dept_id).order_by("name")
    data = [{"id": b.id, "label": getattr(b, "get_name_display", lambda: b.name)()} for b in qs]
    return JsonResponse({"options": data})



def custom_admin_courses(request):
    courses = Course.objects.all().order_by("code")
    slot_choices = Course.SLOT_CHOICES  # for form select

    if request.method == "POST":
        code = request.POST.get("code", "").strip().upper()
        name = request.POST.get("name", "").strip()
        credits = request.POST.get("credits")
        ltpc = request.POST.get("LTPC", "").strip().upper()
        slot = request.POST.get("slot")

        # Basic validation
        if not code or not name or not credits or not slot:
            messages.error(request, "Please fill in all required fields.")
            return redirect("custom_admin_courses")

        try:
            credits = int(credits)
        except ValueError:
            messages.error(request, "Credits must be an integer.")
            return redirect("custom_admin_courses")

        try:
            Course.objects.create(
                code=code, name=name, credits=credits, LTPC=ltpc, slot=slot
            )
            messages.success(request, "Course added successfully.")
        except IntegrityError as e:
            messages.error(request, f"Could not add course: {e}")
        except Exception as e:
            messages.error(request, f"Error: {e}")
        return redirect("custom_admin_courses")

    return render(request, "admin/custom_admin_courses.html", {
        "courses": courses,
        "slot_choices": slot_choices,
    })

def custom_admin_courses_bulk(request):
    if request.method == "POST":
        course_file = request.FILES.get("course_file")
        if not course_file:
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect("custom_admin_courses_bulk")

        temp_file_path = default_storage.save(f"temp/{course_file.name}", course_file)
        added_count = 0
        error_rows = []

        try:
            ext = course_file.name.split('.')[-1].lower()
            if ext == 'csv':
                with default_storage.open(temp_file_path, mode='r') as file:
                    reader = csv.DictReader(file)
                    rows = list(reader)
            elif ext in ['xls', 'xlsx']:
                with default_storage.open(temp_file_path, 'rb') as f:
                    df = pd.read_excel(f)
                rows = df.to_dict(orient='records')
            else:
                messages.error(request, "Unsupported file format. Upload CSV or Excel.")
                return redirect("custom_admin_courses_bulk")

            valid_slots = {val for val, _ in Course.SLOT_CHOICES}

            def safe_str(val):
                return '' if val is None else str(val).strip()

            for idx, row in enumerate(rows, start=2):  # header = 1, data starts at 2
                code = safe_str(row.get('Course Code')).upper()
                name = safe_str(row.get('Course Name'))
                status = safe_str(row.get('Status in course booklet')) or 'Yes'
                ltpc = safe_str(row.get('L-T-P-C')).upper()
                slot = safe_str(row.get('Slot')).upper()
                credit_val = row.get('Credit')
                try:
                    credits = int(float(credit_val)) if credit_val not in [None, ''] else None
                except (ValueError, TypeError):
                    credits = None

                if not (code and name and slot and credits is not None):
                    error_rows.append(f"Row {idx}: Missing required fields.")
                    continue
                if slot not in valid_slots:
                    error_rows.append(f"Row {idx}: Invalid slot '{slot}'.")
                    continue

                try:
                    course, created = Course.objects.get_or_create(
                        code=code,
                        defaults={
                            'name': name,
                            'status': status,
                            'LTPC': ltpc,
                            'slot': slot,
                            'credits': credits,
                        }
                    )
                    if created:
                        added_count += 1
                    else:
                        # Uncomment below to update existing course info if needed
                        # course.name = name
                        # course.status = status
                        # course.LTPC = ltpc
                        # course.slot = slot
                        # course.credits = credits
                        # course.save()
                        pass
                except IntegrityError:
                    error_rows.append(f"Row {idx}: Course with code '{code}' already exists.")
                except Exception as e:
                    error_rows.append(f"Row {idx}: Error saving course - {str(e)}")

        except Exception as e:
            messages.error(request, f"Failed to process file: {str(e)}")
            default_storage.delete(temp_file_path)
            return redirect("custom_admin_courses_bulk")

        default_storage.delete(temp_file_path)

        if added_count:
            messages.success(request, f"Successfully added {added_count} new courses.")
        for err in error_rows:
            messages.error(request, err)
        return redirect("custom_admin_courses_bulk")

    return render(request, "admin/bulk_add_courses.html")



def custom_admin_edit_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    slot_choices = Course.SLOT_CHOICES

    if request.method == "POST":
        course.code = request.POST.get("code", "").strip().upper()
        course.name = request.POST.get("name", "").strip()
        credits = request.POST.get("credits")
        course.LTPC = request.POST.get("LTPC", "").strip().upper()
        course.slot = request.POST.get("slot")

        try:
            course.credits = int(credits)
        except (TypeError, ValueError):
            messages.error(request, "Credits must be an integer.")
            return redirect("custom_admin_edit_course", course_id=course.id)

        try:
            course.save()
            messages.success(request, "Course updated successfully.")
            return redirect("custom_admin_courses")
        except IntegrityError as e:
            messages.error(request, f"Could not update course: {e}")
        except Exception as e:
            messages.error(request, f"Error: {e}")

    return render(request, "admin/custom_admin_edit_course.html", {
        "course": course,
        "slot_choices": slot_choices,
    })


from django.views.decorators.http import require_POST
@require_POST
def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    try:
        course.delete()
        messages.success(request, "Course deleted.")
    except Exception as e:
        messages.error(request, f"Could not delete: {e}")
    return redirect("custom_admin_courses")


def manage_course_faculties(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    faculties = Faculty.objects.select_related("department").order_by("last_name", "first_name")

    if request.method == "POST":
        selected_ids = request.POST.getlist("faculty_ids")  # list of strings
        # Replace all current assignments with selected ones
        try:
            # Ensure integers
            faculty_pks = [int(pk) for pk in selected_ids]
            course.faculties.set(faculty_pks)  # ManyToMany replace
            course.save()
            messages.success(request, "Faculties updated for this course.")
        except Exception as e:
            messages.error(request, f"Could not update faculties: {e}")
        return redirect("course_instructors_assign")

    assigned_ids = set(course.faculties.values_list("id", flat=True))
    return render(request, "admin/custom_admin_course_faculties.html", {
        "course": course,
        "faculties": faculties,
        "assigned_ids": assigned_ids,
    })


def course_branch_index(request):
    courses = Course.objects.all().order_by("code")
    return render(request, "admin/course_branch_index.html", {
        "courses": courses,
    })

@transaction.atomic
def manage_course_branches(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    categories = Category.objects.order_by("code")
    branches = Branch.objects.select_related("department").order_by("name")

    if request.method == "POST":
        for br in branches:
            ids = request.POST.getlist(f"category_{br.id}[]")
            cb, _ = CourseBranch.objects.get_or_create(course=course, branch=br)
            qs = Category.objects.filter(id__in=ids)
            cb.categories.set(qs) if ids else cb.categories.clear()
        return redirect("course_branch_index")

    # Build preselected sets from DB for this course
    cb_for_course = CourseBranch.objects.filter(course=course).prefetch_related("categories")
    selected_map = {
        cb.branch_id: set(cb.categories.values_list("id", flat=True))
        for cb in cb_for_course
    }
    for br in branches:
        br.selected_category_ids = selected_map.get(br.id, set())

    return render(
        request,
        "admin/custom_admin_course_branches.html",
        {"course": course, "branches": branches, "categories": categories},
    )


def bulk_upload_course_branch(request):
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            return JsonResponse({"ok": False, "message": "Please upload a CSV or Excel file.", "errors": []}, status=400)

        temp_path = default_storage.save(f"temp/{upload.name}", upload)
        errors = []
        applied_links = 0
        rows_processed = 0

        try:
            ext = upload.name.split(".")[-1].lower()
            if ext == "csv":
                with default_storage.open(temp_path, mode="r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            elif ext in ["xls", "xlsx"]:
                with default_storage.open(temp_path, "rb") as f:
                    df = pd.read_excel(f, engine="openpyxl")
                rows = df.to_dict(orient="records")
            else:
                return JsonResponse({"ok": False, "message": "Unsupported file format. Upload CSV or Excel.", "errors": []}, status=400)

            if not rows:
                return JsonResponse({"ok": False, "message": "The uploaded file has no data rows.", "errors": []}, status=400)

            headers = set(rows[0].keys())
            must_have = {"Course Code", "Course Name"}
            missing = [h for h in must_have if h not in headers]
            if missing:
                return JsonResponse({"ok": False, "message": "Invalid header row.", "errors": [f"Missing required columns: {', '.join(missing)}"]}, status=400)
            if not any(h in headers for h in CATEGORY_HEADERS):
                return JsonResponse({"ok": False, "message": "No category columns present.", "errors": [f"Provide any of: {', '.join(CATEGORY_HEADERS)}"]}, status=400)

            # Preload references
            branches = list(Branch.objects.all())
            branches_by_code = {b.name.strip().upper(): b for b in branches}  # Branch.name holds the code
            all_branches = branches  # for ALL token
            categories_by_code = {c.code.strip().upper(): c for c in Category.objects.all()}
            missing_cats = [h for h in CATEGORY_HEADERS if h not in categories_by_code]
            if missing_cats:
                return JsonResponse({"ok": False, "message": "Missing Category rows in DB.", "errors": [f"Create Category for codes: {', '.join(missing_cats)}"]}, status=400)

            with transaction.atomic():
                for rix, row in enumerate(rows, start=2):
                    code = _safe_str(row.get("Course Code")).upper()
                    name = _safe_str(row.get("Course Name"))

                    if not code and not name:
                        errors.append(f"Row {rix}: Both Course Code and Course Name are empty; skipped.")
                        continue

                    course = None
                    if code:
                        course = Course.objects.filter(code__iexact=code).first()
                    if course is None and name:
                        course = Course.objects.filter(name__iexact=name).first()

                    if course is None:
                        errors.append(f"Row {rix}: Course not found by code '{code}' or name '{name}'. Skipped.")
                        continue

                    rows_processed += 1

                    # Process each category column
                    for header in CATEGORY_HEADERS:
                        if header not in headers:
                            continue
                        cell = row.get(header)

                        # Skip blanks/NaN entirely
                        if _is_blank_or_nan(cell):
                            continue

                        tokens = _split_codes(cell)
                        # If the cell explicitly contains ALL (case-insensitive), apply to all branches
                        # Detect ALL even if mixed with other separators/tokens
                        has_all = any(tok.upper() == "ALL" for tok in (tokens if tokens else [_safe_str(cell).upper()]))

                        if has_all:
                            target_branches = all_branches
                        else:
                            # Use parsed codes; unknown codes are reported individually
                            target_branches = []
                            for br_code in tokens:
                                if br_code == "ALL":
                                    # Already handled by has_all; skip here
                                    continue
                                br = branches_by_code.get(br_code)
                                if not br:
                                    errors.append(f"Row {rix}: Unknown Branch Code '{br_code}' in column '{header}'.")
                                    continue
                                target_branches.append(br)

                        if not target_branches:
                            continue  # nothing to do for this cell

                        cat = categories_by_code[HEADER_TO_CATEGORY[header]]

                        for br in target_branches:
                            cb, _ = CourseBranch.objects.get_or_create(course=course, branch=br)
                            # Add category if missing; never remove others
                            if not cb.categories.filter(id=cat.id).exists():
                                cb.categories.add(cat)
                                applied_links += 1

            return JsonResponse({
                "ok": True,
                "message": "Processed file.",
                "rows": rows_processed,
                "links_added": applied_links,
                "errors": errors[:100],
            }, status=200)

        except Exception as e:
            return JsonResponse({
                "ok": False,
                "message": f"Failed to process file: {str(e)}",
                "errors": errors[:100]
            }, status=500)
        finally:
            try:
                default_storage.delete(temp_path)
            except Exception:
                pass

    # GET
    return render(request, "admin/bulk_upload_course_branches.html")

def requirements_index(request):
    branches = Branch.objects.select_related("department").order_by("department__code", "name")
    return render(request, "admin/requirements_index.html", {"branches": branches})

def manage_branch_requirements(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    categories = ProgramRequirement.CATEGORY_CHOICES

    # map: category -> required
    existing = {
        r.category: r.required_credits for r in ProgramRequirement.objects.filter(branch=branch)
    }

    if request.method == "POST":
        try:
            with transaction.atomic():
                for cat, _label in categories:
                    req_val = request.POST.get(f"{cat}_req", "").strip()
                    if req_val == "":
                        # Treat blank as 0 to keep it simple; adjust if you prefer to delete rows when blank
                        req_i = 0
                    else:
                        try:
                            req_i = int(req_val)
                        except ValueError:
                            raise ValueError(f"{cat}: required must be an integer.")
                        if req_i < 0:
                            raise ValueError(f"{cat}: required cannot be negative.")

                    ProgramRequirement.objects.update_or_create(
                        branch=branch,
                        category=cat,
                        defaults={"required_credits": req_i},
                    )
            messages.success(request, "Requirements saved.")
        except Exception as e:
            messages.error(request, f"Could not save: {e}")
        return redirect("manage_branch_requirements", branch_id=branch.id)

    # Build rows for template
    rows = []
    for cat, label in categories:
        req_i = existing.get(cat, 0)
        rows.append({"cat": cat, "label": label, "req": req_i})

    return render(request, "admin/manage_branch_requirements.html", {
        "branch": branch,
        "rows": rows,
    })


def bulk_upload_program_requirements(request):
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            return JsonResponse({
                "ok": False,
                "message": "Please upload a CSV or Excel file.",
                "errors": []
            }, status=400)

        temp_file_path = default_storage.save(f"temp/{upload.name}", upload)
        created_count = 0
        updated_count = 0
        error_rows = []

        try:
            # Read rows
            ext = upload.name.split(".")[-1].lower()
            if ext == "csv":
                with default_storage.open(temp_file_path, mode="r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            elif ext in ["xls", "xlsx"]:
                with default_storage.open(temp_file_path, "rb") as f:
                    # Use openpyxl for .xlsx
                    df = pd.read_excel(f, engine="openpyxl")
                rows = df.to_dict(orient="records")
            else:
                return JsonResponse({
                    "ok": False,
                    "message": "Unsupported file format. Upload CSV or Excel.",
                    "errors": []
                }, status=400)

            # Guard: no data
            if not rows:
                return JsonResponse({
                    "ok": False,
                    "message": "The uploaded file contains no data rows.",
                    "errors": []
                }, status=400)

            # Header validation
            headers = set(rows[0].keys())
            required_headers = {"Branch Code"} | set(CATEGORY_MAP.keys())
            missing = [h for h in required_headers if h not in headers]
            if missing:
                return JsonResponse({
                    "ok": False,
                    "message": "Missing required columns.",
                    "errors": [f"Missing required columns: {', '.join(missing)}"]
                }, status=400)

            # Preload Branch by code.
            # IMPORTANT: Branch.name stores the code (choices), e.g., "CSE", "DSE", etc.
            branch_by_code = {b.name.strip().upper(): b for b in Branch.objects.all()}

            # Existing ProgramRequirement map
            existing = {
                (pr.branch_id, pr.category): pr
                for pr in ProgramRequirement.objects.all().only("id", "branch_id", "category", "required_credits")
            }

            def parse_int(val):
                if val is None:
                    return None
                s = str(val).strip()
                if s == "" or s.lower() in ("nan", "none"):
                    return None
                try:
                    return int(float(s))
                except Exception:
                    return None

            to_create = []
            to_update = []

            for idx, row in enumerate(rows, start=2):  # header row = 1
                branch_code = str(row.get("Branch Code", "")).strip().upper()
                if not branch_code:
                    error_rows.append(f"Row {idx}: Missing Branch Code.")
                    continue
                branch = branch_by_code.get(branch_code)
                if not branch:
                    error_rows.append(f"Row {idx}: Unknown Branch Code '{branch_code}'.")
                    continue

                any_valid = False
                for header_key, category_code in CATEGORY_MAP.items():
                    credits = parse_int(row.get(header_key))
                    if credits is None:
                        continue  # blank/invalid => no change
                    any_valid = True
                    key = (branch.id, category_code)
                    pr = existing.get(key)
                    if pr:
                        if pr.required_credits != credits:
                            pr.required_credits = credits
                            to_update.append(pr)
                    else:
                        to_create.append(ProgramRequirement(
                            branch=branch,
                            category=category_code,
                            required_credits=credits,
                        ))

                if not any_valid:
                    error_rows.append(f"Row {idx}: No valid integer credits for any category; row skipped.")

            # Apply DB changes atomically
            with transaction.atomic():
                if to_create:
                    ProgramRequirement.objects.bulk_create(to_create, batch_size=1000)
                    created_count = len(to_create)
                if to_update:
                    ProgramRequirement.objects.bulk_update(to_update, ["required_credits"], batch_size=1000)
                    updated_count = len(to_update)

            return JsonResponse({
                "ok": True,
                "created": created_count,
                "updated": updated_count,
                "errors": error_rows[:50],  # cap to avoid huge payloads
            }, status=200)

        except Exception as e:
            return JsonResponse({
                "ok": False,
                "message": f"Failed to process file: {str(e)}",
                "errors": error_rows[:50] if error_rows else []
            }, status=500)
        finally:
            try:
                default_storage.delete(temp_file_path)
            except Exception:
                # Silent cleanup failure
                pass

    # GET: render page
    return render(request, "admin/bulk_upload_program_requirements.html")


def course_instructors_assign(request):
    courses = (
        Course.objects
        .all()
        .prefetch_related("faculties__department")
        .order_by("code")
    )

    # Build mapping of course.id to list of mini faculty objects (first_name, last_name, email, department name)
    course_faculty_map: dict[int, list[FacultyMini]] = defaultdict(list)
    for course in courses:
        for f in course.faculties.all():
            dept_name = f.department.name if f.department else "—"
            course_faculty_map[course.id].append(
                FacultyMini(
                    first_name=f.first_name,
                    last_name=f.last_name or "",
                    email_id=f.email_id,
                    department=dept_name,
                )
            )

    # Filters for dropdowns
    departments = (
        Department.objects
        .order_by("name")
        .values_list("name", flat=True)
        .distinct()
    )
    slots = (
        Course.objects
        .order_by()
        .values_list("slot", flat=True)
        .distinct()
    )

    return render(
        request,
        "admin/course_instructors_assign.html",  # use the template you created for this page
        {
            "courses": courses,
            "course_faculty_map": course_faculty_map,
            "departments": departments,
            "slots": slots,
        },
    )

def course_instructors_assign_bulk(request):
    if request.method == "POST":
        upload = request.FILES.get("csv_file")
        if not upload:
            return JsonResponse({"ok": False, "message": "Please upload a CSV or Excel file.", "errors": []}, status=400)

        # Optional controls provided by the form
        conflict = (request.POST.get("conflict_policy") or "append").strip().lower()
        validation = (request.POST.get("validation_level") or "strict").strip().lower()
        if conflict not in {"append", "replace"}:
            conflict = "append"
        if validation not in {"strict", "lenient"}:
            validation = "strict"

        temp_path = default_storage.save(f"temp/{upload.name}", upload)
        errors = []
        links_added = 0
        rows_processed = 0

        try:
            ext = upload.name.split(".")[-1].lower()
            # Load rows as a list of dicts with keys 'course_code' and 'faculty_email'
            if ext == "csv":
                with default_storage.open(temp_path, mode="r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            elif ext in ["xls", "xlsx"]:
                with default_storage.open(temp_path, "rb") as f:
                    df = pd.read_excel(f, engine="openpyxl")
                rows = df.to_dict(orient="records")
            else:
                return JsonResponse(
                    {"ok": False, "message": "Unsupported file format. Upload CSV or Excel.", "errors": []},
                    status=400
                )

            if not rows:
                return JsonResponse({"ok": False, "message": "The uploaded file has no data rows.", "errors": []}, status=400)

            # Header validation (case-insensitive normalize)
            sample_keys = {str(k).strip().lower() for k in rows[0].keys()}
            need = {"course_code", "faculty_email"}
            missing = [h for h in need if h not in sample_keys]
            if missing:
                return JsonResponse(
                    {"ok": False, "message": "Invalid header row.", "errors": [f"Missing required columns: {', '.join(need)}"]},
                    status=400
                )

            # Normalize keys to lower for consistent access
            norm_rows = []
            for r in rows:
                norm = {str(k).strip().lower(): v for k, v in r.items()}
                norm_rows.append(norm)

            # Preload references
            codes = set()
            emails = set()
            for rix, row in enumerate(norm_rows, start=2):
                code = str(row.get("course_code") or "").strip()
                email = str(row.get("faculty_email") or "").strip()
                if not code or not email:
                    msg = f"Row {rix}: empty course_code or faculty_email; skipped."
                    if validation == "strict":
                        return JsonResponse({"ok": False, "message": msg, "errors": errors}, status=400)
                    errors.append(msg)
                    continue
                codes.add(code)
                emails.add(email)

            course_by_code = {c.code: c for c in Course.objects.filter(code__in=codes)}
            faculty_by_email = {f.email_id: f for f in Faculty.objects.filter(email_id__in=emails)}

            with transaction.atomic():
                # Replace clears once per course before applying rows
                if conflict == "replace":
                    for code in sorted(codes):
                        course = course_by_code.get(code)
                        if not course:
                            msg = f"Unknown course: {code}"
                            if validation == "strict":
                                return JsonResponse({"ok": False, "message": msg, "errors": errors}, status=400)
                            errors.append(msg)
                            continue
                        course.faculties.clear()

                # Apply rows
                for rix, row in enumerate(norm_rows, start=2):
                    code = str(row.get("course_code") or "").strip()
                    email = str(row.get("faculty_email") or "").strip()
                    if not code or not email:
                        # already recorded earlier when collecting codes/emails
                        continue

                    course = course_by_code.get(code)
                    if not course:
                        msg = f"Row {rix}: Unknown course: {code}"
                        if validation == "strict":
                            return JsonResponse({"ok": False, "message": msg, "errors": errors}, status=400)
                        errors.append(msg)
                        continue

                    faculty = faculty_by_email.get(email)
                    if not faculty:
                        msg = f"Row {rix}: Unknown faculty: {email}"
                        if validation == "strict":
                            return JsonResponse({"ok": False, "message": msg, "errors": errors}, status=400)
                        errors.append(msg)
                        continue

                    rows_processed += 1
                    if not course.faculties.filter(pk=faculty.pk).exists():
                        course.faculties.add(faculty)
                        links_added += 1

            return JsonResponse({
                "ok": True,
                "message": "Processed file.",
                "rows": rows_processed,
                "links_added": links_added,
                "errors": errors[:100],
            }, status=200)

        except Exception as e:
            return JsonResponse({
                "ok": False,
                "message": f"Failed to process file: {str(e)}",
                "errors": errors[:100]
            }, status=500)
        finally:
            try:
                default_storage.delete(temp_path)
            except Exception:
                pass

    # GET
    return render(request, "admin/bulk_add_course_instructors.html")




##Should be changed -- Faculty Advisor
@require_POST
def advisor_request_action(request):
    # Assume Admins act as advisors; refine later to advisor mapping
    admin_email = request.session.get("email_id")
    if not admin_email or not Admins.objects.filter(email_id=admin_email).exists():
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    sc_id = request.POST.get("sc_id")
    action = (request.POST.get("action") or "").lower()

    sc = get_object_or_404(StudentCourse, id=sc_id)

    if sc.status != "INS":
        messages.info(request, "Request not ready for advisor approval.")
        return redirect("custom_admin_home")

    if action == "approve":
        sc.status = "ENR"
        sc.outcome = "UNK"
        sc.save(update_fields=["status", "outcome"])
        messages.success(request, "Student enrolled for this course.")
    elif action == "reject":
        sc.status = "DRP"
        sc.save(update_fields=["status"])
        messages.success(request, "Request rejected.")
    else:
        messages.error(request, "Unknown action.")
    return redirect("custom_admin_home")


def instructor_requests(request):
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    instructor = get_object_or_404(Faculty, email_id=email_id)

    # Filters
    status_filter = request.GET.get("status", "").upper() or "PND"
    course_code = (request.GET.get("course") or "").strip().upper()
    slot = (request.GET.get("slot") or "").strip().upper()

    # ✅ Fetch only active prereg data
    qs = (
        StudentCourse.objects.filter(
            course__faculties__email_id=email_id,
            is_active_pre_reg=True,
        )
        .select_related("course", "student")
        .prefetch_related("course__faculties")
        .order_by("course__code", "student__roll_no")
        .distinct()
    )

    # ✅ Fix: default filter should exclude already enrolled
    if status_filter == "PND":
        qs = qs.filter(status="PND")
    elif status_filter == "INS":
        qs = qs.filter(status="INS")
    elif status_filter == "ENR":
        # Faculty can view ENR if they explicitly select it
        qs = qs.filter(status="ENR")
    elif status_filter == "DRP":
        qs = qs.filter(status="DRP")

    # Optional filters
    if course_code:
        qs = qs.filter(course__code=course_code)
    if slot:
        qs = qs.filter(course__slot=slot)

    # Instructor’s available slots and courses (for filter UI)
    slots = (
        Course.objects.filter(faculties__email_id=email_id)
        .values_list("slot", flat=True)
        .distinct()
    )
    my_courses = (
        Course.objects.filter(faculties__email_id=email_id)
        .order_by("code")
        .values_list("code", "name", "slot")
    )

    context = {
        "instructor": instructor,
        "faculty":instructor,

        "requests": qs,
        "status_filter": status_filter,
        "course_code": course_code,
        "slot_filter": slot,
        "slots": slots,
        "my_courses": my_courses,
    }

    # 🟢 UX improvement: If all ENR, show helpful message
    if not qs.exists() and status_filter == "PND":
        messages.info(
            request,
            "No pending requests. All your students may already be enrolled or none have preregistered yet.",
        )

    return render(request, "instructor/instructor_requests.html", context)



@require_POST
def instructor_request_action(request):
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")
    faculty = get_object_or_404(Faculty, email_id=email_id)

    sc_id = request.POST.get("sc_id")
    action = (request.POST.get("action") or "").lower()

    if not sc_id:
        messages.error(request, "No request selected.")
        return redirect("instructor_requests")

    sc = get_object_or_404(
        StudentCourse.objects.select_related("course", "student"),
        id=sc_id
    )

    # Ownership guard
    if not sc.course.faculties.filter(id=faculty.id).exists():
        messages.error(request, "Not authorized for this course.")
        return redirect("instructor_requests")

    # Only pending can change
    if sc.status != "PND":
        messages.info(request, "Request is not pending.")
        return redirect("instructor_requests")

    # Apply single action
    if action == "approve":
        sc.status = "ENR"
    elif action == "reject":
        sc.status = "DRP"
    else:
        messages.error(request, "Unknown action.")
        return redirect("instructor_requests")

    sc.save(update_fields=["status"])
    messages.success(request, f"Request {action}ed.")
    return redirect("instructor_requests")


@require_POST
def instructor_bulk_action(request):
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")
    faculty = get_object_or_404(Faculty, email_id=email_id)

    ids = request.POST.getlist("sc_ids")
    action = (request.POST.get("action") or "").lower()

    if not ids:
        messages.info(request, "No requests selected.")
        return redirect("instructor_requests")

    if action not in ("approve", "reject"):
        messages.error(request, "Unknown action.")
        return redirect("instructor_requests")

    qs = StudentCourse.objects.filter(id__in=ids).select_related("course", "student")
    updated = 0
    for sc in qs:
        if not sc.course.faculties.filter(id=faculty.id).exists():
            continue
        if sc.status != "PND":
            continue
        sc.status = "ENR" if action == "approve" else "DRP"
        sc.save(update_fields=["status"])
        updated += 1

    messages.success(request, f"{updated} request(s) {action}d.")
    return redirect("instructor_requests")


def instructor_courses(request):
    # Authn: get instructor
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    instructor = get_object_or_404(Faculty, email_id=email_id)


    # ✅ Show only current active pre-registration enrollments
    courses = (
        Course.objects.filter(faculties=instructor)
        .distinct()
        .annotate(
            total_enrolled=Count(
                "enrollments",
                filter=Q(enrollments__status="ENR", enrollments__is_active_pre_reg=True)
            )
        )
        .order_by("code")
    )

    context = {
        "instructor": instructor,
        "faculty": instructor,
        "courses": courses,
    }
    return render(request, "instructor/view_courses.html", context)


def course_roster(request, course_code):
    # --- Auth check ---
    email_id = request.session.get("email_id")
    if not email_id:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    instructor = get_object_or_404(Faculty, email_id=email_id)

    # --- Verify ownership of course ---
    course = get_object_or_404(Course.objects.prefetch_related("faculties"), code=course_code)
    if not course.faculties.filter(id=instructor.id).exists():
        messages.error(request, "You do not have permission to view this course.")
        return redirect("instructor_courses")

    q = (request.GET.get("q") or request.POST.get("q") or "").strip()

    # ==========================
    # Handle POST actions
    # ==========================
    if request.method == "POST":
        action = (request.POST.get("action") or "").lower()

        # --- Add student ---
        if action == "add_student":
            roll_no = (request.POST.get("roll_no") or "").strip().upper()

            if not roll_no:
                messages.error(request, "Roll number is required.")
                return redirect(f"{request.path}{'?q='+q if q else ''}")

            student = Student.objects.filter(roll_no__iexact=roll_no).first()
            if not student:
                messages.error(request, "Student not found.")
                return redirect(f"{request.path}{'?q='+q if q else ''}")

            # Prevent adding if student has already passed the course
            completed = StudentCourse.objects.filter(
                student=student, course=course
            ).filter(Q(outcome__iexact="PAS") | Q(status="CMP")).exists()
            if completed:
                messages.error(request, f"{roll_no} has already completed {course.code}.")
                return redirect(f"{request.path}{'?q='+q if q else ''}")

            with transaction.atomic():
                obj, created = StudentCourse.objects.get_or_create(
                    student=student,
                    course=course,
                    defaults={
                        "status": "ENR",
                        "is_active_pre_reg": True,
                    },
                )
                if created:
                    messages.success(request, f"Enrolled {student.roll_no} in {course.code}.")
                else:
                    if not obj.is_active_pre_reg:
                        obj.is_active_pre_reg = True
                        obj.status = "ENR"
                        obj.save(update_fields=["is_active_pre_reg", "status"])
                        messages.info(request, f"Reactivated {roll_no} in {course.code}.")
                    else:
                        messages.info(request, f"{roll_no} is already active in {course.code}.")
            return redirect(f"{request.path}{'?q='+q if q else ''}")

        # --- Remove student ---
        elif action == "remove_student":
            sc_id = request.POST.get("sc_id")
            try:
                sc_id_int = int(sc_id)
            except (TypeError, ValueError):
                sc_id_int = None

            if not sc_id_int:
                messages.error(request, "Invalid selection.")
            else:
                with transaction.atomic():
                    sc = StudentCourse.objects.filter(
                        id=sc_id_int, course=course, status="ENR", is_active_pre_reg=True
                    ).first()
                    if not sc:
                        messages.info(request, "No active enrollment found.")
                    else:
                        rid = sc.student.roll_no
                        sc.delete()
                        messages.success(request, f"Removed {rid} from {course.code}.")
            return redirect(f"{request.path}{'?q='+q if q else ''}")

        # --- Unknown action ---
        else:
            messages.error(request, "Unknown action.")
            return redirect(f"{request.path}{'?q='+q if q else ''}")

    # ==========================
    # GET: active enrollments
    # ==========================
    enrollments_qs = (
        StudentCourse.objects
        .filter(course__code=course_code, status="ENR", is_active_pre_reg=True)
        .select_related("student", "course")
        .order_by("student__roll_no")
    )

    if q:
        enrollments_qs = enrollments_qs.filter(
            Q(student__roll_no__icontains=q)
            | Q(student__first_name__icontains=q)
            | Q(student__last_name__icontains=q)
        )

    # Optional computed info
    enrollments = []
    for sc in enrollments_qs:
        sc.computed_semester = sc.student.calculate_current_semester()
        enrollments.append(sc)

    context = {
        "instructor": instructor,
        "faculty": instructor,
        "course": course,
        "enrollments": enrollments,
        "q": q,
    }
    return render(request, "instructor/course_roster.html", context)



def custom_admin_preregistration(request):
    return render(request, "admin/preregistration.html")


def admin_prereg_deadline(request):
    # Read current deadline for display
    current_deadline = getattr(settings, "PREREG_DEADLINE", None)
    current_local_iso = ""
    if current_deadline:
        aware = current_deadline
        if timezone.is_naive(aware):
            aware = timezone.make_aware(aware, timezone.get_current_timezone())
        local = timezone.localtime(aware, timezone.get_current_timezone())
        current_local_iso = local.strftime("%Y-%m-%dT%H:%M")

    # Read current temporary override (None => follow deadline)
    temp_state = getattr(settings, TEMP_PREREG_OPEN_FLAG, None)  # True/False/None

    if request.method == "POST":
        # Two possible forms submit to this same endpoint:
        # 1) action=open/close/clear for temporary override
        # 2) deadline=... for deadline updates
        action = (request.POST.get("action") or "").lower()

        if action in {"open", "close", "clear"}:
            if action == "open":
                settings.PREREG_TEMP_OPEN = True
                messages.success(request, "Pre‑registration temporarily set to OPEN.")  # [web:647]
            elif action == "close":
                settings.PREREG_TEMP_OPEN = False
                messages.warning(request, "Pre‑registration temporarily set to CLOSED.")  # [web:647]
            else:
                settings.PREREG_TEMP_OPEN = None
                messages.info(request, "Temporary override cleared. Deadline rule applies.")  # [web:647]
            return redirect("admin_prereg_deadline")

        # Otherwise treat as deadline update (may be empty to clear)
        raw = (request.POST.get("deadline") or "").strip()
        if raw == "":
            settings.PREREG_DEADLINE = None
            messages.info(request, "Deadline cleared. Window will follow temporary override or remain open if none.")  # [web:647]
            return redirect("admin_prereg_deadline")

        naive = parse_datetime(raw)  # "YYYY-MM-DDTHH:MM"
        if not naive:
            messages.error(request, "Invalid date/time format for deadline.")  # [web:647]
            return redirect("admin_prereg_deadline")

        aware = naive
        if timezone.is_naive(aware):
            aware = timezone.make_aware(aware, timezone.get_current_timezone())  # [web:607]
        settings.PREREG_DEADLINE = aware  # runtime assignment [web:635]
        local_disp = timezone.localtime(aware).strftime("%Y-%m-%d %H:%M %Z")
        messages.success(request, f"Deadline set to {local_disp}.")  # [web:647]
        return redirect("admin_prereg_deadline")

    ctx = {
        "current_local_iso": current_local_iso,
        "temp_state": temp_state,  # True / False / None
    }
    return render(request, "admin/prereg_deadline.html", ctx)

def admin_prereg_enrollments(request):
    q = (request.GET.get("q") or "").strip()
    code = (request.GET.get("course") or "").strip().upper()
    slot = (request.GET.get("slot") or "").strip().upper()

    # Filter and search available courses
    courses_qs = Course.objects.all().order_by("slot", "code")
    if q:
        courses_qs = courses_qs.filter(Q(code__icontains=q) | Q(name__icontains=q))
    if slot:
        courses_qs = courses_qs.filter(slot=slot)

    courses = list(courses_qs)
    selected_course = None
    enrolled = []
    current_sem = None

    # Active course selected
    if code:
        selected_course = Course.objects.filter(code__iexact=code).first()
        if selected_course:
            sem_param = request.GET.get("sem")
            base_q = {
                "course": selected_course,
                "is_active_pre_reg": True,  # ✅ show only current-cycle records
                "status__in": ["ENR", "PND"],  # only active statuses
            }
            if sem_param and sem_param.isdigit():
                current_sem = int(sem_param)
                base_q["semester"] = current_sem

            enrolled = (
                StudentCourse.objects.filter(**base_q)
                .select_related("student")
                .order_by("semester", "student__roll_no")
            )

    # Redirect helper
    def redirect_with(extra):
        params = {"q": q, "course": code, "slot": slot}
        params.update(extra)
        clean = {k: v for k, v in params.items() if v}
        return redirect(f"{request.path}?{urlencode(clean)}")

    # ===================== POST SECTION =====================
    if request.method == "POST":
        action = (request.POST.get("action") or "").lower()
        target_code = (request.POST.get("course_code") or "").strip().upper()
        roll_no = (request.POST.get("roll_no") or "").strip().upper()
        sem_raw = (request.POST.get("semester") or "").strip()
        sem = int(sem_raw) if sem_raw.isdigit() else None

        course = Course.objects.filter(code__iexact=target_code).first()
        student = Student.objects.filter(roll_no__iexact=roll_no).first()

        if not (course and student):
            messages.error(request, "Invalid course or student.")
            return redirect_with({})

        # Auto-semester fallback
        if sem is None:
            try:
                sem = student.calculate_current_semester()
            except Exception:
                sem = 1

        # 🧠 prevent re-enrolling in already PAS/CMP courses
        if action == "add":
            completed = StudentCourse.objects.filter(
                student=student,
                course=course
            ).filter(Q(outcome__iexact="PAS") | Q(status="CMP")).exists()
            if completed:
                messages.error(request, f"{roll_no} has already completed {target_code}.")
                return redirect_with({})

            # Prevent duplicate PND record
            if StudentCourse.objects.filter(
                student=student,
                course=course,
                is_active_pre_reg=True,
                status="PND"
            ).exists():
                messages.info(request, f"{roll_no} already has a pending pre-registration for {target_code}.")
                return redirect_with({})

        # ✅ Safe atomic operations
        with transaction.atomic():
            if action == "remove":
                deleted, _ = StudentCourse.objects.filter(
                    student=student,
                    course=course,
                    semester=sem,
                    is_active_pre_reg=True,
                    status__in=["ENR", "PND"]
                ).delete()
                if deleted:
                    messages.success(request, f"Removed {roll_no} from {target_code} (Sem {sem}).")
                else:
                    messages.info(request, "No record found to remove.")
                return redirect_with({})

            if action == "add":
                # Add as ENR (admin manually enrolls, stays active)
                obj, created = StudentCourse.objects.get_or_create(
                    student=student,
                    course=course,
                    semester=sem,
                    defaults={
                        "status": "ENR",
                        "course_mode": "REG",
                        "is_active_pre_reg": True
                    }
                )
                if not created:
                    # Reactivate if found from previous resets
                    obj.is_active_pre_reg = True
                    obj.status = "ENR"
                    obj.save(update_fields=["is_active_pre_reg", "status"])
                    messages.info(request, f"Reactivated {roll_no} in {target_code} (Sem {sem}).")
                else:
                    messages.success(request, f"Added {roll_no} to {target_code} (Sem {sem}).")
                return redirect_with({})

    # ===================== CONTEXT =====================
    context = {
        "q": q,
        "courses": courses,
        "selected_course": selected_course,
        "enrolled": enrolled,
        "current_sem": current_sem,
        "slot": slot,
    }
    return render(request, "admin/prereg_enrollments.html", context)



def admin_prereg_swap(request):
    q = (request.GET.get("q") or "").strip()
    slot = (request.GET.get("slot") or "").strip().upper()

    courses_qs = Course.objects.all().order_by("slot", "code")
    if q:
        courses_qs = courses_qs.filter(Q(code__icontains=q) | Q(name__icontains=q))
    if slot:
        courses_qs = courses_qs.filter(slot=slot)

    courses = list(courses_qs)

    context = {
        "courses": courses,
        "q": q,
        "slot": slot,
    }

    if request.method != "POST":
        return render(request, "admin/prereg_swap.html", context)

    # POST: perform swap
    roll_no = (request.POST.get("roll_no") or "").strip().upper()
    from_code = (request.POST.get("from_code") or "").strip().upper()
    to_code = (request.POST.get("to_code") or "").strip().upper()
    sem_raw = (request.POST.get("semester") or "").strip()

    if not roll_no or not from_code or not to_code:
        messages.error(request, "Roll no, From course and To course are required.")
        return redirect("admin_prereg_swap")

    if from_code == to_code:
        messages.info(request, "From and To courses are identical; nothing to swap.")
        return redirect("admin_prereg_swap")

    try:
        sem = int(sem_raw) if sem_raw else None
    except ValueError:
        sem = None

    student = Student.objects.filter(roll_no__iexact=roll_no).first()
    if not student:
        messages.error(request, "Student not found.")
        return redirect("admin_prereg_swap")

    from_course = Course.objects.filter(code__iexact=from_code).first()
    to_course   = Course.objects.filter(code__iexact=to_code).first()
    if not from_course or not to_course:
        messages.error(request, "Invalid course code(s).")
        return redirect("admin_prereg_swap")

    if from_course.slot != to_course.slot:
        messages.error(request, f"Courses must be in the same slot (got {from_course.slot} vs {to_course.slot}).")
        return redirect("admin_prereg_swap")

    if sem is None:
        try:
            sem = student.calculate_current_semester()
        except Exception:
            sem = 1

    # Enforce: 'from' must be a pre-registered entry (PND) in the current active prereg cycle.
    enrolled_from = StudentCourse.objects.filter(
        student=student, course=from_course, semester=sem, status="PND", is_active_pre_reg=True
    ).first()

    if not enrolled_from:
        messages.error(request, f"Swap denied: student is not pre-registered (PND) in {from_code} for semester {sem}.")
        return redirect("admin_prereg_swap")

    # Prevent swapping into a course already passed
    if StudentCourse.objects.filter(student=student, course=to_course, outcome__iexact="PAS").exists():
        messages.error(request, f"Swap denied: student has already passed {to_code}.")
        return redirect("admin_prereg_swap")

    with transaction.atomic():
        # If student already has a PND or ENR in to_course in same cycle, handle accordingly:
        enrolled_to = StudentCourse.objects.filter(
            student=student, course=to_course, semester=sem, is_active_pre_reg=True
        ).first()

        # Remove the from_course pre-reg (PND)
        enrolled_from.delete()

        if enrolled_to:
            # If there's already a to_course entry, we keep it (no duplicate) and ensure it remains PND (or ENR if desired)
            # If it was PND, leave it as PND. If it was ENR, keep as ENR.
            messages.success(request, f"Swapped {roll_no}: removed {from_code}, kept existing {to_code} (Sem {sem}).")
            return redirect("admin_prereg_swap")

        # Otherwise create a new PRE-REG (PND) entry for to_course (do NOT ENROLL directly)
        to_obj, created = StudentCourse.objects.get_or_create(
            student=student, course=to_course, semester=sem,
            defaults={
                "status": "PND",              # create as pending pre-registration
                "course_mode": "REG",
                "is_active_pre_reg": True
            }
        )
        if not created:
            # If a record existed but wasn't active in this cycle, update to PND + active
            updated = False
            if not to_obj.is_active_pre_reg:
                to_obj.is_active_pre_reg = True
                updated = True
            if to_obj.status != "PND":
                to_obj.status = "PND"
                updated = True
            if updated:
                to_obj.save(update_fields=["status", "is_active_pre_reg"])
        messages.success(request, f"Swapped {roll_no} from {from_code} to {to_code} (Sem {sem}) — created pre-registration (PND).")
        return redirect("admin_prereg_swap")
    

def student_registered_courses(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    student = get_object_or_404(Student, roll_no=roll_no)

    # Optional filters
    sem = (request.GET.get("sem") or "").strip()
    q = (request.GET.get("q") or "").strip()

    sc_qs = (
        StudentCourse.objects
        .filter(student=student, status="ENR")
        .select_related("course")
        .prefetch_related(Prefetch("course__faculties"))
        .order_by("semester", "course__slot", "course__code")
    )

    if sem.isdigit():
        sc_qs = sc_qs.filter(semester=int(sem))

    if q:
        sc_qs = sc_qs.filter(
            Q(course__code__icontains=q) |
            Q(course__name__icontains=q) |
            Q(course__slot__icontains=q)
        )

    # Build a small data list for the template
    items = []
    for sc in sc_qs:
        c = sc.course
        facs = getattr(c, "faculties", None)
        fac_names = []
        if facs:
            for f in facs.all():
                parts = [p for p in [f.first_name, f.last_name] if p]
                if parts:
                    fac_names.append(" ".join(parts))
        items.append({
            "semester": sc.semester,
            "code": c.code,
            "name": c.name,
            "slot": c.slot,
            "credits": c.credits,
            "ltpc": getattr(c, "LTPC", ""),
            "instructors": ", ".join(fac_names) if fac_names else "—",
        })

    context = {
        "student": student,
        "items": items,
        "q": q,
        "sem": sem,
    }
    return render(request, "registration/registered_course.html", context)


def admin_reset_pre_registration(request):
    # Mark all current pre-registrations inactive
    StudentCourse.objects.filter(is_active_pre_reg=True).update(is_active_pre_reg=False)

    # Optional (recommended for clarity): mark unfinalized requests as archived
    StudentCourse.objects.filter(status__in=["PND", "INS"]).update(status="ARC")

    messages.success(request, "Pre-registration has been reset for all students. A new cycle has started.")
    return redirect("admin_prereg_deadline")

def admin_prereg_reports(request):
    q = (request.GET.get("q") or "").strip()
    slot = (request.GET.get("slot") or "").strip().upper()

    courses = Course.objects.all().order_by("slot", "code")
    if q:
        courses = courses.filter(Q(code__icontains=q) | Q(name__icontains=q))
    if slot:
        courses = courses.filter(slot=slot)

    # ✅ Filter only current active pre-registration data
    active_qs = StudentCourse.objects.filter(
        is_active_pre_reg=True,
        status__in=["ENR", "PND"]
    )

    # Aggregate counts
    enr_counts = (
        active_qs.filter(status="ENR")
        .values("course_id").annotate(count=Count("id"))
    )
    pnd_counts = (
        active_qs.filter(status="PND")
        .values("course_id").annotate(count=Count("id"))
    )
    enr_map = {x["course_id"]: x["count"] for x in enr_counts}
    pnd_map = {x["course_id"]: x["count"] for x in pnd_counts}

    # Combine total active counts
    total_map = {cid: enr_map.get(cid, 0) + pnd_map.get(cid, 0) for cid in set(enr_map) | set(pnd_map)}

    context = {
        "courses": courses,
        "q": q,
        "slot": slot,
        "enr_map": enr_map,
        "pnd_map": pnd_map,
        "total_map": total_map,
    }
    return render(request, "admin/prereg_reports.html", context)


def export_course_excel(request, code):
    course = Course.objects.filter(code__iexact=code).first()
    if not course:
        raise Http404("Course not found")

    # ✅ Only export current active-cycle enrollments
    rows = (
        StudentCourse.objects
        .filter(course=course, status="ENR", is_active_pre_reg=True)
        .select_related("student")
        .order_by("student__roll_no", "semester")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Active ENR Students"
    ws.append(["Roll No", "Name", "Semester", "Status"])

    for sc in rows:
        s = sc.student
        full_name = f"{s.first_name or ''} {s.last_name or ''}".strip() or "—"
        ws.append([s.roll_no, full_name, sc.semester, sc.status])

    fname = f"{slugify(course.code)}-active-enrolled.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


def export_course_pdf(request, code):
    course = Course.objects.filter(code__iexact=code).first()
    if not course:
        raise Http404("Course not found")

    # ✅ Only active pre-registration ENR records
    rows = (
        StudentCourse.objects
        .filter(course=course, status="ENR", is_active_pre_reg=True)
        .select_related("student")
        .order_by("semester", "student__roll_no")
    )

    html = render_to_string(
        "admin/prereg_course_pdf.html",
        {"course": course, "rows": rows}
    )
    pdf_bytes = HTML(string=html).write_pdf()
    fname = f"{slugify(course.code)}-active-enrolled.pdf"

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

# views.py (faculty, require faculty session)

def require_faculty(view):
    def _w(request, *a, **k):
        if not request.session.get("email_id"):
            return redirect(f"{getattr(settings,'LOGIN_URL','/login/')}?next={request.path}")
        return view(request, *a, **k)
    return _w

def instructor_schema_courses(request):
    email_id = request.session["email_id"]
    instructor = get_object_or_404(Faculty, email_id=email_id)
    courses = (
        Course.objects.filter(faculties=instructor)
        .annotate(
            total_enrolled=Count(
                "enrollments",
                filter=Q(enrollments__status="ENR", enrollments__is_active_pre_reg=True)
            )
        )
        .order_by("code")
        .distinct()
    )
    return render(request, "instructor/schema_courses.html", {
        "instructor": instructor,
        "courses": courses,
        "faculty":instructor,
    })


# ----------------- Edit assessment scheme (no semester on component) -----------------

def edit_assessment_scheme(request, course_code):
    email = request.session["email_id"]
    faculty = get_object_or_404(Faculty, email_id=email)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)

    ComponentFormSet = modelformset_factory(
        AssessmentComponent,
        fields=["name", "weight", "max_marks"],
        extra=0, can_delete=True
    )
    qs = AssessmentComponent.objects.filter(course=course).order_by("id")

    if request.method == "POST":
        # Add-only branch
        if request.POST.get("intent") == "add":
            add_name = (request.POST.get("add_name") or "").strip()
            add_weight = request.POST.get("add_weight")
            add_max = request.POST.get("add_max") or 100
            if not add_name or add_weight is None:
                messages.error(request, "Name and weight are required.")
                return redirect(request.path)
            AssessmentComponent.objects.update_or_create(
                course=course, name=add_name,
                defaults={"weight": add_weight, "max_marks": add_max},
            )
            messages.success(request, "Component added.")
            return redirect(request.path)

        # Save edits/deletes
        formset = ComponentFormSet(request.POST, queryset=qs)
        if formset.is_valid():
            instances = formset.save(commit=False)

            # Delete marked-for-deletion
            for f in formset.deleted_forms:
                if f.instance.pk:
                    f.instance.delete()

            # Save updates/creates, attach course
            for inst in instances:
                inst.course = course
                inst.save()

            total_w = AssessmentComponent.objects.filter(course=course).aggregate(s=Sum("weight"))["s"] or 0
            if round(float(total_w), 2) != 100.0:
                messages.warning(request, f"Total weight is {total_w}, should be 100.")
            else:
                messages.success(request, "Assessment scheme saved.")
            return redirect(request.path)
        else:
            messages.error(request, "Please correct the errors.")
    else:
        formset = ComponentFormSet(queryset=qs)

    return render(request, "instructor/edit_assessment_schema.html", {
        "course": course,
        "formset": formset,
        "faculty":faculty,
        "sample_headers": ["roll_no", "email"] + [c.name for c in qs],
    })

from decimal import Decimal, InvalidOperation

def _canon(s: str) -> str:
    return re.sub(r'[^a-z0-9]+','', (s or "").lower())

from django.views.decorators.http import require_POST

# ----------------- Update/delete single component -----------------

from django.views.decorators.http import require_POST

@require_POST
def update_component(request, course_code, component_id):
    email = request.session["email_id"]
    faculty = get_object_or_404(Faculty, email_id=email)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)
    comp = get_object_or_404(AssessmentComponent, id=component_id, course=course)
    name = (request.POST.get("name") or "").strip()
    weight = request.POST.get("weight")
    max_marks = request.POST.get("max_marks")
    if not name:
        messages.error(request, "Name is required.")
    else:
        comp.name = name
        if weight is not None:
            comp.weight = weight or 0
        if max_marks is not None:
            comp.max_marks = max_marks or 100
        comp.save()
        messages.success(request, "Component updated.")
    return redirect(reverse('edit_assessment_scheme', args=[course.code]))


@require_POST
def delete_component(request, course_code, component_id):
    email = request.session["email_id"]
    faculty = get_object_or_404(Faculty, email_id=email)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)
    comp = get_object_or_404(AssessmentComponent, id=component_id, course=course)
    comp.delete()
    messages.success(request, "Component removed.")
    return redirect(reverse('edit_assessment_scheme', args=[course.code]))
# ----------------- Faculty course cards + enter marks -----------------

def faculty_marks_courses(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)
    courses = (
        Course.objects
        .filter(
            faculties=faculty,
            enrollments__is_active_pre_reg=True,
            enrollments__status="ENR"
        )
        .order_by('code', 'name')
        .distinct()
    )
    return render(
        request,
        'instructor/marks_course_list.html',
        {'faculty': faculty, 'courses': courses}
    )


def mark_field_name(student_id, comp_id):
    return f"mark_{student_id}_{comp_id}"

def enter_marks(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Identify Faculty (for back button and ownership)
    faculty = None
    fid = request.GET.get('faculty_id')
    if fid:
        faculty = Faculty.objects.filter(id=fid).first()
    if faculty is None:
        faculty = course.faculties.first()

    # --- Fetch Assessment Components ---
    components = list(
        AssessmentComponent.objects.filter(course=course).order_by('id')
    )

    # ✅ Only current active enrollments (after pre-reg reset)
    enrollments_qs = (
        StudentCourse.objects.filter(
            course=course,
            status='ENR',
            is_active_pre_reg=True  # <-- ensures only active-cycle students are shown
        )
        .select_related('student')
        .order_by('student__roll_no')
    )
    enrollments = list(enrollments_qs)
    students = [e.student for e in enrollments]

    # --- Existing marks map (for prefill) ---
    existing = {}
    if students and components:
        for s in AssessmentScore.objects.filter(
            course=course,
            student__in=students,
            component__in=components
        ):
            existing[(s.student_id, s.component_id)] = s

    # --- Prefill baseline from DB ---
    prefill = {
        mark_field_name(stu_id, comp_id): f"{s.marks_obtained}"
        for (stu_id, comp_id), s in existing.items()
    }

    if request.method == 'POST':
        cell_errors, to_create, to_update = [], [], []
        is_upload_all = 'upload_all' in request.POST
        is_upload_component = 'upload_component' in request.POST

        import pandas as pd

        # ---------------- BULK UPLOAD (one file with all components) ----------------
        if is_upload_all:
            file = request.FILES.get('bulk_file')
            if not file:
                cell_errors.append(("File Error", "-", "No file provided"))
            else:
                try:
                    df = pd.read_excel(file) if file.name.endswith(('.xls', '.xlsx')) else pd.read_csv(file)
                    df.columns = [c.strip().lower() for c in df.columns]
                except Exception as ex:
                    cell_errors.append(("File Error", "-", f"Failed to parse file: {ex}"))
                    df = None

                if df is not None:
                    if 'roll_no' not in df.columns:
                        cell_errors.append(("File Error", "-", "Missing 'roll_no' column"))
                    else:
                        by_roll = {e.student.roll_no.strip().lower(): e.student for e in enrollments}
                        for _, row in df.iterrows():
                            roll_key = str(row['roll_no']).strip().lower()
                            student_obj = by_roll.get(roll_key)
                            if not student_obj:
                                cell_errors.append((row.get('roll_no', ''), "N/A", "Not enrolled (inactive or dropped)"))
                                continue
                            for comp in components:
                                cname = comp.name.strip().lower()
                                if cname not in df.columns:
                                    continue
                                val = row[cname]
                                if pd.isna(val):
                                    continue
                                try:
                                    val = float(val)
                                except (ValueError, TypeError):
                                    cell_errors.append((row.get('roll_no', ''), comp.name, "Invalid number"))
                                    continue
                                if val < 0 or val > comp.max_marks:
                                    cell_errors.append((row.get('roll_no', ''), comp.name, f"Must be 0–{comp.max_marks}"))
                                    continue
                                obj = existing.get((student_obj.id, comp.id))
                                if obj:
                                    obj.marks_obtained = val
                                    to_update.append(obj)
                                else:
                                    to_create.append(AssessmentScore(
                                        student=student_obj, course=course, component=comp, marks_obtained=val
                                    ))

            if to_create:
                AssessmentScore.objects.bulk_create(to_create, batch_size=500)
                for obj in to_create:
                    existing[(obj.student_id, obj.component_id)] = obj
            if to_update:
                AssessmentScore.objects.bulk_update(to_update, ['marks_obtained'], batch_size=500)

            prefill = {
                mark_field_name(stu_id, comp_id): f"{s.marks_obtained}"
                for (stu_id, comp_id), s in existing.items()
            }

            if cell_errors:
                return render(request, 'instructor/enter_marks.html', {
                    'course': course, 'components': components, 'enrollments': enrollments,
                    'prefill': prefill, 'cell_errors': cell_errors, 'faculty': faculty
                })
            url = reverse('enter_marks', kwargs={'course_id': course.id})
            if faculty:
                url = f"{url}?faculty_id={faculty.id}"
            return redirect(url)

        # ---------------- PER COMPONENT UPLOAD ----------------
        elif is_upload_component:
            for comp in components:
                file = request.FILES.get(f'component_csv_{comp.id}')
                if not file:
                    continue
                try:
                    df = pd.read_excel(file) if file.name.endswith(('.xls', '.xlsx')) else pd.read_csv(file)
                    df.columns = [c.strip().lower() for c in df.columns]
                except Exception as ex:
                    cell_errors.append((comp.name, "-", f"Failed to parse file: {ex}"))
                    continue
                if 'roll_no' not in df.columns or 'marks' not in df.columns:
                    cell_errors.append((comp.name, "-", "Missing roll_no or marks column"))
                    continue

                by_roll = {e.student.roll_no.strip().lower(): e.student for e in enrollments}
                for _, row in df.iterrows():
                    roll_key = str(row['roll_no']).strip().lower()
                    val = row['marks']
                    student_obj = by_roll.get(roll_key)
                    if not student_obj:
                        cell_errors.append((row.get('roll_no', ''), comp.name, "Not enrolled (inactive or dropped)"))
                        continue
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        cell_errors.append((row.get('roll_no', ''), comp.name, "Invalid number"))
                        continue
                    if val < 0 or val > comp.max_marks:
                        cell_errors.append((row.get('roll_no', ''), comp.name, f"Must be 0–{comp.max_marks}"))
                        continue
                    obj = existing.get((student_obj.id, comp.id))
                    if obj:
                        obj.marks_obtained = val
                        to_update.append(obj)
                    else:
                        to_create.append(AssessmentScore(
                            student=student_obj, course=course, component=comp, marks_obtained=val
                        ))

            if to_create:
                AssessmentScore.objects.bulk_create(to_create, batch_size=500)
                for obj in to_create:
                    existing[(obj.student_id, obj.component_id)] = obj
            if to_update:
                AssessmentScore.objects.bulk_update(to_update, ['marks_obtained'], batch_size=500)

            prefill = {
                mark_field_name(stu_id, comp_id): f"{s.marks_obtained}"
                for (stu_id, comp_id), s in existing.items()
            }

            if cell_errors:
                return render(request, 'instructor/enter_marks.html', {
                    'course': course, 'components': components, 'enrollments': enrollments,
                    'prefill': prefill, 'cell_errors': cell_errors, 'faculty': faculty
                })
            url = reverse('enter_marks', kwargs={'course_id': course.id})
            if faculty:
                url = f"{url}?faculty_id={faculty.id}"
            return redirect(url)

        # ---------------- MANUAL ENTRY GRID ----------------
        else:
            for e in enrollments:
                for comp in components:
                    field = mark_field_name(e.student_id, comp.id)
                    raw = (request.POST.get(field) or '').strip()
                    if raw == '':
                        continue
                    try:
                        val = float(raw)
                    except ValueError:
                        cell_errors.append((e.student.roll_no, comp.name, "Invalid number"))
                        continue
                    if val < 0 or val > comp.max_marks:
                        cell_errors.append((e.student.roll_no, comp.name, f"Must be 0–{comp.max_marks}"))
                        continue
                    obj = existing.get((e.student_id, comp.id))
                    if obj:
                        obj.marks_obtained = val
                        to_update.append(obj)
                    else:
                        to_create.append(AssessmentScore(
                            student=e.student, course=course, component=comp, marks_obtained=val
                        ))

            if to_create:
                AssessmentScore.objects.bulk_create(to_create, batch_size=500)
                for obj in to_create:
                    existing[(obj.student_id, obj.component_id)] = obj
            if to_update:
                AssessmentScore.objects.bulk_update(to_update, ['marks_obtained'], batch_size=500)

            # Prefill again, keeping any manual entries for errors
            prefill = {
                mark_field_name(stu_id, comp_id): f"{s.marks_obtained}"
                for (stu_id, comp_id), s in existing.items()
            }
            for e in enrollments:
                for comp in components:
                    key = mark_field_name(e.student_id, comp.id)
                    if key in request.POST:
                        prefill[key] = request.POST.get(key, '')

            if cell_errors:
                return render(request, 'instructor/enter_marks.html', {
                    'course': course, 'components': components, 'enrollments': enrollments,
                    'prefill': prefill, 'cell_errors': cell_errors, 'faculty': faculty
                })

            url = reverse('enter_marks', kwargs={'course_id': course.id})
            if faculty:
                url = f"{url}?faculty_id={faculty.id}"
            return redirect(url)

    # --- GET: show marks for current-cycle students only ---
    return render(request, 'instructor/enter_marks.html', {
        'course': course,
        'components': components,
        'enrollments': enrollments,
        'prefill': prefill,
        'cell_errors': [],
        'faculty': faculty,
    })



def course_marks_overview(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    email = request.session.get("email_id")
    if not email:
        login_url = getattr(settings, "LOGIN_URL", "/login/")
        return redirect(f"{login_url}?next={request.path}")

    faculty = get_object_or_404(Faculty, email_id=email)

    # Fetch grading policy (if any)
    policy = GradingPolicy.objects.filter(course=course).first()

    components = list(AssessmentComponent.objects.filter(course=course).order_by('id'))

    # Only current active-cycle enrolled students
    enrollments_qs = (
        StudentCourse.objects
        .filter(course=course, status__in=["ENR", "CMP"], is_active_pre_reg=True)
        .select_related('student')
        .order_by('student__roll_no')
    )
    enrollments = list(enrollments_qs)
    students = [e.student for e in enrollments]

    # Scores lookup
    scores = {}
    if students and components:
        for s in AssessmentScore.objects.filter(
            course=course, student__in=students, component__in=components
        ):
            try:
                scores[(s.student_id, s.component_id)] = Decimal(str(s.marks_obtained))
            except (InvalidOperation, TypeError):
                scores[(s.student_id, s.component_id)] = None

    # Component metadata and weights
    comp_info, total_weight = [], Decimal("0")
    for c in components:
        max_d = Decimal(str(c.max_marks))
        w = getattr(c, "weight", None)
        weight_d = Decimal(str(w)) if w is not None else Decimal("1")
        comp_info.append((c, max_d, weight_d))
        total_weight += weight_d

    # Build student rows (attach enrollment to each row for later saving)
    rows, totals_by_student = [], {}
    for e in enrollments:
        stu = e.student
        pairs, weighted_total = [], Decimal("0")
        for comp, max_d, weight_d in comp_info:
            val = scores.get((stu.id, comp.id))
            if val is not None and max_d > 0:
                contrib = (val / max_d) * weight_d
                weighted_total += contrib
            pairs.append({
                "component": comp,
                "value": val,
                "max": comp.max_marks,
                "component_id": comp.id,
            })

        totals_by_student[stu.id] = weighted_total
        percentage = (weighted_total / total_weight * Decimal("100")) if total_weight > 0 else None
        rows.append({
            "student": stu,
            "pairs": pairs,
            "total": weighted_total,
            "percentage": percentage,
            "enrollment": e,   # keep reference for saving
        })

    # Percentile calculation (for relative grading)
    all_totals = [totals_by_student[e.student.id] for e in enrollments]
    n = len(all_totals)
    sorted_totals = sorted(all_totals)

    def percentile_rank(total_value):
        if n <= 1:
            return Decimal("100")
        lo, hi = 0, n
        while lo < hi:
            mid = (lo + hi) // 2
            if sorted_totals[mid] < total_value:
                lo = mid + 1
            else:
                hi = mid
        lower = lo
        # denominator (n - 1) might be zero-guarded above
        return (Decimal(lower) / Decimal(n - 1)) * Decimal("100")

    for r in rows:
        r["percentile"] = percentile_rank(r["total"]) if r["total"] is not None else None

    # Default grade mapping (if you use grade points elsewhere)
    grade_letters = {10: "A", 9: "A-", 8: "B", 7: "B-", 6: "C", 5: "C-", 4: "D", 0: "F"}

    # Default absolute cutoffs (percent). You can override via policy.abs_cutoffs
    default_abs = {"A": 85, "A-": 75, "B": 65, "B-": 55, "C": 45, "D": 40, "F": 0}

    # Apply grading policy and persist to StudentCourse (only if changed)
    rows_to_save = []
    for r in rows:
        pct = (r["percentage"] if r["percentage"] is not None else Decimal("0"))
        perc_percentile = (r["percentile"] if r["percentile"] is not None else Decimal("0"))
        enrollment = r["enrollment"]
        course_mode = (enrollment.course_mode or "").upper()

        grade = None
        outcome = None

        # Pass/Fail mode: we set outcome based on pass cutoff, but don't assign letter grade
        if course_mode == "PF":
            # Determine pass threshold (use policy.abs_cutoffs.Pass if provided else default C threshold)
            abs_cutoffs = (policy.abs_cutoffs if policy else {}) or {}
            pass_threshold = Decimal(str(abs_cutoffs.get("Pass", abs_cutoffs.get("C", default_abs["C"]))))
            outcome = "PAS" if pct >= pass_threshold else "FAI"
            grade = None

        # Audit mode: normally auditable -> pass, keep grade None and outcome PAS
        elif course_mode == "AUD":
            grade = None
            outcome = "PAS"

        else:
            # If policy exists and is REL, use percentile buckets
            if policy and getattr(policy, "mode", "").upper() == "REL":
                rel = policy.rel_buckets or {}
                # safe parsing of bucket numbers; fallback to common defaults
                try:
                    top10 = Decimal(str(rel.get("top10", 10)))
                    next15 = Decimal(str(rel.get("next15", 15)))
                except (InvalidOperation, TypeError):
                    top10 = Decimal("10")
                    next15 = Decimal("15")

                p = perc_percentile
                # Example bucket logic: top10, next15, next25, ...
                if p <= top10:
                    grade = "A"
                elif p <= (top10 + next15):
                    grade = "A-"
                elif p <= Decimal("50"):
                    grade = "B"
                elif p <= Decimal("75"):
                    grade = "B-"
                elif p <= Decimal("90"):
                    grade = "C"
                else:
                    grade = "D"
                outcome = "PAS" if grade != "F" else "FAI"

            else:
                # Absolute grading - compare percentage to cutoffs
                abs_cutoffs = (policy.abs_cutoffs if policy else {}) or default_abs
                # convert to Decimal and provide fallback values
                def get_cutoff(k, fallback):
                    try:
                        return Decimal(str(abs_cutoffs.get(k, fallback)))
                    except (InvalidOperation, TypeError):
                        return Decimal(str(fallback))

                A_th = get_cutoff("A", default_abs["A"])
                Am_th = get_cutoff("A-", default_abs["A-"])
                B_th = get_cutoff("B", default_abs["B"])
                Bm_th = get_cutoff("B-", default_abs["B-"])
                C_th = get_cutoff("C", default_abs["C"])
                D_th = get_cutoff("D", default_abs["D"])
                # grading descending
                if pct >= A_th:
                    grade = "A"
                elif pct >= Am_th:
                    grade = "A-"
                elif pct >= B_th:
                    grade = "B"
                elif pct >= Bm_th:
                    grade = "B-"
                elif pct >= C_th:
                    grade = "C"
                elif pct >= D_th:
                    grade = "D"
                else:
                    grade = "F"
                outcome = "PAS" if grade != "F" else "FAI"

        # Store into row for template rendering
        r["grade"] = grade
        r["outcome"] = outcome

        # Prepare DB update - update StudentCourse only if values changed
        changed = False
        # grade field stored should be None or a string matching choices; allow None
        if (enrollment.grade != (grade if grade is not None else None)):
            enrollment.grade = grade
            changed = True
        # enrollment.outcome stored as "PAS"/"FAI"
        if (enrollment.outcome != (outcome if outcome is not None else None)):
            enrollment.outcome = outcome
            changed = True
        # If we assigned a letter grade, ensure status is CMP
        if grade and enrollment.status != "CMP":
            enrollment.status = "CMP"
            changed = True

        if changed:
            rows_to_save.append(enrollment)

    # Bulk-save changed enrollments (small list, do one-by-one to respect model signals/constraints)
    if rows_to_save:
        with transaction.atomic():
            for enr in rows_to_save:
                # Only update relevant fields
                enr.save(update_fields=["grade", "outcome", "status"])

    max_total_display = total_weight

    return render(request, "instructor/course_marks_overview.html", {
        "course": course,
        "faculty": faculty,
        "components": components,
        "rows": rows,
        "max_total": max_total_display,
        "policy": policy,
    })


from decimal import Decimal
from django.db import transaction

def apply_grading_policy(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    policy = GradingPolicy.objects.filter(course=course).first()

    if not policy:
        messages.error(request, "No grading policy set for this course.")
        return redirect("course_marks_overview", course_id=course.id)

    # Fetch active enrollments (CMP + ENR)
    enrollments = (
        StudentCourse.objects.filter(
            course=course,
            status__in=["ENR", "CMP"],
            is_active_pre_reg=True
        )
        .select_related("student")
        .order_by("student__roll_no")
    )

    components = list(AssessmentComponent.objects.filter(course=course))

    # Score lookup
    score_map = {}
    for s in AssessmentScore.objects.filter(
        course=course,
        student__in=[e.student for e in enrollments],
        component__in=components
    ):
        score_map[(s.student_id, s.component_id)] = Decimal(str(s.marks_obtained))

    # Component metadata
    comp_data = []
    total_weight = Decimal("0")
    for c in components:
        max_d = Decimal(str(c.max_marks))
        w = Decimal(str(getattr(c, "weight", 1)))
        comp_data.append((c, max_d, w))
        total_weight += w

    # Compute totals
    totals = {}
    for e in enrollments:
        stu_total = Decimal("0")
        for comp, max_d, w in comp_data:
            val = score_map.get((e.student.id, comp.id))
            if val is not None and max_d > 0:
                stu_total += (val / max_d) * w
        totals[e.student.id] = stu_total

    # Percentile calculation
    total_list = list(totals.values())
    sorted_totals = sorted(total_list)
    n = len(sorted_totals)

    def percentile_rank(value):
        if n <= 1:
            return Decimal("100")
        lower = sum(1 for x in sorted_totals if x < value)
        return Decimal(lower) / Decimal(n - 1) * Decimal("100")

    # --- APPLY GRADING POLICY ---
    with transaction.atomic():
        for e in enrollments:
            pct = (totals[e.student.id] / total_weight * Decimal("100")) if total_weight else Decimal("0")
            grade = None
            outcome = None

            # --- RELATIVE ---
            if policy.mode == "REL":
                per = percentile_rank(totals[e.student.id])
                rb = policy.rel_buckets

                if per <= rb.get("top10", 10): grade = "A"
                elif per <= rb.get("top10", 10) + rb.get("next15", 15): grade = "A-"
                elif per <= 50: grade = "B"
                elif per <= 75: grade = "B-"
                elif per <= 90: grade = "C"
                else: grade = "D"

                outcome = "PAS" if grade != "F" else "FAI"

            # --- ABSOLUTE ---
            else:
                c = policy.abs_cutoffs
                if pct >= c["A"]: grade = "A"
                elif pct >= c["A-"]: grade = "A-"
                elif pct >= c["B"]: grade = "B"
                elif pct >= c["B-"]: grade = "B-"
                elif pct >= c["C"]: grade = "C"
                elif pct >= c["Pass"]: grade = "D"
                else: grade = "F"
                outcome = "PAS" if grade != "F" else "FAI"

            # Save results
            e.grade = grade
            e.outcome = outcome
            e.status = "CMP"  # mark as completed
            e.save()

    messages.success(request, "Grades updated successfully using current grading policy.")
    return redirect("course_marks_overview", course_id=course.id)



@require_POST
def update_mark_cell(request, course_id):
    try:
        student_id = int(request.POST.get('student_id'))
        component_id = int(request.POST.get('component_id'))

        # FIXED — Python uses strip(), not trim()
        raw_val = (request.POST.get('value') or '').strip()

    except (TypeError, ValueError):
        return HttpResponseBadRequest("Invalid parameters")

    course = get_object_or_404(Course, id=course_id)
    comp = get_object_or_404(AssessmentComponent, id=component_id, course=course)

    # Ensure active enrollment
    enr = StudentCourse.objects.filter(
        course=course, student_id=student_id,
        status__in=['ENR', 'CMP'], is_active_pre_reg=True
    ).first()

    if not enr:
        return JsonResponse({'ok': False, 'error': 'Student not enrolled in active cycle'}, status=400)

    # ======= DELETE MARK =======
    if raw_val == "":
        with transaction.atomic():
            AssessmentScore.objects.filter(
                student_id=student_id,
                course=course,
                component=comp
            ).delete()

        enr.grade = None
        enr.outcome = None
        enr.save(update_fields=["grade", "outcome"])

        return JsonResponse({'ok': True, 'value': '', 'grade': '—', 'outcome': '—'})

    # ======= SAVE MARK =======
    try:
        val = Decimal(raw_val)
    except InvalidOperation:
        return JsonResponse({'ok': False, 'error': 'Invalid number'}, status=400)

    if val < 0 or val > Decimal(str(comp.max_marks)):
        return JsonResponse({'ok': False, 'error': f'0 to {comp.max_marks}'}, status=400)

    with transaction.atomic():
        obj, created = AssessmentScore.objects.select_for_update().get_or_create(
            student_id=student_id,
            course=course,
            component=comp,
            defaults={'marks_obtained': val}
        )
        if not created:
            obj.marks_obtained = val
            obj.save(update_fields=['marks_obtained'])

    # ======= RECALCULATE =======
    components = AssessmentComponent.objects.filter(course=course)
    scores = AssessmentScore.objects.filter(student_id=student_id, course=course)

    total_weight = Decimal("0")
    weighted_sum = Decimal("0")

    comp_map = {c.id: c for c in components}
    score_map = {s.component_id: s.marks_obtained for s in scores}

    for c in components:
        max_m = Decimal(str(c.max_marks))
        w = Decimal(str(c.weight or 1))
        total_weight += w

        if c.id in score_map:
            weighted_sum += (score_map[c.id] / max_m) * w

    # Incomplete marks → clear grade
    if len(score_map) != len(components):
        enr.grade = None
        enr.outcome = None
        enr.save(update_fields=["grade", "outcome"])
        return JsonResponse({'ok': True, 'value': str(val), 'grade': '—', 'outcome': '—'})

    percentage = float(weighted_sum / total_weight * 100)

    policy = GradingPolicy.objects.filter(course=course).first()

    grade = 'F'
    outcome = 'FAI'

    # ABSOLUTE
    if policy and policy.mode == "ABS":
        cutoff = policy.abs_cutoffs
        if percentage >= cutoff["A"]: grade = "A"
        elif percentage >= cutoff["A-"]: grade = "A-"
        elif percentage >= cutoff["B"]: grade = "B"
        elif percentage >= cutoff["B-"]: grade = "B-"
        elif percentage >= cutoff["C"]: grade = "C"
        elif percentage >= cutoff["Pass"]: grade = "D"
        else: grade = "F"

        outcome = "PAS" if grade != "F" else "FAI"

    # RELATIVE
    elif policy and policy.mode == "REL":
        all_scores = []
        all_enrs = StudentCourse.objects.filter(course=course, status='ENR', is_active_pre_reg=True)

        for e in all_enrs:
            sc = AssessmentScore.objects.filter(student=e.student, course=course)
            st_weight = Decimal("0")
            st_sum = Decimal("0")
            ok = True
            for c in components:
                try:
                    m = sc.get(component=c).marks_obtained
                except:
                    ok = False
                    break
                st_sum += (m / Decimal(str(c.max_marks))) * Decimal(str(c.weight or 1))
                st_weight += Decimal(str(c.weight or 1))
            if ok:
                all_scores.append(float(st_sum / st_weight * 100))

        all_scores = sorted(all_scores)
        idx = all_scores.index(percentage)
        percentile = (idx / (len(all_scores) - 1)) * 100 if len(all_scores) > 1 else 100

        rel = policy.rel_buckets
        if percentile <= rel.get("top10", 10): grade = "A"
        elif percentile <= 10 + rel.get("next15", 15): grade = "A-"
        elif percentile <= 50: grade = "B"
        elif percentile <= 75: grade = "B-"
        elif percentile <= 90: grade = "C"
        else: grade = "D"
        outcome = "PAS"

    enr.grade = grade
    enr.outcome = outcome
    enr.save(update_fields=["grade", "outcome"])

    return JsonResponse({
        'ok': True,
        'value': str(val),
        'grade': grade,
        'outcome': outcome
    })



def _canon(s: str) -> str:
    import re
    return re.sub(r'[^a-z0-9]+','', (s or "").lower())

def _iter_rows_as_dicts(uploaded_file):
    """
    Yield rows as dicts keyed by header from CSV/XLSX/XLS.
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()

    # CSV
    if name.endswith(".csv"):
        text = TextIOWrapper(uploaded_file.file, encoding="utf-8", newline="")
        reader = csv.DictReader(text)
        for row in reader:
            yield row
        return

    # Excel
    if name.endswith(".xlsx") or name.endswith(".xls"):
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active  # use first sheet
        rows = ws.iter_rows(values_only=True)
        headers = None
        for idx, r in enumerate(rows):
            if idx == 0:
                headers = [str(h or "").strip() for h in r]
                continue
            row = {}
            for h, v in zip(headers, r):
                row[h] = "" if v is None else v
            yield row
        return

    # Fallback to CSV
    text = TextIOWrapper(uploaded_file.file, encoding="utf-8", newline="")
    reader = csv.DictReader(text)
    for row in reader:
        yield row


def all_courses(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)

    qs = (
        Course.objects
        .filter(faculties=faculty)
        .prefetch_related('faculties')
        .annotate(
            enrolled_count=Count(
                'enrollments',
                filter=Q(enrollments__status='ENR', enrollments__is_active_pre_reg=True),
                distinct=True
            )
        )
        .order_by('code')
    )

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

    for c in qs:
        sem_disp = getattr(c, 'semester_display', None) or getattr(c, 'semester', '') or ''
        c.semester_display = sem_disp
        if not hasattr(c, 'enrolled_count') or c.enrolled_count is None:
            c.enrolled_count = 0

    return render(request, 'instructor/all_courses.html', {
        'faculty': faculty,
        'courses': qs,
    })


def assign_grading_policy(request, course_code):
    email = request.session.get("email_id")
    faculty = get_object_or_404(Faculty, email_id=email)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)

    components = list(AssessmentComponent.objects.filter(course=course))
    total_weight = sum([float(c.weight or 0) for c in components]) if components else None

    # Count active enrolled students
    num_enrolled = StudentCourse.objects.filter(
        course=course, status="ENR", is_active_pre_reg=True
    ).count()

    # Fetch or create grading policy
    policy, _ = GradingPolicy.objects.get_or_create(course=course)

    # POST SAVE
    if request.method == "POST":
        mode = (request.POST.get("mode") or "ABS").upper()

        # Force absolute if < 25 students
        if num_enrolled < 25:
            mode = "ABS"

        abs_cutoffs = {
            "A": int(request.POST.get("abs_A") or 85),
            "A-": int(request.POST.get("abs_Aminus") or 75),
            "B": int(request.POST.get("abs_B") or 65),
            "B-": int(request.POST.get("abs_Bminus") or 55),
            "C": int(request.POST.get("abs_C") or 45),
            "Pass": int(request.POST.get("abs_Pass") or 45),
        }

        rel_buckets = {
            "top10": float(request.POST.get("top10") or 10),
            "next15": float(request.POST.get("next15") or 15),
            "next25": float(request.POST.get("next25") or 25),
            "next25b": float(request.POST.get("next25b") or 25),
            "next15b": float(request.POST.get("next15b") or 15),
            "next10": float(request.POST.get("next10") or 10),
            "rest_min": float(request.POST.get("rest_min") or 4),
        }

        with transaction.atomic():
            policy.mode = mode
            policy.abs_cutoffs = abs_cutoffs
            policy.rel_buckets = rel_buckets
            policy.save()

        messages.success(request, f"Grading policy saved for {course.code}.")
        return redirect("faculty_dashboard")

    # DEFAULTS for first-time policy
    abs_defaults = {"A": 85, "A-": 75, "B": 65, "B-": 55, "C": 45, "Pass": 45}
    rel_defaults = {
        "top10": 10, "next15": 15, "next25": 25,
        "next25b": 25, "next15b": 15,
        "next10": 10, "rest_min": 4,
    }

    abs_cutoffs = {**abs_defaults, **(policy.abs_cutoffs or {})}
    rel_buckets = {**rel_defaults, **(policy.rel_buckets or {})}

    abs_cutoffs_safe = {
        "A": abs_cutoffs.get("A"),
        "A_minus": abs_cutoffs.get("A-"),
        "B": abs_cutoffs.get("B"),
        "B_minus": abs_cutoffs.get("B-"),
        "C": abs_cutoffs.get("C"),
        "Pass": abs_cutoffs.get("Pass"),
    }

    rel_buckets_safe = {k: rel_buckets.get(k) for k in rel_defaults}

    force_absolute = num_enrolled < 25

    return render(request, "instructor/assign_grading.html", {
        "course": course,
        "faculty": faculty,
        "components": components,
        "total_weight": total_weight,
        "policy": policy,
        "abs_cutoffs_safe": abs_cutoffs_safe,
        "rel_buckets_safe": rel_buckets_safe,
        "num_enrolled": num_enrolled,
        "force_absolute": force_absolute,
    })






def grade_results(request, course_code):
    # Auth and course
    email = request.session.get("email_id")
    faculty = get_object_or_404(Faculty, email_id=email)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)

    # Components (shown as columns)
    components = list(AssessmentComponent.objects.filter(course=course))

    # ENR or CMP enrollments, across all semesters (shows completed too)
    enrollments = (
        StudentCourse.objects
        .filter(course=course, status__in=['ENR', 'CMP'], is_active_pre_reg=True)
        .select_related('student')
        .order_by('student__roll_no')
    )

    LETTER_TO_POINTS = {
        "A": 10,
        "A-": 9,
        "B": 8,
        "B-": 7,
        "C": 6,
        "C-": 5,
        "D": 4,
        "F": 0,
    }

    def letter_for_points(points):
        try:
            p = int(points)
        except (TypeError, ValueError):
            return ''
        if p >= 10: return 'A'
        if p == 9:  return 'A-'
        if p == 8:  return 'B'
        if p == 7:  return 'B-'
        if p == 6:  return 'C'
        if p == 5:  return 'C-'
        if p == 4:  return 'D'
        return 'F'

    rows = []
    for sc in enrollments:
        comp_marks = {}
        total_percent = Decimal("0")
        for comp in components:
            s = AssessmentScore.objects.filter(student=sc.student, course=course, component=comp).first()
            mark = s.marks_obtained if s else None
            comp_marks[comp.id] = mark
            if s and comp.max_marks and comp.weight:
                try:
                    total_percent += (Decimal(str(s.marks_obtained)) / Decimal(str(comp.max_marks))) * Decimal(str(comp.weight))
                except (InvalidOperation, TypeError):
                    pass

        pts_int = LETTER_TO_POINTS.get(sc.grade, None)
        letter = sc.grade or (letter_for_points(pts_int) if pts_int is not None else "")

        rows.append({
            'student': sc.student,
            'roll': sc.student.roll_no,
            'name': f"{sc.student.first_name} {sc.student.last_name or ''}".strip(),
            'components': comp_marks,
            'total_percent': float(total_percent.quantize(Decimal('0.01'))),
            'points': pts_int,
            'letter': letter,
            'outcome': sc.outcome or '',
        })

    # Sorting
    sort = (request.GET.get("sort") or "roll_asc").lower()
    reverse = sort.endswith("_desc")
    base = sort.split("_")[0]
    if base == "name":
        rows.sort(key=lambda r: (r['name'].lower(), r['roll'].lower()), reverse=reverse)
    elif base == "roll":
        rows.sort(key=lambda r: r['roll'].lower(), reverse=reverse)
    elif base == "points":
        rows.sort(key=lambda r: (r['points'] if r['points'] is not None else -1, r['name'].lower()), reverse=reverse)
    elif base == "letter":
        rows.sort(key=lambda r: (r['points'] if r['points'] is not None else -1, r['name'].lower()), reverse=reverse)

    return render(request, 'instructor/grade_results.html', {
        'course': course,
        'faculty': faculty,
        'components': components,
        'rows': rows,
        'sort': sort,
    })

def student_result_semester_list(request, roll_no):
    student = get_object_or_404(Student, roll_no=roll_no)
    admin = Admins.objects.first()
    results_visible = admin.is_results_visible() if admin else False

    # Current semester (for highlighting)
    current_sem = student.calculate_current_semester()

    # ✅ Include all semesters that have at least one completed/enrolled course
    semesters = (
        StudentCourse.objects
        .filter(student=student)
        .exclude(status__in=["PND", "INS", "DRP"])  # skip pending, in-progress, or dropped
        .values_list("semester", flat=True)
        .distinct()
        .order_by("semester")
    )

    # ✅ Handle missing surname safely
    full_name = student.first_name
    if getattr(student, "last_name", None):
        full_name += f" {student.last_name}"

    context = {
        "student": student,
        "semesters": semesters,
        "current_sem": current_sem,
        "results_visible": results_visible,
        "full_name": full_name.strip(),
    }
    return render(request, "student/result_semester_list.html", context)

def student_view_results(request, student_id, semester):
    student = get_object_or_404(Student, id=student_id)

    # ✅ Fetch all valid enrollments for the semester
    enrollments = (
        StudentCourse.objects
        .filter(student=student, semester=semester)
        .exclude(status__in=["PND", "INS", "DRP"])  # skip pending/unapproved/dropped
        .select_related("course")
        .order_by("course__code")
    )

    # ✅ Prepare result list for template
    results = []
    for enr in enrollments:
        course = enr.course
        grade = enr.grade or "—"
        points = GRADE_POINTS.get(grade, 0)

        # Display "FS" explicitly as "Fail (Short Attendance)" in table
        display_grade = "FS" if grade == "FS" else grade

        results.append({
            "course_code": course.code,
            "course_name": course.name,
            "credits": course.credits,
            "grade": display_grade,
            "points": points if enr.course_mode == "REG" else "—",
            "outcome": enr.outcome or "UNK",
            "course_mode": enr.course_mode or "REG",
        })

    # ✅ Compute metrics directly using Student model helpers
    sem_metrics = student.calculate_semester_metrics(semester)
    cum_metrics = student.calculate_cumulative_metrics()

    # ✅ Handle full name properly
    full_name = student.first_name
    if getattr(student, "last_name", None):
        full_name += f" {student.last_name}"

    # ✅ Render the results page
    return render(request, "student/result_semester_view.html", {
        "student": student,
        "semester": semester,
        "results": results,
        "metrics": sem_metrics,
        "cg_metrics": cum_metrics,
        "full_name": full_name.strip(),
    })

def student_result_pdf(request, student_id, semester):
    student = get_object_or_404(Student, id=student_id)

    # Fetch semester enrollments
    enrollments = (
        StudentCourse.objects
        .filter(student=student, semester=semester, status__in=['ENR', 'CMP'])
        .select_related('course')
    )

    results = []
    for enr in enrollments:
        course = enr.course
        grade = enr.grade or ''
        points = GRADE_POINTS.get(grade, 0)
        credits = course.credits or 0
        mode = (enr.course_mode or "REG").upper()
        outcome = (enr.outcome or "UNK").upper()

        # Determine outcome display based on course mode
        if mode == "AUD":
            if outcome == "PAS":
                outcome_display = "Pass (Audit)"
            elif outcome == "FAI":
                outcome_display = "Fail (Audit)"
            else:
                outcome_display = "Unknown (Audit)"
        elif mode == "PF":
            if outcome == "PAS":
                outcome_display = "Pass (Pass/Fail)"
            elif outcome == "FAI":
                outcome_display = "Fail (Pass/Fail)"
            else:
                outcome_display = "Unknown (Pass/Fail)"
        else:
            if outcome == "PAS":
                outcome_display = "Pass"
            elif outcome == "FAI":
                outcome_display = "Fail"
            else:
                outcome_display = "Unknown"

        # For Audit/PassFail: no grade/points
        display_grade = grade if mode == "REG" else "-"
        display_points = points if mode == "REG" else "-"

        results.append({
            "course_code": course.code,
            "course_name": course.name,
            "credits": credits,
            "grade": display_grade,
            "points": display_points,
            "outcome_display": outcome_display,  # ✅ use this
            "course_mode": mode,
        })

    # Get metrics from Student methods
    metrics = student.calculate_semester_metrics(semester)
    cg_metrics = student.calculate_cumulative_metrics()

    # Render to HTML for PDF
    html_string = render_to_string(
        "student/marksheet_pdf.html",
        {
            "student": student,
            "semester": semester,
            "results": results,
            "metrics": metrics,
            "cg_metrics": cg_metrics,
            "academic_year": f"{datetime.now().year - 1}-{datetime.now().year}",
            "generated_on": datetime.now().strftime("%B %d, %Y %H:%M"),
        },
    )

    html = HTML(string=html_string)
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    filename = f"Marksheet_{student.roll_no}_Sem_{semester}.pdf"
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return response


def get_main_admin():
    """Fetch main admin record (assuming 1 admin controls visibility)."""
    return Admins.objects.first()


def admin_grade_management(request):
    """
    Admin landing page: List all courses with enrolled students + control result visibility.
    """
    # Get all courses with at least one enrolled student
    courses = Course.objects.annotate(
        enrolled_count=Count('enrollments', filter=Q(enrollments__status__in=['ENR', 'CMP']))
    ).filter(enrolled_count__gt=0)

    # Prepare course data (no grading_type needed)
    course_data = [
        {
            'code': c.code,
            'name': c.name,
            'slot': c.slot,
            'enrolled_count': c.enrolled_count,
        }
        for c in courses
    ]

    # Get unique semesters and slots for filters
    semesters = StudentCourse.objects.values_list('semester', flat=True).distinct().order_by('semester')
    slots = Course.objects.values_list('slot', flat=True).distinct().order_by('slot')

    # ✅ Fetch result visibility settings from Admin model
    admin = get_main_admin()
    if not admin:
        admin = Admins.objects.create(
            first_name="Default", last_name="Admin",
            email_id="admin@example.com", password="admin"
        )

    return render(request, 'admin/assign_grade_management.html', {
        'courses': course_data,
        'semesters': semesters,
        'slots': slots,
        'admin': admin,  # 👈 Pass to template for mode + deadline
    })


def set_result_visibility(request):
    """
    Handle result visibility setting form submission.
    """
    if request.method == "POST":
        admin = get_main_admin()
        mode = request.POST.get("mode")
        deadline = request.POST.get("deadline") or None

        admin.results_mode = mode
        admin.results_deadline = deadline
        admin.save()

        messages.success(request, "Result visibility settings updated successfully!")
    return redirect("admin_grade_management")

def admin_assign_grades(request, course_code):
    """
    Display all enrolled students for a course to assign/edit grades
    Allows editing for both ENR and CMP statuses!
    """
    course = get_object_or_404(Course, code=course_code)
    
    # Show both ENR and CMP
    enrollments = StudentCourse.objects.filter(
        course=course,
        status__in=['ENR', 'CMP']
    ).select_related('student').order_by('student__roll_no')

    return render(request, 'admin/assign_edit_grades.html', {
        'course': course,
        'enrollments': enrollments,
    })

def admin_save_grades(request, course_code):
    """
    Save grades submitted by admin for a course.
    Now updates grades for BOTH ENR and CMP records, and always sets them to CMP.
    """
    if request.method != 'POST':
        return redirect('admin_assign_grades', course_code=course_code)

    course = get_object_or_404(Course, code=course_code)
    
    # Allow grade changes for both ENR and CMP status enrollments
    enrollments = StudentCourse.objects.filter(course=course, status__in=['ENR', 'CMP'])

    updated_count = 0
    for enrollment in enrollments:
        grade_key = f'grade_{enrollment.id}'
        outcome_key = f'outcome_{enrollment.id}'
        grade = request.POST.get(grade_key, '').strip()
        outcome = request.POST.get(outcome_key, 'UNK')

        # Update if grade is provided
        if grade and grade in GRADE_POINTS:
            enrollment.grade = grade
            enrollment.outcome = outcome
            enrollment.status = 'CMP'  # Always set as complete when edited
            enrollment.save(update_fields=['grade', 'outcome', 'status'])
            updated_count += 1

    messages.success(request, f'Successfully updated grades for {updated_count} student(s) in {course.code}.')
    return redirect('admin_assign_grades', course_code=course_code)

def admin_upload_results(request):
    """
    Bulk upload of previous-semester results.
    Uses admin-selected default course_mode (REG / PF / AUD).
    """

    context = {
        "columns": ["roll_no", "course_code", "grade", "outcome"],
        "preview_data": None,
        "summary": None,
        "selected_semester": None,
    }

    if request.method == "POST":
        file = request.FILES.get("file")
        text_data = request.POST.get("csv_text", "").strip()

        semester = int(request.POST.get("semester") or 1)
        default_mode = request.POST.get("default_mode", "REG")  # 🔥 new
        context["selected_semester"] = semester

        # ------------------ Read Uploaded Data ------------------
        try:
            if file:
                if file.name.endswith(".csv"):
                    df = pd.read_csv(file)
                elif file.name.endswith((".xls", ".xlsx")):
                    df = pd.read_excel(file)
                else:
                    messages.error(request, "Invalid file format. Use CSV or Excel.")
                    return redirect("admin_upload_results")
            elif text_data:
                df = pd.read_csv(io.StringIO(text_data))
            else:
                messages.error(request, "Upload a file or paste CSV data.")
                return redirect("admin_upload_results")
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("admin_upload_results")

        df.columns = df.columns.str.lower()

        required = {"roll_no", "course_code", "grade"}
        missing = required - set(df.columns)
        if missing:
            messages.error(request, f"Missing required columns: {', '.join(missing)}")
            return redirect("admin_upload_results")

        if "outcome" not in df.columns:
            df["outcome"] = ""

        df = df.fillna("")
        rows = df.to_dict(orient="records")

        # ---------------- Preview ----------------
        if "preview" in request.POST:
            context["preview_data"] = rows
            messages.info(request, "Preview loaded.")
            return render(request, "admin/admin_upload_results.html", context)

        # ---------------- Final Upload ----------------
        created = updated = skipped = 0
        errors = []
        success = []

        with transaction.atomic():
            for row in rows:
                roll = str(row.get("roll_no", "")).strip()
                code = str(row.get("course_code", "")).strip().upper()
                grade = str(row.get("grade", "")).strip().upper()
                outcome = str(row.get("outcome", "")).strip().upper()

                if not roll or not code or not grade:
                    skipped += 1
                    errors.append(f"Invalid row: {row}")
                    continue

                student = Student.objects.filter(roll_no__iexact=roll).first()
                course = Course.objects.filter(code__iexact=code).first()

                if not student:
                    skipped += 1
                    errors.append(f"Roll not found: {roll}")
                    continue
                if not course:
                    skipped += 1
                    errors.append(f"Course not found: {code}")
                    continue

                # ---------------- Determine Course Mode ----------------
                mode = default_mode  # admin-selected default

                # Override based on grade
                if grade in ["P", "F"]:
                    mode = "PF"
                elif grade == "AUD":
                    mode = "AUD"

                # ---------------- Determine Outcome ----------------
                if mode == "AUD":
                    final_outcome = "PAS"
                elif grade == "F":
                    final_outcome = "FAI"
                else:
                    final_outcome = "PAS"

                sc, new = StudentCourse.objects.update_or_create(
                    student=student,
                    course=course,
                    semester=semester,
                    defaults={
                        "grade": grade,
                        "outcome": final_outcome,
                        "status": "CMP",
                        "course_mode": mode,
                        "is_active_pre_reg": False,
                    }
                )

                if new:
                    created += 1
                    success.append(f"Added {roll}-{code}")
                else:
                    updated += 1
                    success.append(f"Updated {roll}-{code}")

        context["summary"] = {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "success_details": success[:10],
        }

        messages.success(request, f"Upload complete: {created} created, {updated} updated, {skipped} skipped.")
        return render(request, "admin/admin_upload_results.html", context)

    return render(request, "admin/admin_upload_results.html", context)



def result_management_home(request):
    return render(request, "admin/result_module.html")









# For Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# For PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER,TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm,inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def database_management_view(request):
    """
    Main view for database management dashboard
    Loads all data with pagination support
    """

    # -------------------------------
    # Fetch all data with sorting
    # -------------------------------
    students = Student.objects.select_related('branch', 'department').order_by('roll_no')
    faculty = Faculty.objects.select_related('department').order_by('first_name')
    courses = Course.objects.all().order_by('code')
    branches = Branch.objects.select_related('department').order_by('name')
    departments = Department.objects.all().order_by('code')
    enrollments = StudentCourse.objects.select_related('student', 'course').order_by('student__roll_no', 'semester')
    categories = Category.objects.all().order_by('code')
    requirements = ProgramRequirement.objects.select_related('branch').order_by('branch__name')

    # -------------------------------
    # Prepare Students Data
    # -------------------------------
    students_data = []
    for student in students:
        students_data.append({
            'id': student.id,
            'roll_no': student.roll_no,
            'first_name': student.first_name,
            'last_name': student.last_name or '',
            'email': student.email_id,
            'branch': student.branch.name if student.branch else 'N/A',
            'department': student.department.code if student.department else 'N/A',
            'semester': student.calculate_current_semester(),
            'mobile': str(student.mobile_no) if student.mobile_no else 'N/A'
        })

    # -------------------------------
    # Prepare Faculty Data
    # -------------------------------
    faculty_data = []
    for fac in faculty:
        course_count = fac.courses.count()
        faculty_data.append({
            'id': fac.id,
            'first_name': fac.first_name,
            'last_name': fac.last_name or '',
            'email': fac.email_id,
            'department': fac.department.code if fac.department else 'N/A',
            'mobile': str(fac.mobile_no) if fac.mobile_no else 'N/A',
            'courses': course_count
        })

    # -------------------------------
    # Prepare Courses Data
    # -------------------------------
    courses_data = []
    for course in courses:
        enrolled_count = course.enrollments.filter(status__in=['ENR', 'CMP']).count()
        courses_data.append({
            'id': course.id,
            'code': course.code,
            'name': course.name,
            'credits': course.credits,
            'ltpc': course.LTPC,
            'slot': course.slot,
            'status': course.status,
            'enrolled': enrolled_count
        })

    # -------------------------------
    # Prepare Branches Data
    # -------------------------------
    branches_data = []
    for branch in branches:
        student_count = branch.student_set.count()
        course_count = branch.courses.count()
        branches_data.append({
            'id': branch.id,
            'name': branch.name,
            'full_name': dict(Branch.BRANCHES).get(branch.name, branch.name),
            'department': branch.department.code if branch.department else 'N/A',
            'students': student_count,
            'courses': course_count
        })

    # -------------------------------
    # Prepare Departments Data
    # -------------------------------
    departments_data = []
    for dept in departments:
        branch_count = dept.branches.count()
        faculty_count = dept.faculty_set.count()
        student_count = dept.student_set.count()
        departments_data.append({
            'id': dept.id,
            'code': dept.code,
            'name': dept.name,
            'branches': branch_count,
            'faculty': faculty_count,
            'students': student_count
        })

    # -------------------------------
    # Prepare Enrollments Data
    # -------------------------------
    enrollments_data = []
    for enr in enrollments:
        enrollments_data.append({
            'id': enr.id,
            'student': f"{enr.student.first_name} {enr.student.last_name or ''}".strip(),
            'roll_no': enr.student.roll_no,
            'course': enr.course.code,
            'semester': enr.semester or 'N/A',
            'status': enr.status,
            'grade': enr.grade or '-',
            'outcome': enr.outcome,
            'course_mode': enr.course_mode  # ✅ new field replacing is_pass_fail
        })

    # -------------------------------
    # Prepare Categories Data
    # -------------------------------
    categories_data = []
    for cat in categories:
        course_count = cat.course_branches.count()
        categories_data.append({
            'id': cat.id,
            'code': cat.code,
            'label': cat.label,
            'courses': course_count
        })

    # -------------------------------
    # Prepare Requirements Data
    # -------------------------------
    requirements_data = []
    for req in requirements:
        requirements_data.append({
            'id': req.id,
            'branch': req.branch.name,
            'category': req.category,
            'required_credits': req.required_credits
        })

    # -------------------------------
    # Context for Template
    # -------------------------------
    context = {
        'students_json': json.dumps(students_data),
        'faculty_json': json.dumps(faculty_data),
        'courses_json': json.dumps(courses_data),
        'branches_json': json.dumps(branches_data),
        'departments_json': json.dumps(departments_data),
        'enrollments_json': json.dumps(enrollments_data),
        'categories_json': json.dumps(categories_data),
        'requirements_json': json.dumps(requirements_data),
    }

    return render(request, 'admin/database_management.html', context)


def edit_database_record(request, record_type, record_id):
    if request.method == 'POST':
        try:
            # -------------------- STUDENT --------------------
            if record_type == 'student':
                student = get_object_or_404(Student, id=record_id)
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                email = request.POST.get('email')
                mobile = request.POST.get('mobile')
                branch_id = request.POST.get('branch')

                if first_name: student.first_name = first_name
                if last_name: student.last_name = last_name
                if email: student.email_id = email
                if mobile and mobile.isdigit(): student.mobile_no = int(mobile)
                if branch_id: student.branch = Branch.objects.filter(id=branch_id).first()

                student.save()
                messages.success(request, f"✅ Student '{student.first_name}' updated successfully.")
                return redirect('database_management')

            # -------------------- FACULTY --------------------
            elif record_type == 'faculty':
                faculty = get_object_or_404(Faculty, id=record_id)
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                email = request.POST.get('email')
                mobile = request.POST.get('mobile')
                dept_id = request.POST.get('department')

                if first_name: faculty.first_name = first_name
                if last_name: faculty.last_name = last_name
                if email: faculty.email_id = email
                if mobile and mobile.isdigit(): faculty.mobile_no = int(mobile)
                if dept_id: faculty.department = Department.objects.filter(id=dept_id).first()

                faculty.save()
                messages.success(request, f"✅ Faculty '{faculty.first_name}' updated successfully.")
                return redirect('database_management')

            # -------------------- COURSE --------------------
            elif record_type == 'course':
                course = get_object_or_404(Course, id=record_id)
                code = request.POST.get('code')
                name = request.POST.get('name')
                credits = request.POST.get('credits')
                ltpc = request.POST.get('ltpc')
                slot = request.POST.get('slot')
                status = request.POST.get('status')

                if code: course.code = code
                if name: course.name = name
                if credits and credits.isdigit(): course.credits = int(credits)
                if ltpc: course.LTPC = ltpc
                if slot: course.slot = slot
                if status: course.status = status

                course.save()
                messages.success(request, f"✅ Course '{course.code}' updated successfully.")
                return redirect('database_management')

            # -------------------- ENROLLMENT --------------------
            elif record_type == 'enrollment':
                enrollment = get_object_or_404(StudentCourse, id=record_id)
                status = request.POST.get('status')
                grade = request.POST.get('grade')
                outcome = request.POST.get('outcome')
                course_mode = request.POST.get('course_mode')

                if status: enrollment.status = status
                if grade: enrollment.grade = grade
                if outcome: enrollment.outcome = outcome
                if course_mode in ['REG', 'PF', 'AUD']:
                    enrollment.course_mode = course_mode

                enrollment.save()
                messages.info(request, f"ℹ️ Enrollment record for '{enrollment.student.roll_no}' updated.")
                return redirect('database_management')

            # -------------------- REQUIREMENT --------------------
            elif record_type == 'requirement':
                requirement = get_object_or_404(ProgramRequirement, id=record_id)
                required_credits = request.POST.get('required_credits')
                if required_credits and required_credits.isdigit():
                    requirement.required_credits = int(required_credits)
                requirement.save()
                messages.success(request, f"✅ Requirement record updated successfully.")
                return redirect('database_management')

            # -------------------- INVALID TYPE --------------------
            else:
                messages.error(request, "❌ Invalid record type specified.")
                return redirect('database_management')

        except Exception as e:
            messages.error(request, f"⚠️ Error while updating record: {str(e)}")
            return redirect('database_management')

    # -------------------- GET REQUEST: Edit Form --------------------
    context = {}

    if record_type == 'student':
        student = get_object_or_404(Student, id=record_id)
        branches = Branch.objects.all()
        context = {
            'record_type': 'Student',
            'record': student,
            'branches': branches
        }

    elif record_type == 'faculty':
        faculty = get_object_or_404(Faculty, id=record_id)
        departments = Department.objects.all()
        context = {
            'record_type': 'Faculty',
            'record': faculty,
            'departments': departments
        }

    elif record_type == 'course':
        course = get_object_or_404(Course, id=record_id)
        context = {
            'record_type': 'Course',
            'record': course,
            'slots': Course.SLOT_CHOICES
        }

    elif record_type == 'enrollment':
        enrollment = get_object_or_404(StudentCourse, id=record_id)
        context = {
            'record_type': 'Enrollment',
            'record': enrollment,
            'statuses': StudentCourse.STATUS,
            'outcomes': StudentCourse.OUTCOME,
            'grades': StudentCourse.GRADES,
            'modes': StudentCourse.COURSE_MODE,
        }

    elif record_type == 'requirement':
        requirement = get_object_or_404(ProgramRequirement, id=record_id)
        context = {
            'record_type': 'Requirement',
            'record': requirement
        }

    return render(request, 'admin/edit_record.html', context)


@require_http_methods(["POST"])
def delete_database_record(request, record_type, record_id):
    """
    Delete a specific record
    """
    try:
        if record_type == 'student':
            record = get_object_or_404(Student, id=record_id)
        elif record_type == 'faculty':
            record = get_object_or_404(Faculty, id=record_id)
        elif record_type == 'course':
            record = get_object_or_404(Course, id=record_id)
        elif record_type == 'enrollment':
            record = get_object_or_404(StudentCourse, id=record_id)
        elif record_type == 'requirement':
            record = get_object_or_404(ProgramRequirement, id=record_id)
        elif record_type == 'branch':
            record = get_object_or_404(Branch, id=record_id)
        elif record_type == 'department':
            record = get_object_or_404(Department, id=record_id)
        elif record_type == 'category':
            record = get_object_or_404(Category, id=record_id)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid record type'})
        
        record.delete()
        return JsonResponse({'success': True, 'message': f'{record_type.capitalize()} deleted successfully'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_http_methods(["GET"])
def ajax_edit_record(request, record_type, record_id):
    """
    Returns HTML for the edit modal body (AJAX).
    """
    context = {}
    
    if record_type == 'student':
        record = get_object_or_404(Student, id=record_id)
        branches = Branch.objects.all()
        context = {"record_type": "student", "record": record, "branches": branches}

    elif record_type == 'faculty':
        record = get_object_or_404(Faculty, id=record_id)
        departments = Department.objects.all()
        context = {"record_type": "faculty", "record": record, "departments": departments}

    elif record_type == 'course':
        record = get_object_or_404(Course, id=record_id)
        context = {"record_type": "course", "record": record, "slots": Course.SLOT_CHOICES}

    elif record_type == 'enrollment':
        record = get_object_or_404(StudentCourse, id=record_id)
        context = {
            "record_type": "enrollment",
            "record": record,
            "statuses": StudentCourse.STATUS,
            "outcomes": StudentCourse.OUTCOME,
            "grades": StudentCourse.GRADES,
            "modes": StudentCourse.COURSE_MODE,
        }

    elif record_type == 'requirement':
        record = get_object_or_404(ProgramRequirement, id=record_id)
        context = {"record_type": "requirement", "record": record}

    else:
        return JsonResponse({"success": False, "error": "Invalid type"})

    html = render_to_string("admin/database_edit_modal.html", context, request=request)
    return JsonResponse({"success": True, "html": html})

@require_http_methods(["POST"])
def ajax_save_record(request, record_type, record_id):
    try:
        if record_type == "student":
            student = get_object_or_404(Student, id=record_id)
            student.first_name = request.POST.get("first_name", student.first_name)
            student.last_name = request.POST.get("last_name", student.last_name)
            student.email_id = request.POST.get("email", student.email_id)
            branch = request.POST.get("branch")
            if branch:
                student.branch = Branch.objects.filter(id=branch).first()
            student.save()

            return JsonResponse({
                "success": True,
                "updated": {
                    "name": student.first_name + " " + (student.last_name or ""),
                    "email": student.email_id,
                    "branch": student.branch.name if student.branch else "",
                }
            })

        elif record_type == "enrollment":
            enr = get_object_or_404(StudentCourse, id=record_id)

            enr.status = request.POST.get("status", enr.status)
            enr.grade = request.POST.get("grade", enr.grade)
            enr.outcome = request.POST.get("outcome", enr.outcome)
            mode = request.POST.get("course_mode")
            if mode in ["REG", "PF", "AUD"]:
                enr.course_mode = mode
            enr.save()

            return JsonResponse({
                "success": True,
                "updated": {
                    "status": enr.status,
                    "grade": enr.grade,
                    "outcome": enr.outcome,
                    "course_mode": enr.course_mode,
                }
            })

        elif record_type == "course":
            c = get_object_or_404(Course, id=record_id)
            c.code = request.POST.get("code", c.code)
            c.name = request.POST.get("name", c.name)
            c.credits = request.POST.get("credits", c.credits)
            c.slot = request.POST.get("slot", c.slot)
            c.save()
            return JsonResponse({"success": True})

        elif record_type == "requirement":
            r = get_object_or_404(ProgramRequirement, id=record_id)
            credits = request.POST.get("required_credits")
            if credits and credits.isdigit():
                r.required_credits = int(credits)
            r.save()
            return JsonResponse({"success": True})

        return JsonResponse({"success": False, "error": "Invalid type"})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
@require_http_methods(["POST"])
def ajax_delete_record(request, record_type, record_id):
    try:
        if record_type == 'student':
            obj = get_object_or_404(Student, id=record_id)
        elif record_type == 'faculty':
            obj = get_object_or_404(Faculty, id=record_id)
        elif record_type == 'course':
            obj = get_object_or_404(Course, id=record_id)
        elif record_type == 'enrollment':
            obj = get_object_or_404(StudentCourse, id=record_id)
        elif record_type == 'requirement':
            obj = get_object_or_404(ProgramRequirement, id=record_id)
        else:
            return JsonResponse({"success": False, "error": "Invalid type"})

        obj.delete()
        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    


def api_get_record(request, record_type, record_id):
    try:
        if record_type == 'student':
            s = get_object_or_404(Student, id=record_id)
            data = {
                "id": s.id, "roll_no": s.roll_no, "first_name": s.first_name,
                "last_name": s.last_name or "", "email": s.email_id,
                "branch_id": s.branch.id if s.branch else None,
                "department_id": s.department.id if s.department else None,
                "semester": s.semester, "mobile_no": str(s.mobile_no) if s.mobile_no else ""
            }
        elif record_type == 'faculty':
            f = get_object_or_404(Faculty, id=record_id)
            data = {
                "id": f.id, "first_name": f.first_name, "last_name": f.last_name or "",
                "email": f.email_id, "department_id": f.department.id if f.department else None,
                "mobile_no": str(f.mobile_no) if f.mobile_no else ""
            }
        elif record_type == 'course':
            c = get_object_or_404(Course, id=record_id)
            data = {
                "id": c.id, "code": c.code, "name": c.name,
                "credits": c.credits, "ltpc": c.LTPC or "", "slot": c.slot, "status": c.status
            }
        elif record_type == 'enrollment':
            enr = get_object_or_404(StudentCourse, id=record_id)
            data = {
                "id": enr.id,
                "student_id": enr.student.id,
                "student": f"{enr.student.first_name} {enr.student.last_name or ''}".strip(),
                "roll_no": enr.student.roll_no,
                "course_id": enr.course.id,
                "course": enr.course.code,
                "semester": enr.semester,
                "status": enr.status,
                "grade": enr.grade or "",
                "outcome": enr.outcome,
                "course_mode": enr.course_mode
            }
        elif record_type == 'requirement':
            r = get_object_or_404(ProgramRequirement, id=record_id)
            data = {
                "id": r.id, "branch_id": r.branch.id, "category": r.category, "required_credits": r.required_credits
            }
        else:
            return JsonResponse({"ok": False, "error": "invalid type"}, status=400)
        return JsonResponse({"ok": True, "record": data})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@require_POST
def api_update_record(request, record_type, record_id):
    """
    Updates a database record. Accepts form-encoded POST body.
    Returns updated record JSON to update the client arrays.
    """
    data = request.POST.dict()
    try:
        with transaction.atomic():
            if record_type == 'student':
                student = get_object_or_404(Student, id=record_id)
                student.first_name = data.get('first_name', student.first_name)
                student.last_name = data.get('last_name', student.last_name)
                if data.get('email'): student.email_id = data.get('email')
                mobile = data.get('mobile_no')
                if mobile is not None and mobile != '':
                    try:
                        student.mobile_no = int(mobile)
                    except ValueError:
                        pass
                branch_id = data.get('branch_id')
                if branch_id:
                    student.branch = Branch.objects.filter(id=int(branch_id)).first()
                department_id = data.get('department_id')
                if department_id:
                    student.department = Department.objects.filter(id=int(department_id)).first()
                student.save()
                updated = {
                    "id": student.id, "roll_no": student.roll_no, "first_name": student.first_name,
                    "last_name": student.last_name or "", "email": student.email_id,
                    "branch": student.branch.name if student.branch else "N/A",
                    "department": student.department.code if student.department else "N/A",
                    "semester": student.calculate_current_semester(),
                    "mobile": str(student.mobile_no) if student.mobile_no else ""
                }
            elif record_type == 'faculty':
                faculty = get_object_or_404(Faculty, id=record_id)
                faculty.first_name = data.get('first_name', faculty.first_name)
                faculty.last_name = data.get('last_name', faculty.last_name)
                if data.get('email'): faculty.email_id = data.get('email')
                mobile = data.get('mobile_no')
                if mobile is not None and mobile != '':
                    try:
                        faculty.mobile_no = int(mobile)
                    except ValueError:
                        pass
                dept_id = data.get('department_id')
                if dept_id:
                    faculty.department = Department.objects.filter(id=int(dept_id)).first()
                faculty.save()
                updated = {
                    "id": faculty.id, "first_name": faculty.first_name,
                    "last_name": faculty.last_name or "", "email": faculty.email_id,
                    "department": faculty.department.code if faculty.department else "N/A",
                    "mobile": str(faculty.mobile_no) if faculty.mobile_no else "",
                    "courses": faculty.courses.count()
                }
            elif record_type == 'course':
                course = get_object_or_404(Course, id=record_id)
                if data.get('code'): course.code = data.get('code')
                if data.get('name'): course.name = data.get('name')
                if data.get('credits'):
                    try:
                        course.credits = int(data.get('credits'))
                    except ValueError:
                        pass
                if 'ltpc' in data: course.LTPC = data.get('ltpc')
                if 'slot' in data: course.slot = data.get('slot')
                if 'status' in data: course.status = data.get('status')
                course.save()
                updated = {
                    "id": course.id, "code": course.code, "name": course.name,
                    "credits": course.credits, "ltpc": course.LTPC, "slot": course.slot, "status": course.status,
                    "enrolled": course.enrollments.filter(status__in=['ENR','CMP']).count()
                }
            elif record_type == 'enrollment':
                enr = get_object_or_404(StudentCourse, id=record_id)
                if 'status' in data: enr.status = data.get('status')
                if 'grade' in data:
                    enr.grade = data.get('grade') or None
                if 'outcome' in data:
                    enr.outcome = data.get('outcome') or None
                if 'course_mode' in data and data.get('course_mode') in ['REG','PF','AUD']:
                    enr.course_mode = data.get('course_mode')
                enr.save()
                updated = {
                    "id": enr.id, "student": f"{enr.student.first_name} {enr.student.last_name or ''}".strip(),
                    "roll_no": enr.student.roll_no, "course": enr.course.code, "semester": enr.semester,
                    "status": enr.status, "grade": enr.grade or "-", "outcome": enr.outcome or "", "course_mode": enr.course_mode
                }
            elif record_type == 'requirement':
                req = get_object_or_404(ProgramRequirement, id=record_id)
                if data.get('required_credits'):
                    try:
                        req.required_credits = int(data.get('required_credits'))
                    except ValueError:
                        pass
                req.save()
                updated = {
                    "id": req.id, "branch": req.branch.name, "category": req.category, "required_credits": req.required_credits
                }
            else:
                return JsonResponse({"ok": False, "error": "invalid type"}, status=400)

        return JsonResponse({"ok": True, "updated": updated})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@require_POST
def api_delete_record(request, record_type, record_id):
    """
    Deletes the requested record and returns success. Safe wrapped in transaction.
    """
    try:
        with transaction.atomic():
            if record_type == 'student':
                obj = get_object_or_404(Student, id=record_id)
            elif record_type == 'faculty':
                obj = get_object_or_404(Faculty, id=record_id)
            elif record_type == 'course':
                obj = get_object_or_404(Course, id=record_id)
            elif record_type == 'enrollment':
                obj = get_object_or_404(StudentCourse, id=record_id)
            elif record_type == 'requirement':
                obj = get_object_or_404(ProgramRequirement, id=record_id)
            elif record_type == 'branch':
                obj = get_object_or_404(Branch, id=record_id)
            elif record_type == 'department':
                obj = get_object_or_404(Department, id=record_id)
            elif record_type == 'category':
                obj = get_object_or_404(Category, id=record_id)
            else:
                return JsonResponse({"ok": False, "error": "invalid type"}, status=400)

            obj.delete()
        return JsonResponse({"ok": True, "deleted_id": record_id})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


def export_page_view(request):
    """
    Display the export configuration page
    """
    context = {
        'students_count': Student.objects.count(),
        'faculty_count': Faculty.objects.count(),
        'courses_count': Course.objects.count(),
        'enrollments_count': StudentCourse.objects.count(),
        'branches_count': Branch.objects.count(),
        'departments_count': Department.objects.count(),
        'categories_count': Category.objects.count(),
        'requirements_count': ProgramRequirement.objects.count(),
        'branches': Branch.objects.all(),
        'departments': Department.objects.all(),
    }
    return render(request, 'admin/database_export.html', context)


def export_filtered_data(request):
    """
    Export filtered data based on user selections
    """
    data_type = request.GET.get('data_type', '')
    export_format = request.GET.get('format', 'csv')
    
    # Get filtered data based on type
    if data_type == 'students':
        data = filter_students(request)
        headers = ['ID', 'Roll No', 'First Name', 'Last Name', 'Email', 'Branch', 'Department', 'Semester', 'Mobile']
        title = 'Students Data'
        
    elif data_type == 'faculty':
        data = filter_faculty(request)
        headers = ['ID', 'First Name', 'Last Name', 'Email', 'Department', 'Mobile', 'Courses Count']
        title = 'Faculty Data'
        
    elif data_type == 'courses':
        data = filter_courses(request)
        headers = ['ID', 'Code', 'Name', 'Credits', 'LTPC', 'Slot', 'Status', 'Enrolled']
        title = 'Courses Data'
        
    elif data_type == 'enrollments':
        data = filter_enrollments(request)
        headers = ['ID', 'Student', 'Roll No', 'Course', 'Semester', 'Status', 'Grade', 'Outcome']
        title = 'Enrollments Data'
        
    elif data_type == 'branches':
        data = filter_branches(request)
        headers = ['ID', 'Code', 'Full Name', 'Department', 'Students', 'Courses']
        title = 'Branches Data'
        
    elif data_type == 'departments':
        data = filter_departments(request)
        headers = ['ID', 'Code', 'Name', 'Branches', 'Faculty', 'Students']
        title = 'Departments Data'
        
    elif data_type == 'categories':
        data = filter_categories(request)
        headers = ['ID', 'Code', 'Label', 'Courses Count']
        title = 'Categories Data'
        
    elif data_type == 'requirements':
        data = filter_requirements(request)
        headers = ['ID', 'Branch', 'Category', 'Required Credits']
        title = 'Program Requirements Data'
        
    else:
        return HttpResponse('Invalid data type', status=400)
    
    # Export based on format
    if export_format == 'excel':
        return export_to_excel(data, headers, title)
    elif export_format == 'pdf':
        return export_to_pdf(data, headers, title)
    else:  # csv
        return export_to_csv(data, headers, title)

def filter_students(request):
    """Filter students based on query parameters"""
    students = Student.objects.select_related('branch', 'department').all()
    
    # Filter by year (from roll number) - FIXED
    years = request.GET.getlist('year')  # Get multiple years
    if years:
        # Extract year patterns: 2024 -> "24", 2023 -> "23", etc.
        year_patterns = [year[2:] for year in years]  # ["24", "23"]
        # Filter students whose roll_no contains any of these patterns
        from django.db.models import Q
        year_query = Q()
        for pattern in year_patterns:
            # Match roll numbers like "B24001", "b24002", etc.
            year_query |= Q(roll_no__iregex=r'^[a-zA-Z]' + pattern + r'\d+$')
        students = students.filter(year_query)
    
    # Filter by branch - UPDATED for multiple values
    branches = request.GET.getlist('branch')
    if branches:
        students = students.filter(branch__name__in=branches)
    
    
    # Prepare data rows
    data = []
    student_list = students if isinstance(students, list) else students
    for student in student_list:
        data.append([
            student.id,
            student.roll_no,
            student.first_name,
            student.last_name or '',
            student.email_id,
            student.branch.name if student.branch else 'N/A',
            student.department.code if student.department else 'N/A',
            student.calculate_current_semester(),
            str(student.mobile_no) if student.mobile_no else 'N/A'
        ])
    
    return data

def filter_faculty(request):
    """Filter faculty based on query parameters"""
    faculty = Faculty.objects.select_related('department').all()
    
    # Filter by department - UPDATED for multiple values
    departments = request.GET.getlist('department')
    if departments:
        faculty = faculty.filter(department__code__in=departments)
    
    # Filter by minimum courses
    min_courses = request.GET.get('min_courses')
    if min_courses:
        faculty = faculty.annotate(course_count=Count('courses')).filter(course_count__gte=int(min_courses))
    
    # Prepare data rows
    data = []
    for fac in faculty:
        data.append([
            fac.id,
            fac.first_name,
            fac.last_name or '',
            fac.email_id,
            fac.department.code if fac.department else 'N/A',
            str(fac.mobile_no) if fac.mobile_no else 'N/A',
            fac.courses.count()
        ])
    
    return data


def filter_courses(request):
    """Filter courses based on query parameters"""
    courses = Course.objects.all()
    
    # Filter by slot - UPDATED for multiple values
    slots = request.GET.getlist('slot')
    if slots:
        courses = courses.filter(slot__in=slots)
    
    # Filter by credits - UPDATED for multiple values
    credits_list = request.GET.getlist('credits')
    if credits_list:
        credits_ints = [int(c) for c in credits_list]
        courses = courses.filter(credits__in=credits_ints)
    
    # Filter by status - UPDATED for multiple values
    statuses = request.GET.getlist('status')
    if statuses:
        courses = courses.filter(status__in=statuses)
    
    # Prepare data rows
    data = []
    for course in courses:
        enrolled_count = course.enrollments.filter(status__in=['ENR', 'CMP']).count()
        data.append([
            course.id,
            course.code,
            course.name,
            course.credits,
            course.LTPC,
            course.slot,
            course.status,
            enrolled_count
        ])
    
    return data


def filter_enrollments(request):
    """Filter enrollments based on query parameters"""
    enrollments = StudentCourse.objects.select_related('student', 'course').all()
    
    # Filter by semester - UPDATED for multiple values
    semesters = request.GET.getlist('semester')
    if semesters:
        semester_ints = [int(s) for s in semesters]
        enrollments = enrollments.filter(semester__in=semester_ints)
    
    # Filter by status - UPDATED for multiple values
    statuses = request.GET.getlist('status')
    if statuses:
        enrollments = enrollments.filter(status__in=statuses)
    
    # Filter by outcome - UPDATED for multiple values
    outcomes = request.GET.getlist('outcome')
    if outcomes:
        enrollments = enrollments.filter(outcome__in=outcomes)
    
    # Filter by grade - UPDATED for multiple values
    grades = request.GET.getlist('grade')
    if grades:
        enrollments = enrollments.filter(grade__in=grades)
    
    # Prepare data rows
    data = []
    for enr in enrollments:
        data.append([
            enr.id,
            f"{enr.student.first_name} {enr.student.last_name or ''}".strip(),
            enr.student.roll_no,
            enr.course.code,
            enr.semester or 'N/A',
            enr.status,
            enr.grade or '-',
            enr.outcome
        ])
    
    return data


def filter_requirements(request):
    """Filter requirements based on query parameters"""
    requirements = ProgramRequirement.objects.select_related('branch').all()
    
    # Filter by branch - UPDATED for multiple values
    branches = request.GET.getlist('branch')
    if branches:
        requirements = requirements.filter(branch__name__in=branches)
    
    # Filter by category - UPDATED for multiple values
    categories = request.GET.getlist('category')
    if categories:
        requirements = requirements.filter(category__in=categories)
    
    data = []
    for req in requirements:
        data.append([
            req.id,
            req.branch.name,
            req.category,
            req.required_credits
        ])
    
    return data


def export_to_csv(data, headers, title):
    """Export data to CSV format"""
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{title.replace(' ', '_')}_{timestamp}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(data)
    
    return response


def export_to_excel(data, headers, title):
    """Export data to Excel format with styling"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{title.replace(' ', '_')}_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    
    # Styling
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(response)
    return response


def export_to_pdf(data, headers, title):
    """Export data to PDF format with wrapped text and optional email exclusion"""
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{title.replace(' ', '_')}_{timestamp}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=20,
        alignment=TA_CENTER
    )

    title_para = Paragraph(f"{title}<br/>{datetime.now().strftime('%B %d, %Y %H:%M')}", title_style)
    elements.append(title_para)
    elements.append(Spacer(1, 0.3 * inch))

    # ------------------------------------
    # 🧹 Remove email column for Students
    # ------------------------------------
    if "Student" in title or "Students" in title:
        if "Email" in headers:
            email_index = headers.index("Email")
            headers.pop(email_index)
            for i in range(len(data)):
                if len(data[i]) > email_index:
                    data[i].pop(email_index)

    # Limit to 100 rows
    table_data = [headers]
    for row in data[:100]:
        wrapped_row = []
        for cell in row:
            # Convert everything to string safely
            text = str(cell) if cell is not None else ""
            # Wrap long text
            wrapped_row.append(Paragraph(text.replace("\n", "<br/>"), styles["Normal"]))
        table_data.append(wrapped_row)

    # ------------------------------------
    # Dynamic column width adjustment
    # ------------------------------------
    total_width = landscape(A4)[0] - 60
    col_count = len(headers)

    # Heuristic: wider columns for name/code, smaller for numeric fields
    col_widths = []
    for header in headers:
        if any(k in header.lower() for k in ["name", "course", "branch", "department"]):
            col_widths.append(total_width * 0.18)  # ~18% width for long text
        elif any(k in header.lower() for k in ["roll", "id"]):
            col_widths.append(total_width * 0.12)
        elif any(k in header.lower() for k in ["credits", "sem", "status", "grade"]):
            col_widths.append(total_width * 0.1)
        else:
            col_widths.append(total_width * 0.12)

    # Normalize if total exceeds width
    width_sum = sum(col_widths)
    if width_sum > total_width:
        scale = total_width / width_sum
        col_widths = [w * scale for w in col_widths]

    # ------------------------------------
    # Create table
    # ------------------------------------
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # ------------------------------------
    # Add note if data truncated
    # ------------------------------------
    if len(data) > 100:
        elements.append(Spacer(1, 0.3 * inch))
        note = Paragraph(
            f"<i>Note: Showing first 100 of {len(data)} records. For full export, please use Excel or CSV format.</i>",
            styles["Normal"]
        )
        elements.append(note)

    doc.build(elements)
    return response


def faculty_view_courses_for_grades(request, faculty_id):
    faculty = get_object_or_404(Faculty, id=faculty_id)

    # ✅ Only count active pre-registration enrollments
    courses = (
        Course.objects.filter(faculties=faculty)
        .annotate(
            enrolled_count=Count(
                'enrollments',
                filter=Q(
                    enrollments__status__in=['ENR', 'CMP'],
                    enrollments__is_active_pre_reg=True  # ✅ Added
                ),
                distinct=True
            )
        )
        .filter(enrolled_count__gt=0)
        .order_by('code')
    )

    return render(request, 'instructor/view_results_courses.html', {
        'faculty': faculty,
        'courses': courses,
    })


def faculty_view_course_grades(request, faculty_id, course_code):
    faculty = get_object_or_404(Faculty, id=faculty_id)
    course = get_object_or_404(Course, code=course_code, faculties=faculty)
    policy = GradingPolicy.objects.filter(course=course).first()

    # Only current active enrollments (ENR or CMP so faculty can see both)
    enrollments = (
        StudentCourse.objects
        .filter(course=course, status__in=['ENR', 'CMP'], is_active_pre_reg=True)
        .select_related('student')
        .order_by('student__roll_no')
    )

    # Optional sort query param (?sort=asc/desc)
    sort_order = request.GET.get('sort', 'asc').lower()
    grade_order = {"A": 1, "A-": 2, "B": 3, "B-": 4, "C": 5, "C-": 6, "D": 7, "F": 8}
    reverse = sort_order == "desc"

    results = []
    for enr in enrollments:
        grade = enr.grade or ""
        outcome = enr.outcome or ""
        course_mode = (enr.course_mode or "").upper()

        if course_mode == "PF":
            display_grade = "Pass/Fail Mode"
            display_outcome = "Pass" if outcome and outcome.upper() in ["PAS", "PASS"] else "Fail"
        elif course_mode == "AUD":
            display_grade = "Audit"
            display_outcome = "Pass" if outcome and outcome.upper() in ["PAS", "PASS"] else "Fail"
        else:
            display_grade = grade or "—"
            display_outcome = "Pass" if outcome and outcome.upper() in ["PAS", "PASS"] else "Fail"

        results.append({
            "student": enr.student,
            "grade": display_grade,
            "outcome": display_outcome,
            "mode": course_mode,
        })

    # Sorting by grade order where possible (non-standard values go to end)
    def grade_sort_key(r):
        g = r["grade"]
        return grade_order.get(g, 999)

    results.sort(key=grade_sort_key, reverse=reverse)

    return render(request, "instructor/faculty_course_grades.html", {
        "faculty": faculty,
        "course": course,
        "results": results,
        "policy": policy,
        "sort_order": sort_order,
    })


@require_POST
@transaction.atomic
def update_student_grade(request, course_code):
    """
    AJAX endpoint to update either grade or outcome of a student in a course.
    Handles both Absolute/Relative and Pass/Fail (PF) modes.
    """
    student_id = request.POST.get("student_id")
    new_grade = request.POST.get("grade", "").strip()
    new_outcome = request.POST.get("outcome", "").strip()

    # --- Validation ---
    if not student_id:
        return JsonResponse({"ok": False, "error": "Missing student ID"}, status=400)
    if not new_grade and not new_outcome:
        return JsonResponse({"ok": False, "error": "Missing parameters"}, status=400)

    # --- Fetch student-course enrollment ---
    course = get_object_or_404(Course, code=course_code)
    enr = StudentCourse.objects.filter(student_id=student_id, course=course).first()
    if not enr:
        return JsonResponse({"ok": False, "error": "Enrollment not found"}, status=404)

    policy = GradingPolicy.objects.filter(course=course).first()
    course_mode = (enr.course_mode or "").upper()

    # --- Handle different update types ---
    if new_grade:  # ✅ Grade field updated
        enr.grade = new_grade
        # If the grading policy is Pass/Fail, then outcome should follow automatically
        if course_mode == "PF":
            enr.outcome = "PAS" if new_grade.lower().startswith("p") else "FAI"
        enr.save(update_fields=["grade", "outcome"])

        return JsonResponse({
            "ok": True,
            "grade": new_grade,
            "outcome": enr.outcome,
        })

    elif new_outcome:  # ✅ Outcome dropdown changed
        outcome_map = {
            "pass": "PAS",
            "fail": "FAI",
            "fs": "FS",  # Fail by short attendance
        }
        outcome_code = outcome_map.get(new_outcome.lower(), "FAI")
        enr.outcome = outcome_code

        # If the course mode is Pass/Fail, set grade to P/F automatically
        if course_mode == "PF":
            enr.grade = "P" if outcome_code == "PAS" else "F"

        enr.save(update_fields=["outcome", "grade"])
        return JsonResponse({
            "ok": True,
            "grade": enr.grade,
            "outcome": new_outcome.title(),
        })

    return JsonResponse({"ok": False, "error": "Invalid request"}, status=400)


def bulk_preregistration(request):
    context = {
        "columns": ["roll_no"] + [f"slot_{s}" for s in SLOTS],
        "preview_data": None,
    }

    if request.method == "POST":
        try:
            file = request.FILES.get("file")
            text_data = request.POST.get("csv_text", "").strip()
            enroll_mode = request.POST.get("enroll_mode", "PND")  # Default: PND
            parsed_data = []

            # --- STEP 1: Read CSV or Excel ---
            if file:
                if file.name.endswith(".csv"):
                    df = pd.read_csv(file)
                elif file.name.endswith((".xls", ".xlsx")):
                    df = pd.read_excel(file)
                else:
                    messages.error(request, "Please upload only CSV or Excel files.")
                    return redirect("bulk_preregistration")
            elif text_data:
                df = pd.read_csv(io.StringIO(text_data))
            else:
                messages.error(request, "Please provide a CSV/Excel file or paste data.")
                return redirect("bulk_preregistration")

            # --- STEP 2: Validate required columns ---
            expected_cols = ["roll_no"] + [f"slot_{s}" for s in SLOTS]
            missing_cols = [c for c in expected_cols if c not in df.columns]
            if missing_cols:
                messages.error(request, f"Missing columns: {', '.join(missing_cols)}")
                return redirect("bulk_preregistration")

            df = df.fillna("")  # Replace NaN with blank
            parsed_data = df.to_dict(orient="records")

            # --- STEP 3: If Preview Button Clicked ---
            if "preview" in request.POST:
                context["preview_data"] = parsed_data
                context["enroll_mode"] = enroll_mode
                messages.info(request, f"Preview loaded in {enroll_mode} mode. Review before submitting.")
                return render(request, "admin/prereg_bulk.html", context)

            # --- STEP 4: Process Data ---
            created, skipped = 0, 0
            current_year = datetime.now().year
            current_month = datetime.now().month
            semester_in_year = 1 if current_month >= 7 else 2

            def compute_sem(roll):
                try:
                    admission_year = 2000 + int(roll[1:3])
                    return (current_year - admission_year) * 2 + semester_in_year
                except:
                    return 1

            with transaction.atomic():
                for row in parsed_data:
                    roll_no = str(row.get("roll_no")).strip().lower()
                    student = Student.objects.filter(roll_no=roll_no).first()
                    if not student:
                        skipped += 1
                        continue

                    semester = student.calculate_current_semester()


                    for slot in SLOTS:
                        code = str(row.get(f"slot_{slot}", "")).strip().upper()
                        if not code:
                            continue

                        course = Course.objects.filter(code__iexact=code, slot=slot).first()
                        if not course:
                            skipped += 1
                            continue

                        obj, created_now = StudentCourse.objects.get_or_create(
                            student=student,
                            course=course,
                            semester=semester,
                            defaults={"status": enroll_mode, "course_mode": "REG"},
                        )
                        if created_now:
                            created += 1

            messages.success(
                request,
                f"✅ Bulk preregistration complete ({enroll_mode} mode): {created} added, {skipped} skipped."
            )
            return redirect("bulk_preregistration")

        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect("bulk_preregistration")

    return render(request, "admin/prereg_bulk.html", context)


# optional things

def parse_time_range(time_str):
    """
    Convert strings like '8–8:50 AM' or '2–5 PM' to (start_time, end_time) objects.
    """
    try:
        start_str, end_str = time_str.replace("–", "-").split("-")
        period = "AM" if "AM" in end_str else "PM"
        start_str = start_str.strip() + " " + period
        end_str = end_str.strip()
        if "AM" not in end_str and "PM" not in end_str:
            end_str += " " + period
        start_time = datetime.strptime(start_str.strip(), "%I:%M %p" if ":" in start_str else "%I %p").time()
        end_time = datetime.strptime(end_str.strip(), "%I:%M %p" if ":" in end_str else "%I %p").time()
        return start_time, end_time
    except Exception as e:
        print("❌ Time parsing error:", e, "for string:", time_str)
        return None, None


def student_timetable(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        return redirect("login")

    student = get_object_or_404(Student, roll_no=roll_no)

    # 🔹 Get student's enrolled courses
    enrolled_courses = (
        StudentCourse.objects
        .filter(student=student, status="ENR")
        .select_related("course")
    )

    enrolled_course_ids = [sc.course.id for sc in enrolled_courses if sc.course]
    if not enrolled_course_ids:
        return render(request, "timetable.html", {
            "timetable": {},
            "role": "student",
            "message": "No enrolled courses found.",
        })

    # 🔹 Auto-generate timetable entries if missing
    with transaction.atomic():
        for sc in enrolled_courses:
            course = sc.course
            slot = (course.slot or "").strip().upper()
            if not slot or slot not in SLOT_SCHEDULE:
                continue

            existing = Timetable.objects.filter(course=course)
            if not existing.exists():
                for day, time_range in SLOT_SCHEDULE[slot]:
                    start_time, end_time = parse_time_range(time_range)
                    if start_time and end_time:
                        Timetable.objects.create(
                            course=course,
                            faculty=course.faculties.first() if hasattr(course, "faculties") and course.faculties.exists() else None,
                            day=day,
                            start_time=start_time,
                            end_time=end_time,
                            location="TBD",
                            remarks="Auto-generated by slot system"
                        )

    # 🔹 Fetch timetable after generation
    timetable_data = (
        Timetable.objects
        .filter(course_id__in=enrolled_course_ids)
        .select_related("course", "faculty")
        .order_by("day", "start_time")
    )

    # 🔹 Organize by day for UI
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    timetable_by_day = {day: [] for day in days}

    for entry in timetable_data:
        timetable_by_day[entry.day].append({
            "course": f"{entry.course.code} – {entry.course.name}",
            "slot": entry.course.slot or "—",
            "faculty": f"{entry.faculty.first_name} {entry.faculty.last_name}" if entry.faculty else "N/A",
            "time": f"{entry.start_time.strftime('%I:%M %p')} – {entry.end_time.strftime('%I:%M %p')}",
            "location": entry.location or "—",
            "remarks": entry.remarks or "",
        })

    # 🔹 Remove empty days to avoid blank cards
    timetable_by_day = {k: v for k, v in timetable_by_day.items() if v}

    return render(request, "timetable.html", {
        "timetable": timetable_by_day,
        "role": "student",
    })


def faculty_course_list(request):
    email_id = request.session.get("email_id")
    if not email_id:
        return redirect("login")

    faculty = get_object_or_404(Faculty, email_id=email_id)
    courses = faculty.courses.all().distinct()

    return render(request, "instructor/course_list.html", {"faculty": faculty, "courses": courses})


def faculty_timetable_manage(request, course_id):
    email_id = request.session.get("email_id")
    if not email_id:
        return redirect("login")

    faculty = get_object_or_404(Faculty, email_id=email_id)
    course = get_object_or_404(Course, id=course_id, faculties=faculty)

    # Handle ADD / EDIT / DELETE
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            day = request.POST.get("day")
            start_time = request.POST.get("start_time")
            end_time = request.POST.get("end_time")
            location = request.POST.get("location")
            remarks = request.POST.get("remarks")

            Timetable.objects.create(
                course=course,
                faculty=faculty,
                day=day,
                start_time=start_time,
                end_time=end_time,
                location=location or None,
                remarks=remarks or None,
            )
            messages.success(request, f"✅ Added class on {day} at {start_time}.")
            return redirect("faculty_timetable_manage", course_id=course.id)

        elif action == "edit":
            entry_id = request.POST.get("entry_id")
            timetable_entry = get_object_or_404(Timetable, id=entry_id, course=course)

            timetable_entry.day = request.POST.get("day")
            timetable_entry.start_time = request.POST.get("start_time")
            timetable_entry.end_time = request.POST.get("end_time")
            timetable_entry.location = request.POST.get("location")
            timetable_entry.remarks = request.POST.get("remarks")
            timetable_entry.save()

            messages.success(request, f"✏️ Updated timetable entry successfully.")
            return redirect("faculty_timetable_manage", course_id=course.id)

        elif action == "delete":
            entry_id = request.POST.get("entry_id")
            Timetable.objects.filter(id=entry_id, course=course).delete()
            messages.success(request, "🗑️ Timetable entry deleted.")
            return redirect("faculty_timetable_manage", course_id=course.id)

    # GET — show timetable
    timetable = Timetable.objects.filter(course=course).order_by("day", "start_time")
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    return render(
        request,
        "instructor/timetable_manage.html",
        {"course": course,"faculty":faculty, "timetable": timetable, "days": days},
    )


def student_attendance(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        return redirect("login")

    student = get_object_or_404(Student, roll_no=roll_no)
    enrolled_courses = StudentCourse.objects.filter(student=student, status="ENR").select_related("course")
    attendance_data = []

    for enr in enrolled_courses:
        att = Attendance.objects.filter(student=student, course=enr.course).first()
        attendance_data.append({
            "course": enr.course.name,
            "code": enr.course.code,
            "attended": att.attended_classes if att else 0,
            "total": att.total_classes if att else 0,
            "percent": att.attendance_percent if att else 0,
        })

    return render(request, "attendance.html", {"attendance_data": attendance_data, "student": student})

def faculty_attendance(request, course_id):
    email_id = request.session.get("email_id")
    if not email_id:
        return redirect("login")

    faculty = get_object_or_404(Faculty, email_id=email_id)
    course = get_object_or_404(Course, id=course_id, faculties=faculty)
    enrollments = StudentCourse.objects.filter(course=course, status="ENR").select_related("student")

    # --- CSV Upload Handling ---
    if request.method == "POST" and "attendance_file" in request.FILES:
        file = request.FILES["attendance_file"]
        try:
            df = pd.read_csv(file)
        except Exception as e:
            messages.error(request, f"Failed to read CSV file: {e}")
            return redirect("faculty_attendance", course_id=course.id)

        with transaction.atomic():
            if "roll no" in [c.lower() for c in df.columns]:
                # ✅ Case 1: matrix-style
                df.columns = [c.strip().lower() for c in df.columns]
                roll_col = "roll no"
                date_cols = [c for c in df.columns if c != roll_col]
                roll_map = {e.student.roll_no.lower(): e.student for e in enrollments}

                for _, row in df.iterrows():
                    roll = str(row[roll_col]).strip().lower()
                    student = roll_map.get(roll)
                    if not student:
                        continue

                    total_classes = len(date_cols)
                    attended = sum(int(row[c]) for c in date_cols if str(row[c]).strip() in ["1", "P", "p"])
                    obj, _ = Attendance.objects.get_or_create(student=student, course=course)
                    obj.total_classes = total_classes
                    obj.attended_classes = attended
                    obj.save()

                messages.success(request, f"Attendance imported successfully for {len(df)} students.")

            elif "date" in [c.lower() for c in df.columns] and "roll nos present" in [c.lower() for c in df.columns]:
                # ✅ Case 2: date + roll list
                roll_map = {e.student.roll_no.lower(): e.student for e in enrollments}
                attendance_count = {s.student.roll_no.lower(): 0 for s in enrollments}
                total_dates = len(df)

                for _, row in df.iterrows():
                    present_rolls = [r.strip().lower() for r in str(row["Roll Nos Present"]).split(",")]
                    for roll in present_rolls:
                        if roll in attendance_count:
                            attendance_count[roll] += 1

                for roll, attended in attendance_count.items():
                    student = roll_map.get(roll)
                    if not student:
                        continue
                    obj, _ = Attendance.objects.get_or_create(student=student, course=course)
                    obj.total_classes = total_dates
                    obj.attended_classes = attended
                    obj.save()

                messages.success(request, f"Attendance uploaded for {total_dates} dates.")

            else:
                messages.error(request, "Unrecognized CSV format. Check column names.")
        return redirect("faculty_attendance", course_id=course.id)

    # --- Manual Form Submission ---
    elif request.method == "POST":
        for enr in enrollments:
            total = request.POST.get(f"total_{enr.student.id}", 0)
            attended = request.POST.get(f"attended_{enr.student.id}", 0)
            total = int(total) if total else 0
            attended = int(attended) if attended else 0
            obj, _ = Attendance.objects.get_or_create(student=enr.student, course=course)
            obj.total_classes = total
            obj.attended_classes = attended
            obj.save()
        messages.success(request, "Attendance updated successfully!")
        return redirect("faculty_attendance", course_id=course.id)

    # --- GET View ---
    data = []
    for enr in enrollments:
        att = Attendance.objects.filter(student=enr.student, course=course).first()
        data.append({
            "student": enr.student,
            "faculty": faculty,
            "attended": att.attended_classes if att else 0,
            "total": att.total_classes if att else 0,
            "percent": att.attendance_percent if att else 0,
        })

    return render(request, "instructor/attendance_manage.html", {"course": course, "faculty": faculty, "data": data})



def student_fees(request):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        return redirect("login")

    student = get_object_or_404(Student, roll_no=roll_no)
    pending_fees = FeeRecord.objects.filter(student=student, status="Pending").order_by("semester")
    paid_fees = FeeRecord.objects.filter(student=student, status="Paid").order_by("-semester")

    total_due = sum([(f.amount_due - f.amount_paid) for f in pending_fees if (f.amount_due - f.amount_paid) > 0])
    total_paid = sum([f.amount_paid for f in paid_fees])

    return render(request, "student_fees.html", {
        "student": student,
        "due_fees": pending_fees,
        "paid_fees": paid_fees,
        "total_due": total_due,
        "total_paid": total_paid,
    })


def mock_payment(request, fee_id):
    """Simulate payment gateway and mark as paid."""
    roll_no = request.session.get("roll_no")
    if not roll_no:
        return redirect("login")

    fee = get_object_or_404(FeeRecord, id=fee_id, student__roll_no=roll_no)

    # Simulate payment success
    fee.status = "Paid"
    fee.amount_paid = fee.amount_due
    fee.payment_time = timezone.now()
    fee.save()

    return JsonResponse({"message": "Payment successful!", "redirect": "/fees/receipt/" + str(fee.id)})


def download_fee_receipt(request, fee_id):
    roll_no = request.session.get("roll_no")
    if not roll_no:
        return redirect("login")

    student = get_object_or_404(Student, roll_no=roll_no)
    fee = get_object_or_404(FeeRecord, id=fee_id, student=student)

    # Prepare PDF
    response = HttpResponse(content_type="application/pdf")
    filename = f"Fee_Receipt_Sem{fee.semester}_{student.roll_no}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Register Hindi-compatible font
    font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansDevanagari-Regular.ttf")
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("NotoSans", font_path))
        font_name = "NotoSans"
    else:
        font_name = "Helvetica"

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Header
    p.setFont(font_name, 16)
    p.drawCentredString(width / 2, height - 80, "भारतीय प्रौद्योगिकी संस्थान मंडी")
    p.drawCentredString(width / 2, height - 100, "Indian Institute of Technology Mandi")
    p.setFont(font_name, 13)
    p.drawCentredString(width / 2, height - 120, "Kamand, Himachal Pradesh (175075)")
    p.setFont(font_name, 15)
    p.drawCentredString(width / 2, height - 150, "FEE PAYMENT RECEIPT")

    # Details
    y = height - 200
    p.setFont("Helvetica", 12)
    p.drawString(60, y, f"Name: {student.first_name} {student.last_name or ''}")
    y -= 18
    p.drawString(60, y, f"Roll No: {student.roll_no}")
    y -= 18
    p.drawString(60, y, f"Semester: {fee.semester}")
    y -= 18
    p.drawString(60, y, f"Amount Paid: ₹{fee.amount_paid}")
    y -= 18
    p.drawString(60, y, f"Payment Date: {fee.payment_time.strftime('%d %B %Y, %I:%M %p')}")

    y -= 30
    p.setFont("Helvetica-Oblique", 11)
    p.drawString(60, y, "This is a system-generated receipt. No signature required.")

    p.showPage()
    p.save()
    return response

#---------------------------





def fee_receipt_pdf(request, fee_id):
    """
    Generates a simple PDF receipt for a paid FeeRecord.
    Only student who owns the fee (or staff) can download.
    """
    roll_no = request.session.get("roll_no")
    fee = get_object_or_404(FeeRecord, id=fee_id)

    # Permission: either the student (session) or admin/staff (you can expand)
    if roll_no:
        if fee.student.roll_no != roll_no:
            return HttpResponseForbidden("Not authorized to download this receipt.")
    # else: you might allow staff to download by other checks (not implemented here)

    if fee.status != "Paid":
        return HttpResponse("Receipt only available for paid fees.", status=400)

    # Create PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    margin = 20 * mm
    x = margin
    y = height - margin

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, "Sootrank — Fee Receipt")
    c.setFont("Helvetica", 10)
    c.drawString(x, y - 18, f"Receipt No: {fee.receipt_number()}")
    c.drawString(x, y - 32, f"Date: {fee.payment_time.strftime('%Y-%m-%d %H:%M') if fee.payment_time else fee.created_at.strftime('%Y-%m-%d')}")

    y -= 60
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Student Details")
    c.setFont("Helvetica", 10)
    c.drawString(x, y - 16, f"Name: {fee.student.first_name} {fee.student.last_name or ''}")
    c.drawString(x, y - 32, f"Roll No: {fee.student.roll_no}")
    c.drawString(x, y - 48, f"Branch: {fee.student.branch.name if fee.student.branch else ''}")

    y -= 80
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Payment Details")
    c.setFont("Helvetica", 10)
    c.drawString(x, y - 16, f"Semester: {fee.semester}")
    c.drawString(x, y - 32, f"Amount Paid: ₹{fee.amount_paid}")
    c.drawString(x, y - 48, f"Payment ID: {fee.razorpay_payment_id or '—'}")
    c.drawString(x, y - 64, f"Order ID: {fee.razorpay_order_id or '—'}")

    y -= 100
    c.setFont("Helvetica", 9)
    c.drawString(x, y, "This is a computer-generated receipt from Sootrank. No signature required.")

    c.showPage()
    c.save()
    buffer.seek(0)

    filename = f"receipt_{fee.receipt_number()}.pdf"
    return HttpResponse(buffer, content_type="application/pdf", headers={
        "Content-Disposition": f'attachment; filename="{filename}"'
    })


def admin_fee_dashboard(request):
    """
    Main page — list of all students and their fee records.
    """
    students = Student.objects.all().order_by("roll_no").prefetch_related("fees")

    # Filter by semester if needed
    sem_filter = request.GET.get("semester")
    if sem_filter:
        students = students.filter(fees__semester=sem_filter).distinct()

    total_records = FeeRecord.objects.count()
    paid_count = FeeRecord.objects.filter(status="Paid").count()
    pending_count = FeeRecord.objects.filter(status="Pending").count()

    context = {
        "students": students,
        "total_records": total_records,
        "paid_count": paid_count,
        "pending_count": pending_count,
    }
    return render(request, "admin/fees_manage.html", context)


def admin_fee_add(request):
    """
    Add a new fee record manually.
    """
    if request.method == "POST":
        roll_no = request.POST.get("roll_no")
        semester = request.POST.get("semester")
        amount_due = request.POST.get("amount_due")

        student = Student.objects.filter(roll_no=roll_no).first()
        if not student:
            messages.error(request, f"Student with Roll No {roll_no} not found.")
            return redirect("admin_fee_add")

        FeeRecord.objects.create(
            student=student,
            semester=semester,
            amount_due=amount_due,
            status="Pending"
        )
        messages.success(request, f"Fee record added for {roll_no} (Sem {semester}).")
        return redirect("admin_fee_dashboard")

    return render(request, "admin/add_fee.html")


def admin_fee_update(request, fee_id):
    """
    Edit an existing fee record — mark paid, update amount, etc.
    """
    fee = get_object_or_404(FeeRecord, id=fee_id)
    if request.method == "POST":
        fee.amount_due = request.POST.get("amount_due")
        fee.amount_paid = request.POST.get("amount_paid")
        fee.status = request.POST.get("status")
        fee.semester = request.POST.get("semester")
        fee.save()
        messages.success(request, "Fee record updated successfully.")
        return redirect("admin_fee_dashboard")

    return render(request, "admin/update_fee.html", {"fee": fee})


def admin_fee_upload_csv(request):
    """
    Upload a CSV file with columns:
    roll_no, semester, amount_due, amount_paid, status
    """
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a valid CSV file.")
            return redirect("admin_fee_upload_csv")

        fs = FileSystemStorage()
        filename = fs.save(csv_file.name, csv_file)
        file_path = fs.path(filename)

        created, updated, failed = 0, 0, 0
        with open(file_path, mode="r", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            with transaction.atomic():
                for row in reader:
                    roll_no = row.get("roll_no")
                    semester = row.get("semester")
                    amount_due = row.get("amount_due")
                    amount_paid = row.get("amount_paid", "0")
                    status = row.get("status", "Pending")

                    if not roll_no or not semester or not amount_due:
                        failed += 1
                        continue

                    student = Student.objects.filter(roll_no=roll_no).first()
                    if not student:
                        failed += 1
                        continue

                    obj, created_now = FeeRecord.objects.update_or_create(
                        student=student,
                        semester=semester,
                        defaults={
                            "amount_due": amount_due,
                            "amount_paid": amount_paid or 0,
                            "status": status or "Pending",
                        },
                    )
                    if created_now:
                        created += 1
                    else:
                        updated += 1

        messages.success(request, f"CSV processed — Added: {created}, Updated: {updated}, Skipped: {failed}")
        return redirect("admin_fee_dashboard")

    return render(request, "admin/fee_upload_csv.html")

def admin_timetable_dashboard(request):
    """
    Centralized Timetable Management Dashboard:
    - Lists all courses
    - Sorts by courses with defined timetables first
    - Provides search + add/edit modal data
    """
    # Annotate each course with the number of timetable entries
    courses = (
        Course.objects.annotate(timetable_count=Count("timetables"))
        .prefetch_related("timetables", "faculties")
        .order_by("-timetable_count", "code")  # Courses with timetables first
    )

    days = Timetable.DAYS

    context = {
        "courses": courses,
        "days": days,
    }

    return render(request, "admin/admin_timetable.html", context)


def admin_timetable_add(request):
    """
    Handle adding a new timetable entry from admin dashboard
    """
    if request.method == "POST":
        course_id = request.POST.get("course")
        faculty_id = request.POST.get("faculty")
        day = request.POST.get("day")
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")
        location = request.POST.get("location")
        remarks = request.POST.get("remarks")

        if not course_id or not day or not start_time or not end_time:
            messages.error(request, "Please fill all required fields.")
            return redirect("admin_timetable_dashboard")

        try:
            course = Course.objects.get(id=course_id)
            faculty = Faculty.objects.get(id=faculty_id) if faculty_id else None

            # Conflict check
            exists = Timetable.objects.filter(
                course=course,
                day=day,
                start_time=start_time,
                end_time=end_time
            ).exists()
            if exists:
                messages.warning(request, "A timetable entry for this course and time already exists.")
                return redirect("admin_timetable_dashboard")

            Timetable.objects.create(
                course=course,
                faculty=faculty,
                day=day,
                start_time=start_time,
                end_time=end_time,
                location=location,
                remarks=remarks,
            )
            messages.success(request, "Timetable entry added successfully.")
        except Exception as e:
            messages.error(request, f"Error adding timetable: {e}")

    return redirect("admin_timetable_dashboard")


def admin_timetable_delete(request, timetable_id):
    """
    Delete a timetable entry by ID
    """
    try:
        timetable = Timetable.objects.get(id=timetable_id)
        timetable.delete()
        messages.success(request, "Timetable entry deleted successfully.")
    except Timetable.DoesNotExist:
        messages.error(request, "Timetable entry not found.")
    return redirect("admin_timetable_dashboard")


def admin_timetable_edit(request, timetable_id):
    """
    Edit an existing timetable entry (central admin control)
    """
    timetable = get_object_or_404(Timetable, id=timetable_id)

    if request.method == "POST":
        try:
            course_id = request.POST.get("course")
            faculty_id = request.POST.get("faculty")
            day = request.POST.get("day")
            start_time = request.POST.get("start_time")
            end_time = request.POST.get("end_time")
            location = request.POST.get("location")
            remarks = request.POST.get("remarks")

            if not course_id or not day or not start_time or not end_time:
                messages.error(request, "All required fields must be filled.")
                return redirect("admin_timetable_dashboard")

            # Conflict check (excluding current record)
            exists = Timetable.objects.filter(
                course_id=course_id,
                day=day,
                start_time=start_time,
                end_time=end_time
            ).exclude(id=timetable_id).exists()
            if exists:
                messages.warning(request, "A conflicting timetable entry already exists.")
                return redirect("admin_timetable_dashboard")

            timetable.course_id = course_id
            timetable.faculty_id = faculty_id if faculty_id else None
            timetable.day = day
            timetable.start_time = start_time
            timetable.end_time = end_time
            timetable.location = location
            timetable.remarks = remarks
            timetable.save()

            messages.success(request, "Timetable entry updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating timetable: {e}")

    return redirect("admin_timetable_dashboard")


def admin_attendance_dashboard(request):
    """
    Shows all courses with attendance summary
    Courses with enrolled students appear at the top
    """
    # Annotate courses safely (avoid division by zero)
    courses = (
        Course.objects.annotate(
            total_students=Count("attendance_records__student", distinct=True),
            avg_attendance=Avg(
                Case(
                    When(
                        attendance_records__total_classes__gt=0,
                        then=ExpressionWrapper(
                            (F("attendance_records__attended_classes") * 100.0) /
                            F("attendance_records__total_classes"),
                            output_field=FloatField(),
                        )
                    ),
                    default=Value(0.0),
                    output_field=FloatField(),
                )
            ),
        )
        .order_by("-total_students", "-avg_attendance", "code")
    )

    return render(request, "admin/admin_attendance_dashboard.html", {"courses": courses})


def admin_course_attendance(request, course_id):
    """
    Shows attendance for each student in the selected course
    """
    course = get_object_or_404(Course, id=course_id)
    enrollments = StudentCourse.objects.filter(course=course)
    student_ids = [e.student.id for e in enrollments]
    
    # Prefill attendance records if missing
    for sid in student_ids:
        Attendance.objects.get_or_create(student_id=sid, course=course)

    records = (
        Attendance.objects.filter(course=course)
        .select_related("student")
        .order_by("student__roll_no")
    )

    if request.method == "POST":
        for rec in records:
            total = request.POST.get(f"total_{rec.id}")
            attended = request.POST.get(f"attended_{rec.id}")
            if total is not None and attended is not None:
                rec.total_classes = int(total)
                rec.attended_classes = int(attended)
                rec.save()
        messages.success(request, "Attendance updated successfully.")
        return redirect("admin_course_attendance", course_id=course.id)

    return render(request, "admin/admin_course_attendance.html", {
        "course": course,
        "records": records
    })

def faculty_course_report(request):
    """Show all courses taught by the logged-in faculty with export options."""
    email_id = request.session.get("email_id")
    if not email_id:
        return redirect("login")

    faculty = get_object_or_404(Faculty, email_id=email_id)
    courses = Course.objects.filter(faculties=faculty).order_by("code")

    report_data = []
    for course in courses:
        count = StudentCourse.objects.filter(course=course, status="ENR").count()
        report_data.append({
            "course": course,
            "count": count,
        })

    return render(request, "instructor/faculty_report.html", {
        "faculty": faculty,
        "courses": report_data,
    })


def export_course_enrollments_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    enrollments = (
        StudentCourse.objects.filter(course=course, status="ENR")
        .select_related("student")
        .order_by("student__roll_no")
    )

    data = [
        {
            "Roll No": e.student.roll_no,
            "Name": f"{e.student.first_name} {e.student.last_name or ''}",
            "Email": e.student.email_id,
            "Branch": e.student.branch.name if e.student.branch else "",
            "Status": e.status,
        }
        for e in enrollments
    ]
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Enrollments")

    output.seek(0)
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{course.code}_enrollments.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



def export_course_enrollments_pdf(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    enrollments = (
        StudentCourse.objects.filter(course=course, status="ENR")
        .select_related("student")
        .order_by("student__roll_no")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{course.code}_enrollments.pdf"'
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    width, height = A4
    y = height - 50
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y, f"Course Enrollment Report: {course.code} — {course.name}")
    y -= 30
    pdf.setFont("Helvetica", 11)
    pdf.drawString(50, y, f"Total Enrolled Students: {enrollments.count()}")
    y -= 40
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, "Roll No")
    pdf.drawString(150, y, "Name")
    pdf.drawString(350, y, "Branch")
    pdf.setFont("Helvetica", 10)
    y -= 20

    for e in enrollments:
        if y < 60:
            pdf.showPage()
            y = height - 50
        pdf.drawString(50, y, e.student.roll_no)
        pdf.drawString(150, y, f"{e.student.first_name} {e.student.last_name or ''}")
        pdf.drawString(350, y, e.student.branch.name if e.student.branch else "")
        y -= 20

    pdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()
    response.write(pdf_data)
    return response


